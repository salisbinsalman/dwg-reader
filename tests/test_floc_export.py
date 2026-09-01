#!/usr/bin/env python3
"""Unit tests for FLOC path/description + SAP export (no Bedrock)."""

from __future__ import annotations

import csv
import json
import re
import tempfile
import unittest
from pathlib import Path

from dwg_reader.dwg_floc_context import (
    abbrev_data,
    abbreviate_pltxt,
    build_tplnr,
    description_from_nearby,
    floc_paths_for_function,
    load_floc_context_for_input,
    load_sml_abbreviations,
    merge_floc_context,
    normalize_pltxt,
)
from dwg_reader.dwg_pid_hierarchy_ai import rows_from_ai, title_context
from dwg_reader.export_sap_floc import (
    SAP_COLUMNS,
    _clean_line_description,
    _strip_pipe_class_codes,
    _strip_trailing_spec,
    build_floc_rows,
    collect_functions,
    evaluate_against_gt,
    load_function_positions,
    write_floc_workbook,
)


ROOT = Path(__file__).resolve().parents[1]
ABBREV_JSON = ROOT / "standards/sml_abbreviations.json"
HIERARCHY_CSV = ROOT / "outputs/Broke System.hierarchy_orchestrator.csv"


class FlocContextTests(unittest.TestCase):
    def test_tplnr_function_path(self) -> None:
        paths = floc_paths_for_function("35-24L009")
        self.assertEqual(paths["plant"], "5001")
        self.assertEqual(paths["line"], "5001-PM03")
        self.assertEqual(paths["process"], "5001-PM03-BR")
        self.assertEqual(paths["subprocess"], "5001-PM03-BR-BR1")
        self.assertEqual(paths["function"], "5001-PM03-BR-BR1-35-24L009")
        self.assertLessEqual(len(paths["function"]), 30)

    def test_build_tplnr_truncates(self) -> None:
        s = build_tplnr("5001", "PM03", "BR", "BR1", "35-24L009EXTRA")
        self.assertEqual(len(s), 30)

    def test_normalize_pltxt(self) -> None:
        d = normalize_pltxt("35-24L009 Broke Roll Pulper")
        self.assertTrue(d.startswith("35-24L009"))
        self.assertIn("PLPR", d)
        self.assertLessEqual(len(d), 40)

    def test_description_fallback(self) -> None:
        d = description_from_nearby("35-24P519", ["BROKE ROLL PULPER PUMP", "35-24P519"])
        self.assertTrue(d.startswith("35-24P519"))
        self.assertIn("PMP", d)
        self.assertLessEqual(len(d), 40)

    def test_plant_copies_to_maintenance_and_planning(self) -> None:
        ctx = merge_floc_context({
            "plant": "6001",
            "line_code": "TM01",
            "process_code": "WU",
            "sub_process": "WUC",
        })
        self.assertEqual(ctx["plant"], "6001")
        self.assertEqual(ctx["maintenance_plant"], "6001")
        self.assertEqual(ctx["planning_plant"], "6001")

    def test_explicit_maintenance_plant_not_overwritten(self) -> None:
        ctx = merge_floc_context({"plant": "6001", "maintenance_plant": "5001"})
        self.assertEqual(ctx["maintenance_plant"], "5001")
        self.assertEqual(ctx["planning_plant"], "6001")

    def test_broke_defaults_stay_5001(self) -> None:
        ctx = merge_floc_context(None)
        self.assertEqual(ctx["plant"], "5001")
        self.assertEqual(ctx["maintenance_plant"], "5001")
        self.assertEqual(ctx["planning_plant"], "5001")

    def test_wu12_map_entry_sets_swerk_6001(self) -> None:
        ctx = load_floc_context_for_input(
            Path("inputs/GORB18779.05_SH12(12)_Code 14 - P&ID Ventil Unit WU12_SWE Shotton_CE.dwg")
        )
        self.assertEqual(ctx["plant"], "6001")
        self.assertEqual(ctx["maintenance_plant"], "6001")
        self.assertEqual(ctx["planning_plant"], "6001")


class SmlAbbreviationJsonTests(unittest.TestCase):
    def test_json_exists_and_has_conveyor(self) -> None:
        self.assertTrue(ABBREV_JSON.exists())
        data = abbrev_data(ABBREV_JSON)
        self.assertIn("abbreviations", data)
        self.assertEqual(data["abbreviations"]["CONVEYOR"], "CVYR")
        self.assertGreaterEqual(len(data["abbreviations"]), 300)

    def test_load_sml_abbreviations_from_json(self) -> None:
        mapping = load_sml_abbreviations(ABBREV_JSON)
        self.assertEqual(mapping["PUMP"], "PMP")
        self.assertEqual(mapping["CONVEYOR"], "CVYR")


class AbbreviatePltxtTests(unittest.TestCase):
    def test_review_case_conveyor(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L004 BROKE CONVEYOR 2"),
            "35-24L004 BROKE CVYR 2",
        )

    def test_conveyor_plural(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L006 BROKE CONVEYORS 3"),
            "35-24L006 BROKE CVYR 3",
        )

    def test_broke_stays_full(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L004 BROKE CONVEYOR 2"),
            "35-24L004 BROKE CVYR 2",
        )
        self.assertIn("BROKE", abbreviate_pltxt("35-24L009 BROKE ROLL PULPER"))

    def test_domain_phrase_broke_roll_pulper(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L009 BROKE ROLL PULPER"),
            "35-24L009 BROKE ROLL PLPR",
        )

    def test_domain_phrase_size_press_pulper(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L003 SIZE PRESS PULPER"),
            "35-24L003 SIZE PRESS PLPR",
        )

    def test_pump_abbreviated_after_phrase(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24P519 BROKE ROLL PLPR PUMP"),
            "35-24P519 BROKE ROLL PLPR PMP",
        )

    def test_sml_phrase_hand_valve(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24HV-548 HAND VALVE"),
            "35-24HV-548 HV",
        )

    def test_sml_phrase_normally_closed(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("DRAIN VALVE NORMALLY CLOSED"),
            "DRN VLV NC",
        )

    def test_preserves_tag_and_number(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L006 BROKE CONVEYOR 3"),
            "35-24L006 BROKE CVYR 3",
        )

    def test_already_abbreviated_tokens_unchanged(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24P519 BROKE ROLL PLPR PMP"),
            "35-24P519 BROKE ROLL PLPR PMP",
        )

    def test_motor_tank_valve_shower(self) -> None:
        self.assertEqual(abbreviate_pltxt("PUMP MOTOR"), "PMP MTR")
        self.assertEqual(abbreviate_pltxt("STORAGE TANK"), "STOR TNK")
        self.assertEqual(abbreviate_pltxt("CONTROL VALVE"), "CTRL VLV")
        self.assertEqual(abbreviate_pltxt("SHOWER S1"), "SHW S1")

    def test_overflow_uses_sml_standard(self) -> None:
        self.assertEqual(abbreviate_pltxt("LINE OVERFLOW"), "LN OVRFL")

    def test_max_length_40(self) -> None:
        long_text = "35-24L001 BROKE ROLL PULPER PUMP MOTOR AGITATOR VALVE SHOWER"
        self.assertLessEqual(len(abbreviate_pltxt(long_text)), 40)

    def test_emergency_stop_phrase(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24ES-508 WINDER AREA EMERGENCY STOP"),
            "35-24ES-508 WINDER AREA E-STOP",
        )

    def test_idempotent_on_normalized_text(self) -> None:
        once = abbreviate_pltxt("35-24L004 BROKE CONVEYOR 2")
        twice = abbreviate_pltxt(once)
        self.assertEqual(once, twice)


class RowsFromAiTests(unittest.TestCase):
    def test_deterministic_mask_and_description(self) -> None:
        ctx = title_context({"title_block": []}, "Broke System")
        parsed = {
            "sub_process": "BR1",
            "function": "35-24L009",
            "description": "35-24L009 BROKE ROLL PLPR",
            "rows": [
                {"equipment": "35-24-189", "subequipment": "", "description": "LN OVFL"},
                {"equipment": "", "subequipment": "35-24-194", "description": "HV DRN"},
            ],
        }
        rows, order = rows_from_ai(1, ctx, "35-24L009", parsed, parent={"nearby_descriptions": ["BROKE ROLL PULPER"]})
        self.assertGreater(order, 1)
        masks = [r["MASK"] for r in rows if r["MASK"]]
        self.assertIn("5001-PM03-BR", masks)
        self.assertIn("5001-PM03-BR-BR1", masks)
        self.assertIn("5001-PM03-BR-BR1-35-24L009", masks)
        fn_row = next(r for r in rows if r["FUNCTION"] == "35-24L009")
        self.assertEqual(fn_row["MASK"], "5001-PM03-BR-BR1-35-24L009")
        self.assertTrue(fn_row["DESCRIPTION"].startswith("35-24L009"))
        self.assertLessEqual(len(fn_row["DESCRIPTION"]), 40)
        child = next(r for r in rows if r["EQUIPMENT"] == "35-24-189")
        self.assertEqual(child["MASK"], "")


class StripPipeClassCodesTests(unittest.TestCase):
    """Issue 4 — pipe fluid-class codes and nominal-diameter specs stripped from descriptions."""

    def test_strips_waa_and_dn300(self) -> None:
        self.assertEqual(_strip_pipe_class_codes("WAA LN DN300"), "LN")

    def test_strips_waa_and_dn400(self) -> None:
        self.assertEqual(_strip_pipe_class_codes("WAA LN DN400"), "LN")

    def test_strips_wfl_dn15(self) -> None:
        self.assertEqual(_strip_pipe_class_codes("WFL DN15"), "")

    def test_strips_pp_spec_mid_string(self) -> None:
        self.assertEqual(_strip_pipe_class_codes("PP-200 PLPR DIS LN"), "PLPR DIS LN")

    def test_strips_pp_250_variant(self) -> None:
        self.assertEqual(_strip_pipe_class_codes("PP-250 PROC LN"), "PROC LN")

    def test_leaves_functional_description_unchanged(self) -> None:
        self.assertEqual(_strip_pipe_class_codes("BROKE THICKENER"), "BROKE THICKENER")
        self.assertEqual(_strip_pipe_class_codes("PRESS PLPR DISCHARGE LN"), "PRESS PLPR DISCHARGE LN")

    def test_leaves_conveyor_number_unchanged(self) -> None:
        # Plain integer must not be stripped
        self.assertEqual(_strip_pipe_class_codes("BROKE CVYR 2"), "BROKE CVYR 2")

    def test_collect_functions_strips_pipe_class_codes_in_description(self) -> None:
        """collect_functions must return cleaned descriptions (Issue 4)."""
        rows = [
            {"FUNCTION": "35-24-149", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
             "DESCRIPTION": "35-24-149 WAA LN DN300"},
            {"FUNCTION": "35-24-150", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
             "DESCRIPTION": "35-24-150 WAA LN DN400"},
            {"FUNCTION": "35-24L002", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
             "DESCRIPTION": "35-24L002 BROKE THICKENER"},
        ]
        fns = collect_functions(rows, filter_utility_lines=False)
        by_tag = {fn: desc for fn, _, desc in fns}
        self.assertNotIn("WAA", by_tag["35-24-149"], "WAA fluid-class code must be stripped")
        self.assertNotIn("DN300", by_tag["35-24-149"], "DN300 size code must be stripped")
        self.assertNotIn("WAA", by_tag["35-24-150"])
        self.assertNotIn("DN400", by_tag["35-24-150"])
        self.assertIn("THICKENER", by_tag["35-24L002"], "functional description preserved")


class UtilityLineAndCrossUnitFilterTests(unittest.TestCase):
    """Issues 2 and 6 — WFL utility lines and cross-unit pipelines excluded from FLOC."""

    def _make_row(self, fn: str, desc: str = "") -> Dict:
        return {"FUNCTION": fn, "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
                "DESCRIPTION": desc or fn}

    def test_wfl_lines_excluded_by_default(self) -> None:
        """Issue 2: WFL flush lines must not appear as top-level FLOC function nodes."""
        rows = [
            self._make_row("35-24L001", "35-24L001 PRESS PLPR"),
            self._make_row("35-24-008", "35-24-008 WFL DN15"),
            self._make_row("35-24-010", "35-24-010 WFL DN15"),
            self._make_row("35-24T601", "35-24T601 COUCH PIT TNK"),
        ]
        fns = collect_functions(rows)
        tags = [fn for fn, _, _ in fns]
        self.assertIn("35-24L001", tags)
        self.assertIn("35-24T601", tags)
        self.assertNotIn("35-24-008", tags, "WFL flush line must be filtered out")
        self.assertNotIn("35-24-010", tags, "WFL flush line must be filtered out")

    def test_filter_disabled_keeps_wfl_lines(self) -> None:
        rows = [
            self._make_row("35-24-008", "35-24-008 WFL DN15"),
        ]
        fns = collect_functions(rows, filter_utility_lines=False)
        self.assertEqual(len(fns), 1)
        self.assertEqual(fns[0][0], "35-24-008")

    def test_non_wfl_process_pipelines_kept(self) -> None:
        """Genuine process pipelines (no WFL keyword) must survive filtering."""
        rows = [
            self._make_row("35-24L001", "35-24L001 PRESS PLPR"),
            self._make_row("35-24-056", "35-24-056 BROKE DIV HDR"),
            self._make_row("35-24-215", "35-24-215 THKND BROKE FD LN"),
        ]
        fns = collect_functions(rows)
        tags = [fn for fn, _, _ in fns]
        self.assertIn("35-24-056", tags, "process pipeline without WFL must be kept")
        self.assertIn("35-24-215", tags, "process pipeline without WFL must be kept")

    def test_cross_unit_pipelines_excluded(self) -> None:
        """Issue 6: numeric pipeline tags from a different area-unit are filtered out."""
        rows = [
            self._make_row("35-24L001", "35-24L001 PRESS PLPR"),
            self._make_row("35-24T601", "35-24T601 COUCH PIT TNK"),
            self._make_row("35-24-056", "35-24-056 BROKE DIV HDR"),    # unit-24, keep
            self._make_row("35-25-034", "35-25-034 WAF LN DN600"),     # unit-25, cross-unit
            self._make_row("35-25-072", "35-25-072 WAF LN DN350"),     # unit-25, cross-unit
        ]
        fns = collect_functions(rows)
        tags = [fn for fn, _, _ in fns]
        self.assertIn("35-24L001", tags)
        self.assertIn("35-24T601", tags)
        self.assertIn("35-24-056", tags, "same-unit pipeline must remain")
        self.assertNotIn("35-25-034", tags, "cross-unit pipeline 35-25-034 must be filtered")
        self.assertNotIn("35-25-072", tags, "cross-unit pipeline 35-25-072 must be filtered")

    def test_gor_tags_not_cross_unit_filtered(self) -> None:
        """GOR tags (no NN-NN prefix) must never be filtered by cross-unit logic."""
        rows = [
            self._make_row("WU12", "WU12 VENTIL UNIT"),
            self._make_row("WU13", "WU13 VENTIL UNIT"),
        ]
        fns = collect_functions(rows)
        tags = [fn for fn, _, _ in fns]
        self.assertIn("WU12", tags, "GOR tag must not be filtered as cross-unit")
        self.assertIn("WU13", tags, "GOR tag must not be filtered as cross-unit")


class CleanLineDescriptionTests(unittest.TestCase):
    """Issue 1 — flow substance codes translated; bare pipe sizes stripped."""

    def test_translates_waf_to_white_wtr(self) -> None:
        result = _clean_line_description("35-24-016 WAF 250 DIST HDR")
        self.assertNotIn("WAF", result)
        self.assertIn("WHITE WTR", result)
        self.assertIn("DIST HDR", result)
        self.assertNotIn("250", result)

    def test_translates_waa_to_cloudy_filt(self) -> None:
        result = _clean_line_description("35-24-149 WAA LN DN300")
        self.assertNotIn("WAA", result)
        self.assertIn("CLOUDY FILT", result)
        self.assertNotIn("DN300", result)

    def test_translates_wfl_to_seal_wtr(self) -> None:
        result = _clean_line_description("35-24-008 WFL LN DN15")
        self.assertNotIn("WFL", result)
        self.assertIn("SEAL WTR", result)
        self.assertNotIn("DN15", result)

    def test_strips_bare_dn_size_after_code(self) -> None:
        result = _clean_line_description("35-24-017 WAF 300 PROC LN")
        self.assertNotIn("300", result)
        self.assertIn("WHITE WTR", result)
        self.assertIn("PROC", result)

    def test_strips_dn_prefixed_size(self) -> None:
        result = _clean_line_description("35-24-054 WAF LN DN500")
        self.assertNotIn("DN500", result)
        self.assertNotIn("WAF", result)
        self.assertIn("WHITE WTR", result)

    def test_leaves_conveyor_number_unchanged(self) -> None:
        self.assertEqual(_clean_line_description("BROKE CVYR 2"), "BROKE CVYR 2")

    def test_leaves_functional_description_unchanged(self) -> None:
        self.assertEqual(_clean_line_description("BROKE THICKENER"), "BROKE THICKENER")
        self.assertEqual(_clean_line_description("PRESS PLPR DISCHARGE LN"), "PRESS PLPR DISCHARGE LN")

    def test_strips_pp_spec(self) -> None:
        result = _clean_line_description("35-24-095 PP-200 PLPR DIS LN")
        self.assertNotIn("PP-200", result)
        self.assertIn("PLPR DIS LN", result)

    def test_strips_dangling_arrow_and_sizes(self) -> None:
        result = _clean_line_description("35-24-095 PRESS PLPR > PP-200 DN150")
        self.assertNotIn(">", result)
        self.assertNotIn("PP-200", result)
        self.assertNotIn("DN", result)
        self.assertIn("PRESS", result)

    def test_strips_bare_mm_size(self) -> None:
        """B-05: leftover 15MM (no DN prefix) must be stripped."""
        result = _clean_line_description("35-24-005 CLG WTR DIST SPUR 15MM")
        self.assertNotIn("15MM", result)
        self.assertIn("CLG WTR", result)

    def test_translates_wfc_to_clg_wtr(self) -> None:
        result = _clean_line_description("35-24-200 WFC LN DN80")
        self.assertIn("CLG WTR", result)
        self.assertNotIn("WFC", result)

    def test_collect_functions_uses_translated_substance(self) -> None:
        """collect_functions must produce translated substance labels in descriptions."""
        rows = [
            {"FUNCTION": "35-24-016", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
             "DESCRIPTION": "35-24-016 WAF 250 DIST HDR"},
            {"FUNCTION": "35-24-054", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
             "DESCRIPTION": "35-24-054 WAF LN DN500"},
            {"FUNCTION": "35-24L002", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
             "DESCRIPTION": "35-24L002 BROKE THICKENER"},
        ]
        fns = collect_functions(rows, filter_utility_lines=False)
        by_tag = {fn: desc for fn, _, desc in fns}
        self.assertNotIn("WAF", by_tag["35-24-016"])
        self.assertIn("WHITE WTR", by_tag["35-24-016"])
        self.assertNotIn("250", by_tag["35-24-016"])
        self.assertNotIn("WAF", by_tag["35-24-054"])
        self.assertIn("WHITE WTR", by_tag["35-24-054"])
        self.assertIn("THICKENER", by_tag["35-24L002"])

    def test_custom_flow_codes_override_defaults(self) -> None:
        """flow_codes parameter overrides the module-level default."""
        custom = {"FOO": "FOO WTR", "BAR": "BAR STM"}
        result = _clean_line_description("FOO DIST HDR", flow_codes=custom)
        self.assertIn("FOO WTR", result)
        # Default code WAF must NOT translate when custom codes are provided and don't have it
        result2 = _clean_line_description("WAF DIST HDR", flow_codes=custom)
        self.assertNotIn("WHITE WTR", result2)  # WAF not in custom, stays as WAF
        self.assertIn("WAF", result2)

    def test_empty_flow_codes_skips_translation(self) -> None:
        """Empty codes dict disables substance translation entirely."""
        result = _clean_line_description("WAF DIST HDR", flow_codes={})
        self.assertIn("WAF", result)  # no translation, WAF stays


class StripTrailingSpecTests(unittest.TestCase):
    def test_strips_hyphenated_model(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L008 WINDER PLPR HP-33G2"), "35-24L008 WINDER PLPR")
        self.assertEqual(_strip_trailing_spec("35-24L002 BROKE THICKENER GT-2540"), "35-24L002 BROKE THICKENER")
        self.assertEqual(_strip_trailing_spec("35-24L010 BROKE SCREEN PROFS-250HC"), "35-24L010 BROKE SCREEN")

    def test_strips_mixed_alphanumeric(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L009 BROKE ROLL PLPR 141BDTPD"), "35-24L009 BROKE ROLL PLPR")

    def test_strips_stacked_vendor_codes(self) -> None:
        self.assertEqual(
            _strip_trailing_spec("35-24L008 WINDER PLPR HP-33G2 800BDTPD"),
            "35-24L008 WINDER PLPR",
        )
        self.assertEqual(
            _strip_trailing_spec("35-24L010 BROKE SCREEN PROFS-250HC"),
            "35-24L010 BROKE SCREEN",
        )

    def test_strips_m3_volume_spec(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24T603 BROKE COLLECTION TNK 40M3"), "35-24T603 BROKE COLLECTION TNK")
        self.assertEqual(_strip_trailing_spec("35-24T602 BROKE STOR TWR 4000M3"), "35-24T602 BROKE STOR TWR")

    def test_keeps_plain_integer(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L006 BROKE CONVEYOR 3"), "35-24L006 BROKE CONVEYOR 3")
        self.assertEqual(_strip_trailing_spec("35-24L004 BROKE CONVEYOR 2"), "35-24L004 BROKE CONVEYOR 2")

    def test_keeps_clean_description(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L003 SIZE PRESS PLPR"), "35-24L003 SIZE PRESS PLPR")

    def test_strips_semicolon_separated_specs(self) -> None:
        self.assertEqual(
            _strip_trailing_spec("35-24L007 SLABBING PLPR; HP-33G2; BDTPD"),
            "35-24L007 SLABBING PLPR",
        )
        self.assertEqual(
            _strip_trailing_spec("35-24L007 SLABBING PLPR; HP-33G2; 800 BDTPD"),
            "35-24L007 SLABBING PLPR",
        )
        self.assertEqual(
            _strip_trailing_spec("35-24L005 REEL PLPR; HP-63G2; 2621 ADTPD"),
            "35-24L005 REEL PLPR",
        )

    def test_semicolon_keeps_non_spec_tail(self) -> None:
        self.assertEqual(
            _strip_trailing_spec("35-24L007 SLABBING PLPR; DUTY"),
            "35-24L007 SLABBING PLPR DUTY",
        )


class ExportFlocTests(unittest.TestCase):
    def test_build_and_write_workbook(self) -> None:
        functions = [
            ("35-24L009", "5001-PM03-BR-BR1-35-24L009", "35-24L009 BROKE ROLL PLPR"),
            ("35-24P519", "5001-PM03-BR-BR1-35-24P519", "35-24P519 BROKE ROLL PLPR PMP"),
        ]
        floc_rows = build_floc_rows(functions)
        self.assertEqual(len(floc_rows), 6)
        self.assertEqual(floc_rows[0]["TPLNR"], "5001")
        self.assertEqual(floc_rows[3]["TPLNR"], "5001-PM03-BR-BR1")
        self.assertEqual(floc_rows[4]["TPLMA"], "5001-PM03-BR-BR1")
        self.assertEqual(floc_rows[4]["TPLNR"], "5001-PM03-BR-BR1-35-24L009")
        self.assertEqual(floc_rows[4]["FLTYP"], "M")
        self.assertEqual(floc_rows[4]["SWERK"], "5001")
        self.assertEqual(floc_rows[4]["ABCKZ"], "D")
        self.assertEqual(floc_rows[4]["INGRP"], "P01")
        self.assertLessEqual(len(floc_rows[4]["PLTXT"]), 40)

        template = ROOT / "docs/examples/final-output-template.xlsx"
        self.assertTrue(template.exists())
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_floc_workbook(template, out, floc_rows)
            self.assertTrue(out.exists())
            from openpyxl import load_workbook

            wb = load_workbook(out)
            ws = wb["Functional Location"]
            self.assertEqual(ws.cell(8, 2).value, "5001")
            self.assertEqual(ws.cell(12, 2).value, "5001-PM03-BR-BR1-35-24L009")
            fltyp_col = 2 + SAP_COLUMNS.index("FLTYP")
            eqart_col = 2 + SAP_COLUMNS.index("EQART")
            self.assertEqual(ws.cell(12, fltyp_col).number_format, "@")
            self.assertEqual(ws.cell(12, eqart_col).number_format, "@")
            self.assertIsInstance(ws.cell(12, fltyp_col).value, str)

    def test_line_function_pltxt_gets_ln_prefix(self) -> None:
        """Pipe-line FUNCTION tags must get LN … PLTXT (same rule as Equipment)."""
        functions = [
            ("35-24-194", "5001-PM03-BR-BR1-35-24-194", "35-24-194 SEALING WATER TO HOSE"),
            ("35-24-095", "5001-PM03-BR-BR1-35-24-095", "35-24-095 PP-200 PLPR LN"),
            ("168L-522", "5001-TM01-WU-WUC-168L-522", "168L-522 PIPE DN65"),
        ]
        floc_rows = build_floc_rows(functions)
        line_rows = [r for r in floc_rows if r["TPLNR"].endswith(("35-24-194", "35-24-095", "168L-522"))]
        self.assertEqual(len(line_rows), 3)
        for r in line_rows:
            self.assertTrue(
                r["PLTXT"].startswith("LN "),
                msg=f"{r['TPLNR']} PLTXT={r['PLTXT']!r}",
            )
            self.assertNotIn("DN", r["PLTXT"])
            self.assertNotIn("PP-200", r["PLTXT"])
        self.assertTrue(any(r["PLTXT"].startswith("LN 35-24-194") for r in line_rows))
        self.assertTrue(any(r["PLTXT"].startswith("LN 35-24-095") for r in line_rows))
        self.assertTrue(any(r["PLTXT"].startswith("LN 168L-522") for r in line_rows))
        wu = next(r for r in line_rows if r["TPLNR"].endswith("168L-522"))
        self.assertEqual(wu["PLTXT"], "LN 168L-522 PIPE")

    def test_line_function_pltxt_strips_duplicate_tag(self) -> None:
        functions = [
            ("35-24-185", "5001-PM03-BR-BR1-35-24-185", "LN 35-24-185 35-24-185 WW > BROKE ROLL P"),
        ]
        floc_rows = build_floc_rows(functions)
        row = next(r for r in floc_rows if r["TPLNR"].endswith("35-24-185"))
        self.assertEqual(row["PLTXT"], "LN 35-24-185 WW > BROKE ROLL P")
        self.assertEqual(row["PLTXT"].count("35-24-185"), 1)

    def test_tissue_floc_swerk_is_6001(self) -> None:
        functions = [("WU12", "6001-TM01-WU-WUC-WU12", "WU12 VENTIL UNIT")]
        ctx = merge_floc_context({
            "plant": "6001",
            "line_code": "TM01",
            "process_code": "WU",
            "sub_process": "WUC",
            "site_name": "SHOTTON MILL LTD",
            "line_name": "TISSUE MACHINE 1",
            "process_name": "VENTILATION",
        })
        floc_rows = build_floc_rows(functions, ctx=ctx)
        self.assertTrue(all(r["SWERK"] == "6001" for r in floc_rows))
        fn = next(r for r in floc_rows if r["TPLNR"].endswith("WU12"))
        self.assertEqual(fn["IWERK"], "6001")
        self.assertEqual(fn["TPLNR"], "6001-TM01-WU-WUC-WU12")

    def test_winder_floc_strips_vendor_model(self) -> None:
        functions = [
            ("35-24L008", "5001-PM03-BR-BR1-35-24L008", "35-24L008 WINDER PLPR HP-33G2 800BDTPD"),
        ]
        floc_rows = build_floc_rows(functions)
        row = next(r for r in floc_rows if r["TPLNR"].endswith("35-24L008"))
        self.assertEqual(row["PLTXT"], "35-24L008 WINDER PLPR")
        self.assertNotIn("HP-33G2", row["PLTXT"])

    def test_slabbing_floc_strips_semicolon_specs(self) -> None:
        functions = [
            ("35-24L007", "5001-PM03-BR-BR1-35-24L007", "35-24L007 SLABBING PLPR; HP-33G2; BDTPD"),
        ]
        floc_rows = build_floc_rows(functions)
        row = next(r for r in floc_rows if r["TPLNR"].endswith("35-24L007"))
        self.assertEqual(row["PLTXT"], "35-24L007 SLABBING PLPR")
        self.assertNotIn("HP-33G2", row["PLTXT"])
        self.assertNotIn("BDTPD", row["PLTXT"])
        self.assertNotIn(";", row["PLTXT"])

    def test_non_line_function_pltxt_no_ln_prefix(self) -> None:
        functions = [
            ("35-24L009", "5001-PM03-BR-BR1-35-24L009", "35-24L009 BROKE ROLL PLPR"),
        ]
        floc_rows = build_floc_rows(functions)
        fn = next(r for r in floc_rows if r["TPLNR"].endswith("35-24L009"))
        self.assertFalse(fn["PLTXT"].startswith("LN "))
        self.assertTrue(fn["PLTXT"].startswith("35-24L009"))

    def test_line_function_floc_gets_mech_workcenter(self) -> None:
        """M2-5: pipeline FUNCTION FLOCs must have GEWRK=MECH and EQART=2100."""
        functions = [
            ("35-24-016", "5001-PM03-BR-BR1-35-24-016", "35-24-016 WHITE WTR DIST HDR"),
            ("35-24-054", "5001-PM03-BR-BR1-35-24-054", "35-24-054 WHITE WTR"),
        ]
        floc_rows = build_floc_rows(functions)
        line_rows = [r for r in floc_rows if r["TPLNR"].endswith("35-24-016") or r["TPLNR"].endswith("35-24-054")]
        for r in line_rows:
            self.assertEqual(r["GEWRK"], "MECH", msg=f"{r['TPLNR']} GEWRK={r['GEWRK']!r}")
            self.assertEqual(r["EQART"], "2100", msg=f"{r['TPLNR']} EQART={r['EQART']!r}")

    def test_function_rows_populate_gewrk(self) -> None:
        functions = [
            ("35-24L001", "5001-PM03-BR-BR1-35-24L001", "35-24L001 PRESS PLPR"),
            ("35-24L004", "5001-PM03-BR-BR1-35-24L004", "35-24L004 BROKE CVYR 2"),
            ("35-24P519", "5001-PM03-BR-BR1-35-24P519", "35-24P519 BROKE ROLL PLPR PMP"),
            ("35-24T602", "5001-PM03-BR-BR1-35-24T602", "35-24T602 BROKE STOR TWR 4000M3"),
            ("35-24ES-508", "5001-PM03-BR-BR1-35-24ES-508", "35-24ES-508 WINDER AREA E-STOP"),
        ]
        floc_rows = build_floc_rows(functions)
        fn_rows = floc_rows[4:]
        self.assertEqual(fn_rows[0]["GEWRK"], "MECH")
        self.assertEqual(fn_rows[0]["EQART"], "2005")
        self.assertEqual(fn_rows[1]["GEWRK"], "MECH")  # CVYR → conveyors
        self.assertEqual(fn_rows[2]["GEWRK"], "MECH")  # pump tag prefix
        self.assertEqual(fn_rows[3]["GEWRK"], "MECH")  # TWR → tank
        self.assertEqual(fn_rows[4]["GEWRK"], "ELEC")  # ES → switch
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "h.csv"
            with p.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(
                    f,
                    fieldnames=["SUB-PROCESS", "FUNCTION", "EQUIPMENT", "SUB-EQUIPMENT", "MASK", "DESCRIPTION"],
                )
                w.writeheader()
                w.writerow(
                    {
                        "SUB-PROCESS": "BR1",
                        "FUNCTION": "35-24L004",
                        "EQUIPMENT": "",
                        "SUB-EQUIPMENT": "",
                        "MASK": "5001-PM03-BR-BR1-35-24L004",
                        "DESCRIPTION": "35-24L004 BROKE CONVEYOR 2",
                    }
                )
            rows = list(csv.DictReader(p.open(encoding="utf-8")))
            fns = collect_functions(rows)
            self.assertEqual(fns[0][2], "35-24L004 BROKE CVYR 2")

    def test_collect_functions_and_gt_eval(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "h.csv"
            with p.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(
                    f,
                    fieldnames=["SUB-PROCESS", "FUNCTION", "EQUIPMENT", "SUB-EQUIPMENT", "MASK", "DESCRIPTION"],
                )
                w.writeheader()
                w.writerow(
                    {
                        "SUB-PROCESS": "BR1",
                        "FUNCTION": "35-24L009",
                        "EQUIPMENT": "",
                        "SUB-EQUIPMENT": "",
                        "MASK": "5001-PM03-BR-BR1-35-24L009",
                        "DESCRIPTION": "35-24L009 BROKE ROLL PLPR",
                    }
                )
            rows = list(csv.DictReader(p.open(encoding="utf-8")))
            fns = collect_functions(rows)
            self.assertEqual(fns[0][0], "35-24L009")
            floc_rows = build_floc_rows(fns)
            gt = ROOT / "resources/gt_hierarchy_broke_system.xlsx"
            if gt.exists():
                report = evaluate_against_gt(floc_rows, gt)
                self.assertGreaterEqual(report["functions_compared"], 1)
                self.assertEqual(report["mask_hit"], 1)


class RealHierarchyAbbreviationTests(unittest.TestCase):
    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_conveyor_descriptions_abbreviated_in_export(self) -> None:
        rows = list(csv.DictReader(HIERARCHY_CSV.open(encoding="utf-8")))
        fns = collect_functions(rows)
        by_tag = {tag: desc for tag, _mask, desc in fns}
        for tag in ("35-24L004", "35-24L006"):
            if tag in by_tag:
                self.assertIn("CVYR", by_tag[tag], msg=f"{tag} -> {by_tag[tag]}")
                self.assertNotIn("CONVEYOR", by_tag[tag], msg=f"{tag} -> {by_tag[tag]}")

    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_no_unabbreviated_pump_motor_in_function_descriptions(self) -> None:
        rows = list(csv.DictReader(HIERARCHY_CSV.open(encoding="utf-8")))
        fns = collect_functions(rows)
        offenders = []
        for tag, _mask, desc in fns:
            words = set(desc.split())
            if "PUMP" in words or "MOTOR" in words or "CONVEYOR" in words:
                offenders.append((tag, desc))
        self.assertEqual(offenders, [], msg=json.dumps(offenders[:10], indent=2))


class LoadFunctionPositionsTests(unittest.TestCase):
    def _write_inventory(self, functions: list) -> Path:
        tmp = Path(tempfile.mktemp(suffix=".json"))
        tmp.write_text(json.dumps({"functions": functions}), encoding="utf-8")
        return tmp

    def test_extracts_tag_and_x(self) -> None:
        tmp = self._write_inventory([
            {"function": "35-24L001", "x": 100.0, "y": 200.0},
            {"function": "35-24P518", "x": 500.5, "y": 150.0},
        ])
        result = load_function_positions(tmp)
        tmp.unlink()
        self.assertAlmostEqual(result["35-24L001"], 100.0)
        self.assertAlmostEqual(result["35-24P518"], 500.5)

    def test_missing_file_returns_empty(self) -> None:
        result = load_function_positions(Path("/nonexistent/path.json"))
        self.assertEqual(result, {})

    def test_entry_without_x_skipped(self) -> None:
        tmp = self._write_inventory([
            {"function": "35-24L001", "y": 200.0},  # no x field
            {"function": "35-24P518", "x": 500.0, "y": 150.0},
        ])
        result = load_function_positions(tmp)
        tmp.unlink()
        self.assertNotIn("35-24L001", result)
        self.assertIn("35-24P518", result)

    def test_normalises_tag_to_uppercase(self) -> None:
        tmp = self._write_inventory([{"function": "35-24l001", "x": 100.0, "y": 0.0}])
        result = load_function_positions(tmp)
        tmp.unlink()
        self.assertIn("35-24L001", result)


class FunctionPositionOrderingTests(unittest.TestCase):
    """collect_functions sorts by X-coordinate when positions are provided."""

    def _rows(self, *tags: str) -> list:
        return [
            {"FUNCTION": t, "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "", "DESCRIPTION": f"{t} DESC"}
            for t in tags
        ]

    def test_sorts_functions_by_x_coordinate(self) -> None:
        rows = self._rows("35-24P518", "35-24L001", "35-24T601")
        positions = {"35-24P518": 500.0, "35-24L001": 100.0, "35-24T601": 300.0}
        result = collect_functions(rows, filter_utility_lines=False, positions=positions)
        tags = [r[0] for r in result]
        self.assertEqual(tags, ["35-24L001", "35-24T601", "35-24P518"])

    def test_unknowns_appended_at_end_in_original_order(self) -> None:
        rows = self._rows("35-24P518", "35-24L001", "35-24T601")
        positions = {"35-24L001": 100.0}  # P518 and T601 have no known position
        result = collect_functions(rows, filter_utility_lines=False, positions=positions)
        tags = [r[0] for r in result]
        self.assertEqual(tags[0], "35-24L001")
        self.assertIn("35-24P518", tags[1:])
        self.assertIn("35-24T601", tags[1:])

    def test_no_positions_preserves_csv_order(self) -> None:
        rows = self._rows("35-24P518", "35-24L001", "35-24T601")
        result = collect_functions(rows, filter_utility_lines=False, positions=None, sort_by_tag_number=False)
        tags = [r[0] for r in result]
        self.assertEqual(tags, ["35-24P518", "35-24L001", "35-24T601"])

    def test_default_sorts_by_tag_number(self) -> None:
        """Without positions, functions default to numeric tag order."""
        rows = self._rows("35-24P518", "35-24L001", "35-24T601")
        result = collect_functions(rows, filter_utility_lines=False)
        tags = [r[0] for r in result]
        # L001→1, P518→518, T601→601 → numeric order
        self.assertEqual(tags, ["35-24L001", "35-24P518", "35-24T601"])

    def test_sorting_applied_after_utility_line_filter(self) -> None:
        rows = [
            {"FUNCTION": "35-24P518", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "", "DESCRIPTION": "P518 DESC"},
            {"FUNCTION": "35-24-010", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "", "DESCRIPTION": "35-24-010 WFL DN15"},
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "", "DESCRIPTION": "L001 DESC"},
        ]
        positions = {"35-24P518": 500.0, "35-24L001": 100.0, "35-24-010": 50.0}
        result = collect_functions(rows, positions=positions)
        tags = [r[0] for r in result]
        # WFL line filtered; remaining sorted by X: L001 (100) before P518 (500)
        self.assertNotIn("35-24-010", tags)
        self.assertEqual(tags, ["35-24L001", "35-24P518"])

    @unittest.skipUnless(
        (Path(__file__).resolve().parents[1] / "outputs/jsons/Broke System.pid_inventory.json").exists(),
        "requires Broke System inventory JSON",
    )
    def test_real_broke_system_sorted_before_pumps(self) -> None:
        """L-coded machines should not all precede P-coded pumps after position sort."""
        import csv as csv_mod
        inv_path = Path(__file__).resolve().parents[1] / "outputs/jsons/Broke System.pid_inventory.json"
        hier_csv = Path(__file__).resolve().parents[1] / "outputs/Broke System.hierarchy_orchestrator.csv"
        if not hier_csv.exists():
            self.skipTest("requires hierarchy orchestrator CSV")
        rows = list(csv_mod.DictReader(hier_csv.open(encoding="utf-8")))
        positions = load_function_positions(inv_path)
        result = collect_functions(rows, positions=positions)
        tags = [r[0] for r in result]
        # With position sort, a pump should appear before some L-line in the list.
        # (Without position sort all L-codes precede all P-codes.)
        p_indices = [i for i, t in enumerate(tags) if t.startswith("35-24P")]
        l_indices = [i for i, t in enumerate(tags) if re.match(r"35-24L\d", t)]
        if p_indices and l_indices:
            # At least one pump must appear before the last L-machine (interleaved)
            self.assertLess(min(p_indices), max(l_indices))

    def test_collect_functions_passes_flow_codes_from_ecosystem(self) -> None:
        """collect_functions uses ecosystem.standard.flow_substance_codes when ecosystem provided."""
        from dwg_reader.dwg_ecosystem import Ecosystem
        eco = Ecosystem(
            name="valmet",
            standard_id="valmet_ps21",
            standard={"flow_substance_codes": {"WAF": "CUSTOM WTR"}},
        )
        rows = [
            {"FUNCTION": "35-24-016", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "MASK": "",
             "DESCRIPTION": "35-24-016 WAF 300 DIST HDR"},
        ]
        result = collect_functions(rows, filter_utility_lines=False, ecosystem=eco)
        descs = [desc for _, _, desc in result]
        self.assertTrue(any("CUSTOM WTR" in d for d in descs), f"Expected CUSTOM WTR in {descs}")


    def test_unknown_non_line_eqart_blank(self) -> None:
        functions = [
            ("35-24Z999", "5001-PM03-BR-BR1-35-24Z999", "35-24Z999 MYSTERY WIDGET"),
        ]
        row = next(r for r in build_floc_rows(functions) if r["TPLNR"].endswith("35-24Z999"))
        self.assertEqual(row["EQART"], "")

    def test_line_function_unknown_type_still_2100(self) -> None:
        functions = [
            ("35-24-095", "5001-PM03-BR-BR1-35-24-095", "35-24-095 UNKNOWN THING"),
        ]
        row = next(r for r in build_floc_rows(functions) if r["TPLNR"].endswith("35-24-095"))
        self.assertEqual(row["EQART"], "2100")

    def test_pump_pltxt_strips_conveyor_tokens(self) -> None:
        functions = [
            ("35-24P507", "5001-PM03-BR-BR1-35-24P507", "35-24P507 BROKE CVYR 3 PMP"),
        ]
        row = next(r for r in build_floc_rows(functions) if r["TPLNR"].endswith("35-24P507"))
        self.assertNotIn("CVYR", row["PLTXT"])
        self.assertIn("PMP", row["PLTXT"])

    def test_pump_pltxt_inherits_parent_vessel_duty(self) -> None:
        functions = [
            ("35-24L005", "5001-PM03-BR-BR1-35-24L005", "35-24L005 REEL PLPR"),
            ("35-24P507", "5001-PM03-BR-BR1-35-24P507", "35-24P507 BROKE CVYR 3 PMP"),
        ]
        row = next(r for r in build_floc_rows(functions) if r["TPLNR"].endswith("35-24P507"))
        self.assertNotIn("CVYR", row["PLTXT"])
        self.assertIn("REEL", row["PLTXT"])
        self.assertIn("PMP", row["PLTXT"])


if __name__ == "__main__":
    unittest.main()
