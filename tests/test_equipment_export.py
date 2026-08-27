#!/usr/bin/env python3
"""Unit tests for SAP Equipment export (no Bedrock)."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from dwg_reader.dwg_floc_context import (
    format_line_eqktx,
    format_valve_eqktx,
    infer_valve_type,
    is_line_equipment_tag,
    is_phantom_motor_line_tag,
    is_valve_equipment,
    is_valve_tag,
    normalize_pltxt,
    strip_valve_prefix,
)
from dwg_reader.dwg_object_type import classify_equipment, lookup
from dwg_reader.export_sap_equipment import (
    SAP_COLUMNS,
    _is_driven_equipment,
    _motor_eqktx,
    _motor_tag_for,
    _valve_hint,
    build_equipment_rows,
    write_equipment_workbook,
)

ROOT = Path(__file__).resolve().parents[1]
HIERARCHY_CSV = ROOT / "outputs/Broke System.hierarchy_orchestrator.csv"


class ObjectTypeClassifierTests(unittest.TestCase):
    """Tests for dwg_object_type.classify_equipment."""

    def _cls(self, tag, desc):
        return classify_equipment(tag, desc)

    # --- tag prefix rules ---
    def test_level_controller_tag(self):
        # LC = Level Controller → 1204 INSTRUMENT LEVEL (not generic 1200)
        code, wc = self._cls("35-24LC-576", "35-24LC-576 LVL CTRL")
        self.assertEqual(code, "1204")
        self.assertEqual(wc, "INST")

    def test_control_valve_hv_tag(self):
        # HV = Hand Valve → 201 VALVE HAND
        code, wc = self._cls("35-24HV-626", "35-24HV-626 DN300 HND VLV")
        self.assertEqual(code, "201")
        self.assertEqual(wc, "MECH")

    def test_flow_valve_fv_tag(self):
        code, wc = self._cls("35-24FV-570", "35-24FV-570 FLOW VLV")
        self.assertEqual(code, "202")
        self.assertEqual(wc, "MECH")

    def test_switch_xs_tag(self):
        code, wc = self._cls("35-24XS-588", "35-24XS-588 MCS SIG")
        self.assertEqual(code, "1108")

    # --- description keyword rules ---
    def test_motor_keyword(self):
        code, wc = self._cls("35-24-001.1", "35-24-001.1 PRESS PLPR ROTOR 1 MTR")
        self.assertEqual(code, "1101")
        self.assertEqual(wc, "ELEC")

    def test_pump_keyword(self):
        code, wc = self._cls("35-24P519", "35-24P519 PMP")
        self.assertEqual(code, "701")
        self.assertEqual(wc, "MECH")

    def test_pulper_keyword(self):
        code, wc = self._cls("35-24L009", "35-24L009 BROKE ROLL PLPR")
        self.assertEqual(code, "2005")
        self.assertEqual(wc, "MECH")

    def test_hand_valve_keyword(self):
        code, wc = self._cls("35-24-207", "35-24-207 CIRC VLV 003-50")
        self.assertEqual(code, "201")
        self.assertEqual(wc, "MECH")

    def test_pipe_line_keyword(self):
        code, wc = self._cls("35-24-149", "35-24-149 WAA DN300 INLET LINE")
        self.assertEqual(code, "2100")
        self.assertEqual(wc, "MECH")

    def test_thickener_keyword(self):
        code, wc = self._cls("35-24L002", "35-24L002 BROKE THICKENER")
        self.assertEqual(code, "2009")
        self.assertEqual(wc, "MECH")

    # --- lookup helper ---
    def test_lookup(self):
        name, wc = lookup("701")
        self.assertEqual(name, "PUMP CENTRIFUGAL")
        self.assertEqual(wc, "MECH")

    def test_fallback(self):
        code, wc = classify_equipment("35-24-999", "35-24-999 UNKNOWN THING")
        self.assertEqual(code, "9999")
        self.assertEqual(wc, "")

    def test_pump_single_letter_tag(self):
        # "35-24P519" — single-letter no-dash format → P → 701 without needing PMP in description
        code, wc = self._cls("35-24P519", "35-24P519 OUTLET")
        self.assertEqual(code, "701")
        self.assertEqual(wc, "MECH")

    def test_motor_suffix_tag(self):
        # ".N" suffix convention marks motor sub-equipment even with a thin description
        code, wc = self._cls("35-24-001.1", "35-24-001.1 ROTOR 1")
        self.assertEqual(code, "1101")
        self.assertEqual(wc, "ELEC")

    def test_motor_suffix_beats_pulper_keywords(self):
        # L008.1 copied the winder-pulper parent text; still a motor.
        code, wc = self._cls("35-24-008.1", "35-24-008.1 WINDER PLPR")
        self.assertEqual(code, "1101")
        self.assertEqual(wc, "ELEC")

    def test_plpr_context_disambiguation(self):
        # PLPR in a child description is context (parent system name), not the equipment type
        code, wc = classify_equipment("35-24-095", "35-24-095 PRESS PLPR OUTLET PP-200")
        self.assertEqual(code, "2100")   # pipe line, not pulper
        self.assertEqual(wc, "MECH")
        code, _ = classify_equipment("35-24-101", "35-24-101 PRESS PLPR SUCT PP-150")
        self.assertEqual(code, "2100")
        # Motor on a pulper still classifies correctly (MTR rule fires first)
        code, wc = classify_equipment("35-24-001.1", "35-24-001.1 PRESS PLPR ROTOR 1 MTR")
        self.assertEqual(code, "1101")
        self.assertEqual(wc, "ELEC")


class FormatLineEqktxTests(unittest.TestCase):
    def test_review_examples(self) -> None:
        self.assertEqual(
            format_line_eqktx(
                "35-24-095",
                normalize_pltxt("35-24-095 PRESS PLPR PP-200 LINE"),
            ),
            "LN 35-24-095 PRESS PLPR PP-200",
        )
        self.assertEqual(
            format_line_eqktx(
                "35-24-096",
                normalize_pltxt("35-24-096 PRESS PLPR PP-900 LINE"),
            ),
            "LN 35-24-096 PRESS PLPR PP-900",
        )

    def test_is_line_equipment_tag(self) -> None:
        self.assertTrue(is_line_equipment_tag("35-24-095"))
        self.assertTrue(is_line_equipment_tag("55-30-001"))
        self.assertTrue(is_line_equipment_tag("168L-522"))
        self.assertFalse(is_line_equipment_tag("35-24-001.1"))
        self.assertFalse(is_line_equipment_tag("35-24LC-576"))
        self.assertFalse(is_line_equipment_tag("35-24HV-548"))
        self.assertFalse(is_line_equipment_tag("35-24P519"))

    def test_phantom_motor_line_tag(self) -> None:
        self.assertTrue(is_phantom_motor_line_tag("35-24-1300", "35-24-1300 PMP MTR"))
        self.assertTrue(is_phantom_motor_line_tag("35-24-1300", "LN 35-24-1300 PMP MTR"))
        self.assertFalse(is_phantom_motor_line_tag("35-24-119", "35-24-119 SUCT LN"))
        self.assertFalse(is_phantom_motor_line_tag("35-24P501", "35-24P501 COUCH PIT PMP"))
        self.assertFalse(is_phantom_motor_line_tag("35-24-501.1", "35-24-501.1 COUCH PIT PMP MTR"))

    def test_idempotent_when_already_prefixed(self) -> None:
        text = "LN 35-24-095 PRESS PLPR PP-200"
        self.assertEqual(format_line_eqktx("35-24-095", text), text)

    def test_strips_duplicate_tag_after_ln_prefix(self) -> None:
        self.assertEqual(
            format_line_eqktx("35-24-185", "LN 35-24-185 35-24-185 WW > BROKE ROLL P"),
            "LN 35-24-185 WW > BROKE ROLL P",
        )

    def test_overflow_line_with_embedded_ln(self) -> None:
        self.assertEqual(
            format_line_eqktx("35-24-189", normalize_pltxt("35-24-189 LN OVFL")),
            "LN 35-24-189 OVFL",
        )

    def test_skips_valve_sub_equipment(self) -> None:
        text = normalize_pltxt("35-24-207 VLV ON 35-24-096")
        self.assertEqual(
            format_line_eqktx("35-24-207", text, hequi="35-24-096"),
            text,
        )

    def test_skips_valve_description_without_hequi(self) -> None:
        text = normalize_pltxt("35-24-234 VLV ON 35-24-093")
        self.assertEqual(format_line_eqktx("35-24-234", text), text)

    def test_applies_without_trailing_line_word(self) -> None:
        self.assertEqual(
            format_line_eqktx("35-24-149", "35-24-149 WAA DN300 INLET"),
            "LN 35-24-149 WAA DN300 INLET",
        )

    def test_max_length_40(self) -> None:
        long_desc = "35-24-095 " + "PRESS PLPR PP-200 EXTRA WORDS " * 3
        out = format_line_eqktx("35-24-095", normalize_pltxt(long_desc))
        self.assertLessEqual(len(out), 40)
        self.assertTrue(out.startswith("LN 35-24-095"))
        self.assertFalse(out.endswith("-"))
        self.assertFalse(out.endswith(" "))

    def test_drops_peer_tag_to_keep_destination(self) -> None:
        out = format_line_eqktx(
            "35-24-100",
            normalize_pltxt(
                "35-24-100 35-24-108 SUCTION LINE TO BROKE CONVEYOR 1",
                max_len=80,
            ),
        )
        self.assertTrue(out.startswith("LN 35-24-100"))
        self.assertLessEqual(len(out), 40)
        self.assertNotIn("35-24-108", out)
        self.assertFalse(out.endswith("-"))
        self.assertTrue(
            any(tok in out for tok in ("SUCT", "CVYR", "BROKE")),
            out,
        )

    def test_strips_truncated_peer_tag_remnant(self) -> None:
        self.assertEqual(
            format_line_eqktx("35-24-100", "LN 35-24-100 35-24-"),
            "LN 35-24-100",
        )


class EquipmentExportTests(unittest.TestCase):
    def test_build_rows_hequi_and_posnr(self) -> None:
        rows = [
            {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L009 BROKE ROLL PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-189", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-189 LN OVFL"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-194", "DESCRIPTION": "35-24-194 HV DRN"},
            {"FUNCTION": "", "EQUIPMENT": "35-24P519", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P519 PMP"},
        ]
        out = build_equipment_rows(rows)
        # 35-24L009(fn-equip), 35-24-189, 35-24-194, 35-24P519, motor(pump), motor(pulper)
        # Pulpers (PLPR in description) match the winder_pulper driven_pattern → get motors
        self.assertEqual(len(out), 6)
        by_tag = {r["EQUNR"]: r for r in out}

        # Function-level equipment for 35-24L009
        self.assertEqual(by_tag["35-24L009"]["POSNR"], "0001")
        self.assertEqual(by_tag["35-24L009"]["HEQUI"], "")
        self.assertEqual(by_tag["35-24L009"]["TPLNR"], "5001-PM03-BR-BR1-35-24L009")

        # Regular equipment children (use by_tag to avoid index brittleness)
        self.assertEqual(by_tag["35-24-189"]["EQUNR"], "35-24-189")
        self.assertEqual(by_tag["35-24-189"]["HEQUI"], "")
        self.assertEqual(by_tag["35-24-189"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24-189"]["TPLNR"], "5001-PM03-BR-BR1-35-24L009")
        self.assertEqual(by_tag["35-24-189"]["EQTYP"], "P")
        self.assertEqual(by_tag["35-24-189"]["EQART"], "2100")  # LN keyword → pipeline
        self.assertEqual(by_tag["35-24-189"]["GEWRK"], "MECH")
        self.assertEqual(by_tag["35-24-189"]["SWERK"], "5001")
        self.assertEqual(by_tag["35-24-189"]["ABCKZ"], "D")
        self.assertEqual(by_tag["35-24-189"]["INGRP"], "P01")
        self.assertLessEqual(len(by_tag["35-24-189"]["EQKTX"]), 40)
        # pump tag (35-24P519 PMP) → PUMP CENTRIFUGAL
        self.assertEqual(by_tag["35-24P519"]["EQART"], "701")
        self.assertEqual(by_tag["35-24P519"]["GEWRK"], "MECH")

        self.assertEqual(by_tag["35-24-194"]["EQUNR"], "35-24-194")
        self.assertEqual(by_tag["35-24-194"]["HEQUI"], "35-24-189")
        self.assertEqual(by_tag["35-24-194"]["POSNR"], "0010")

        self.assertEqual(by_tag["35-24P519"]["EQUNR"], "35-24P519")
        self.assertEqual(by_tag["35-24P519"]["HEQUI"], "")
        self.assertEqual(by_tag["35-24P519"]["POSNR"], "0020")

        # implicit motor injected as sub-equipment of the pump
        self.assertEqual(by_tag["35-24-519.1"]["EQUNR"], "35-24-519.1")
        self.assertEqual(by_tag["35-24-519.1"]["HEQUI"], "35-24P519")
        self.assertEqual(by_tag["35-24-519.1"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24-519.1"]["EQART"], "1101")
        self.assertEqual(by_tag["35-24-519.1"]["GEWRK"], "ELEC")
        self.assertEqual(by_tag["35-24-519.1"]["TPLNR"], by_tag["35-24P519"]["TPLNR"])

    def test_motor_tag_for_pump(self) -> None:
        self.assertEqual(_motor_tag_for("35-24P518"), "35-24-518.1")
        self.assertEqual(_motor_tag_for("35-24P519"), "35-24-519.1")

    def test_motor_tag_for_agitator(self) -> None:
        self.assertEqual(_motor_tag_for("35-24L404"), "35-24-404.1")
        self.assertEqual(_motor_tag_for("35-24L499"), "35-24-499.1")

    def test_motor_tag_for_tissue(self) -> None:
        from dwg_reader.dwg_ecosystem import detect
        eco = detect("GORA68210")
        self.assertEqual(_motor_tag_for("124P-001", ecosystem=eco), "124P-001-M1")

    def test_motor_tag_for_line_tag_returns_empty(self) -> None:
        # 35-24-095 (line/pipe tag, no letter code) → no motor derivable
        self.assertEqual(_motor_tag_for("35-24-095"), "")

    def test_is_driven_pump(self) -> None:
        self.assertTrue(_is_driven_equipment("35-24P501"))
        self.assertTrue(_is_driven_equipment("35-24P518"))

    def test_is_driven_agitator_range(self) -> None:
        self.assertTrue(_is_driven_equipment("35-24L401"))
        self.assertTrue(_is_driven_equipment("35-24L499"))

    def test_is_not_driven_below_range(self) -> None:
        self.assertFalse(_is_driven_equipment("35-24L002"))
        self.assertFalse(_is_driven_equipment("35-24L400"))

    def test_is_not_driven_line_tag(self) -> None:
        self.assertFalse(_is_driven_equipment("35-24-095"))

    def test_is_driven_screen_by_desc(self) -> None:
        self.assertTrue(_is_driven_equipment("35-24L010", desc="35-24L010 BROKE SCRN"))

    def test_is_driven_screen_full_word(self) -> None:
        self.assertTrue(_is_driven_equipment("35-24L010", desc="35-24L010 BROKE SCREEN"))

    def test_is_driven_gearbox_by_desc(self) -> None:
        self.assertTrue(_is_driven_equipment("35-24L015", desc="35-24L015 GRBX INLINE"))

    def test_is_driven_gearbox_full_word(self) -> None:
        self.assertTrue(_is_driven_equipment("35-24L015", desc="35-24L015 GEARBOX"))

    def test_is_driven_agitator_by_desc_outside_range(self) -> None:
        # L010 is outside L401-L499 but description says AGITATOR → should be driven
        self.assertTrue(_is_driven_equipment("35-24L010", desc="35-24L010 AGITATOR"))

    def test_is_driven_agitator_keyword_agit(self) -> None:
        self.assertTrue(_is_driven_equipment("35-24L010", desc="35-24L010 AGIT TANK"))

    def test_is_not_driven_l010_without_desc(self) -> None:
        # L010 outside range; no desc → not driven by tag alone
        self.assertFalse(_is_driven_equipment("35-24L010"))

    def test_is_not_driven_plain_process_equip(self) -> None:
        self.assertFalse(_is_driven_equipment("35-24L002", desc="35-24L002 BROKE THICKENER"))

    def test_is_driven_agitator_rotor(self) -> None:
        # Pulper rotor sub-equipment tags (35-24L009.1) are always driven
        self.assertTrue(_is_driven_equipment("35-24L009.1"))
        self.assertTrue(_is_driven_equipment("35-24L009.2"))

    def test_motor_tag_for_rotor(self) -> None:
        # Strip letter prefix, keep decimal suffix: 35-24L009.1 → 35-24-009.1
        self.assertEqual(_motor_tag_for("35-24L009.1"), "35-24-009.1")
        self.assertEqual(_motor_tag_for("35-24L009.2"), "35-24-009.2")

    def test_implicit_motor_for_rotor_in_equipment_column(self) -> None:
        """Agitator rotors (L009.1) in the EQUIPMENT column must get a motor injected."""
        rows = [
            {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L009 BROKE ROLL PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24L009.1", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L009.1 BROKE ROLL PLPR RTR 1"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24-009.1", by_tag, "rotor motor must be injected")
        motor = by_tag["35-24-009.1"]
        self.assertEqual(motor["HEQUI"], "35-24L009.1", "motor parent must be the rotor, not the pulper")
        self.assertEqual(motor["EQART"], "1101")
        self.assertEqual(motor["GEWRK"], "ELEC")

    def test_implicit_motor_for_rotor_in_subequipment_column(self) -> None:
        """Agitator rotors in the SUB-EQUIPMENT column (HEQUI set) must also get a motor."""
        rows = [
            {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L009 BROKE ROLL PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24L009", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L009 BROKE ROLL PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24L009.1", "DESCRIPTION": "35-24L009.1 BROKE ROLL PLPR RTR 1"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24-009.1", by_tag, "rotor motor must be injected even as sub-equipment")
        motor = by_tag["35-24-009.1"]
        self.assertEqual(motor["HEQUI"], "35-24L009.1")

    def test_implicit_motor_for_screen(self) -> None:
        """Screens (SCRN in description) must have a motor injected when absent."""
        rows = [
            {"FUNCTION": "35-24L004", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "PULPER"},
            {"FUNCTION": "", "EQUIPMENT": "35-24L010", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L010 BROKE SCRN"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24-010.1", by_tag, "screen motor must be injected")
        motor = by_tag["35-24-010.1"]
        self.assertEqual(motor["HEQUI"], "35-24L010")
        self.assertEqual(motor["EQART"], "1101")
        self.assertEqual(motor["GEWRK"], "ELEC")

    def test_implicit_motor_for_gearbox(self) -> None:
        """Gearboxes (GRBX in description) must have a motor injected when absent."""
        rows = [
            {"FUNCTION": "35-24L004", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "PULPER"},
            {"FUNCTION": "", "EQUIPMENT": "35-24L015", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L015 GRBX"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24-015.1", by_tag, "gearbox motor must be injected")
        motor = by_tag["35-24-015.1"]
        self.assertEqual(motor["HEQUI"], "35-24L015")
        self.assertEqual(motor["EQART"], "1101")

    def test_motor_not_duplicated_when_in_hierarchy(self) -> None:
        """Motor already on diagram must not be injected a second time."""
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24P518", "SUB-EQUIPMENT": "", "DESCRIPTION": "PUMP"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-518.1", "DESCRIPTION": "MOTOR"},
        ]
        out = build_equipment_rows(rows)
        motor_rows = [r for r in out if r["EQUNR"] == "35-24-518.1"]
        self.assertEqual(len(motor_rows), 1, "motor must not be emitted twice")

    def test_implicit_motor_for_agitator(self) -> None:
        """L401–L499 agitator tags trigger motor injection."""
        rows = [
            {"FUNCTION": "35-24L004", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "PULPER"},
            {"FUNCTION": "", "EQUIPMENT": "35-24L404", "SUB-EQUIPMENT": "", "DESCRIPTION": "AGITATOR"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24-404.1", by_tag, "agitator motor must be injected")
        motor = by_tag["35-24-404.1"]
        self.assertEqual(motor["HEQUI"], "35-24L404")
        self.assertEqual(motor["EQART"], "1101")

    def test_motor_eqktx_strips_parent_type_tokens(self) -> None:
        eqktx = _motor_eqktx(
            "35-24-406.1",
            "35-24L406",
            "35-24L406 BROKE COLLECTION AGI TNK",
        )
        self.assertIn("BROKE", eqktx.upper())
        self.assertIn("MTR", eqktx.upper())
        self.assertNotIn("TNK", eqktx.upper())
        self.assertNotIn("AGI", eqktx.upper())
        self.assertLessEqual(len(eqktx), 40)

    def test_motor_eqktx_strips_pmp_token(self) -> None:
        eqktx = _motor_eqktx("35-24-518.1", "35-24P518", "35-24P518 BROKE REJECT PMP")
        self.assertIn("BROKE", eqktx.upper())
        self.assertTrue("REJ" in eqktx.upper())
        self.assertIn("MTR", eqktx.upper())
        self.assertNotIn("PMP", eqktx.upper())

    def test_motor_eqktx_inherits_parent_description(self) -> None:
        """Motor text follows Example for motors.docx: parent desc + MTR."""
        self.assertIn(
            "BROKE",
            _motor_eqktx("35-24-518.1", "35-24P518", "35-24P518 BROKE REJECT PMP").upper(),
        )
        self.assertTrue(
            _motor_eqktx("35-24-518.1", "35-24P518", "35-24P518 BROKE REJECT PMP")
            .upper()
            .endswith("MTR")
            or " MTR" in _motor_eqktx("35-24-518.1", "35-24P518", "35-24P518 BROKE REJECT PMP").upper()
        )
        rows = [
            {"FUNCTION": "35-24T606", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "TANK"},
            {
                "FUNCTION": "",
                "EQUIPMENT": "35-24L404",
                "SUB-EQUIPMENT": "",
                "DESCRIPTION": "35-24L404 BROKE REJECT AGITATOR TANK",
            },
        ]
        out = build_equipment_rows(rows)
        motor = next(r for r in out if r["EQUNR"] == "35-24-404.1")
        self.assertEqual(motor["HEQUI"], "35-24L404")
        self.assertIn("BROKE", motor["EQKTX"].upper())
        self.assertIn("MTR", motor["EQKTX"].upper())
        self.assertNotIn("TNK", motor["EQKTX"].upper())
        self.assertNotEqual(motor["EQKTX"].upper(), "35-24-404.1 MTR")

    def test_sub_equipment_posnr_resets_per_parent(self) -> None:
        """Review issue 3: subs reset POSNR; top-level continues separately."""
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "LINE A"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24LV2-576", "DESCRIPTION": "LVL VLV"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-096", "SUB-EQUIPMENT": "", "DESCRIPTION": "LINE B"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-207", "DESCRIPTION": "VLV 1"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-217", "DESCRIPTION": "VLV 2"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-1105", "DESCRIPTION": "VLV 3"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-093", "SUB-EQUIPMENT": "", "DESCRIPTION": "LINE C"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertEqual(by_tag["35-24-095"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24LV2-576"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24-096"]["POSNR"], "0020")
        self.assertEqual(by_tag["35-24-207"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24-217"]["POSNR"], "0020")
        self.assertEqual(by_tag["35-24-1105"]["POSNR"], "0030")
        self.assertEqual(by_tag["35-24-093"]["POSNR"], "0030")

    def test_line_eqktx_in_export(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-095 PRESS PLPR PP-200 LINE"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-096", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-096 PRESS PLPR PP-900 LINE"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-207", "DESCRIPTION": "35-24-207 VLV ON 35-24-096"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertEqual(by_tag["35-24-095"]["EQKTX"], "LN 35-24-095 PRESS PLPR")
        self.assertEqual(by_tag["35-24-096"]["EQKTX"], "LN 35-24-096 PRESS PLPR")
        self.assertNotIn("DN", by_tag["35-24-095"]["EQKTX"])
        self.assertNotIn("PP-200", by_tag["35-24-095"]["EQKTX"])
        self.assertFalse(by_tag["35-24-207"]["EQKTX"].startswith("LN "))

    def test_pipeline_eqktx_keeps_destination_not_peer_tag(self) -> None:
        rows = [
            {"FUNCTION": "35-24L011", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "BROKE CVYR 1"},
            {
                "FUNCTION": "",
                "EQUIPMENT": "35-24-100",
                "SUB-EQUIPMENT": "",
                "DESCRIPTION": "35-24-100 35-24-108 SUCTION LINE PP-250 TO BROKE CONVEYOR 1",
            },
        ]
        out = build_equipment_rows(rows)
        eqktx = next(r["EQKTX"] for r in out if r["EQUNR"] == "35-24-100")
        self.assertTrue(eqktx.startswith("LN 35-24-100"))
        self.assertLessEqual(len(eqktx), 40)
        self.assertNotIn("35-24-108", eqktx)
        self.assertFalse(eqktx.endswith("-"))
        self.assertTrue(any(tok in eqktx for tok in ("SUCT", "CVYR", "BROKE")), eqktx)

    def test_drops_nameplate_misread_as_pipeline(self) -> None:
        rows = [
            {"FUNCTION": "35-24P501", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P501 COUCH PIT PMP"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-1300", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-1300 PMP MTR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-119", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-119 SUCT LN"},
        ]
        out = build_equipment_rows(rows)
        tags = {r["EQUNR"] for r in out}
        self.assertNotIn("35-24-1300", tags)
        self.assertIn("35-24-501.1", tags)
        self.assertIn("35-24-119", tags)
        suction = next(r for r in out if r["EQUNR"] == "35-24-119")
        self.assertTrue(suction["EQKTX"].startswith("LN 35-24-119"))

    def test_limit_functions(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-001.1", "SUB-EQUIPMENT": "", "DESCRIPTION": "MTR"},
            {"FUNCTION": "35-24L002", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-002.1", "SUB-EQUIPMENT": "", "DESCRIPTION": "MTR"},
        ]
        out = build_equipment_rows(rows, limit_functions=1)
        # With function equipment: 35-24L001 (fn-equip) + 35-24-001.1 (child) = 2 rows
        self.assertEqual(len(out), 2)
        equnrs = {r["EQUNR"] for r in out}
        self.assertIn("35-24L001", equnrs)  # function-level equipment
        self.assertIn("35-24-001.1", equnrs)  # regular child
        # 35-24L001 equipment installed at its own FLOC
        self.assertTrue(any("35-24L001" in r["TPLNR"] for r in out if r["EQUNR"] == "35-24L001"))

    def test_machine_function_emitted_as_equipment(self) -> None:
        """M4-8: machine FUNCTION tags (L, P, T) must also appear as Equipment records."""
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L001 PRESS PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "LN SUCT"},
            {"FUNCTION": "35-24P501", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P501 COUCH PIT PMP"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        # Machine functions must be in Equipment sheet
        self.assertIn("35-24L001", by_tag, "pulper must appear as Equipment record")
        self.assertIn("35-24P501", by_tag, "pump must appear as Equipment record")
        # They are installed at their own FLOC
        self.assertIn("35-24L001", by_tag["35-24L001"]["TPLNR"])

    def test_machine_function_equipment_not_emitted_when_disabled(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L001 PRESS PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "LN SUCT"},
        ]
        out = build_equipment_rows(rows, include_function_equipment=False)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertNotIn("35-24L001", by_tag, "function equipment must be absent when disabled")
        self.assertIn("35-24-095", by_tag)

    def test_pipeline_function_not_emitted_as_equipment(self) -> None:
        """35-24-NNN line FUNCTIONs must NOT be emitted as Equipment records."""
        rows = [
            {"FUNCTION": "35-24-095", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "LN 35-24-095 SUCT"},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "HV VLV"},
        ]
        out = build_equipment_rows(rows)
        equnrs = {r["EQUNR"] for r in out}
        self.assertNotIn("35-24-095", equnrs, "pipeline FUNCTION must not be emitted as Equipment")

    def test_write_workbook(self) -> None:
        template = ROOT / "docs/examples/SML-Equipment Template RW.xlsx"
        self.assertTrue(template.exists())
        rows = build_equipment_rows(
            [
                {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
                {"FUNCTION": "", "EQUIPMENT": "35-24-189", "SUB-EQUIPMENT": "", "DESCRIPTION": "OVFL"},
            ]
        )
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "eq.xlsx"
            write_equipment_workbook(template, out, rows)
            self.assertTrue(out.exists())
            from openpyxl import load_workbook

            wb = load_workbook(out)
            self.assertIn("Equipment", wb.sheetnames)
            ws = wb["Equipment"]
            # Row 7 is now function-level Equipment for 35-24L009
            self.assertEqual(ws.cell(7, 2).value, "5001-PM03-BR-BR1-35-24L009")
            self.assertEqual(ws.cell(7, 3).value, "35-24L009")  # function-level Equipment, NEW
            self.assertEqual(ws.cell(8, 3).value, "35-24-189")  # was row 7, now row 8
            eqart_col = 2 + SAP_COLUMNS.index("EQART")
            posnr_col = 2 + SAP_COLUMNS.index("POSNR")
            self.assertEqual(ws.cell(8, eqart_col).number_format, "@")
            self.assertEqual(ws.cell(8, posnr_col).number_format, "@")
            self.assertIsInstance(ws.cell(8, eqart_col).value, str)
            self.assertIsInstance(ws.cell(8, posnr_col).value, str)


class RealHierarchyLineEqktxTests(unittest.TestCase):
    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_all_numeric_line_equipment_prefixed_with_ln(self) -> None:
        import re

        from dwg_reader.export_sap_equipment import read_hierarchy_csv

        hierarchy = read_hierarchy_csv(HIERARCHY_CSV)
        out = build_equipment_rows(hierarchy, limit_functions=5)
        line_pat = re.compile(r"^35-24-\d+$")
        valve_markers = ("VLV", "VALVE", " HV", " FV", " LV", " ON 35-24-")

        missing_ln = []
        has_line_word = []
        for row in out:
            tag = row["EQUNR"]
            eqktx = row["EQKTX"]
            if not line_pat.match(tag):
                continue
            if eqktx.startswith("HV "):
                continue
            if any(m in eqktx for m in valve_markers):
                continue
            if re.search(r"\b(NC|DRN|CHK|FLS|SMP|PRV|SV)\b", eqktx):
                continue
            if not eqktx.startswith("LN "):
                missing_ln.append((tag, eqktx))
            if eqktx.endswith(" LINE") or eqktx.endswith(" LN"):
                has_line_word.append((tag, eqktx))

        self.assertEqual(missing_ln, [], msg=f"missing LN prefix: {missing_ln[:10]}")
        self.assertEqual(has_line_word, [], msg=f"trailing LINE/LN: {has_line_word[:10]}")

    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_l001_review_line_examples(self) -> None:
        from dwg_reader.export_sap_equipment import read_hierarchy_csv

        out = build_equipment_rows(
            read_hierarchy_csv(HIERARCHY_CSV),
            limit_functions=1,
        )
        by_tag = {r["EQUNR"]: r for r in out}
        # Pipeline tags under 35-24L001 (first function in current CSV)
        self.assertTrue(by_tag["35-24-095"]["EQKTX"].startswith("LN 35-24-095"))
        self.assertTrue(by_tag["35-24-096"]["EQKTX"].startswith("LN 35-24-096"))
        # Level controller tag must NOT get LN prefix (not a pipeline)
        self.assertFalse(by_tag["35-24LC-576"]["EQKTX"].startswith("LN "))


class ValveFormattingTests(unittest.TestCase):
    """Tests for Issue 5 — valve HV-scheme equipment text."""

    # --- is_valve_tag ---

    def test_is_valve_tag_hv(self) -> None:
        self.assertTrue(is_valve_tag("35-24HV-548"))

    def test_is_valve_tag_fv(self) -> None:
        self.assertTrue(is_valve_tag("35-24FV-570"))

    def test_is_valve_tag_lv(self) -> None:
        self.assertTrue(is_valve_tag("35-24LV-622"))

    def test_is_valve_tag_lv_with_digit(self) -> None:
        # 35-24LV2-576 — LV2 is the prefix (second level valve on same line)
        self.assertTrue(is_valve_tag("35-24LV2-576"))

    def test_is_valve_tag_xv(self) -> None:
        self.assertTrue(is_valve_tag("35-24XV-669"))

    def test_is_valve_tag_kv(self) -> None:
        self.assertTrue(is_valve_tag("35-24KV-573"))

    def test_is_valve_tag_plain_line_false(self) -> None:
        self.assertFalse(is_valve_tag("35-24-207"))

    def test_is_valve_tag_pump_false(self) -> None:
        self.assertFalse(is_valve_tag("35-24P519"))

    def test_is_valve_tag_switch_false(self) -> None:
        # XS = switch, not a valve prefix
        self.assertFalse(is_valve_tag("35-24XS-588"))

    def test_is_valve_tag_level_controller_false(self) -> None:
        self.assertFalse(is_valve_tag("35-24LC-576"))

    # --- strip_valve_prefix ---

    def test_strip_hv(self) -> None:
        # Rob review example 1
        self.assertEqual(strip_valve_prefix("35-24HV-548"), "35-24-548")

    def test_strip_fv(self) -> None:
        self.assertEqual(strip_valve_prefix("35-24FV-570"), "35-24-570")

    def test_strip_lv_with_digit(self) -> None:
        # Position digit preserved so LV2-576 and a hypothetical LV1-576 stay distinct.
        self.assertEqual(strip_valve_prefix("35-24LV2-576"), "35-24-2-576")

    def test_strip_lv1(self) -> None:
        self.assertEqual(strip_valve_prefix("35-24LV1-560"), "35-24-1-560")

    def test_strip_lv2_distinct_from_lv1(self) -> None:
        # LV1-560 and LV2-560 must not both collapse to 35-24-560.
        self.assertNotEqual(
            strip_valve_prefix("35-24LV1-560"),
            strip_valve_prefix("35-24LV2-560"),
        )

    def test_strip_plain_tag_unchanged(self) -> None:
        # Plain line tag — no embedded letters to strip
        self.assertEqual(strip_valve_prefix("35-24-137"), "35-24-137")

    def test_strip_plain_207_unchanged(self) -> None:
        self.assertEqual(strip_valve_prefix("35-24-207"), "35-24-207")

    # --- infer_valve_type ---

    def test_infer_av_from_desc_keyword(self) -> None:
        self.assertEqual(infer_valve_type("35-24HV-548", "35-24HV-548 AV"), "AV")

    def test_infer_av_from_auto_keyword(self) -> None:
        self.assertEqual(infer_valve_type("35-24HV-548", "35-24HV-548 AUTO VLV"), "AV")

    def test_infer_av_from_fv_prefix(self) -> None:
        # No AV/AUTO in description; FV prefix → AV
        self.assertEqual(infer_valve_type("35-24FV-570", "35-24FV-570 PLPR FLOW VLV"), "AV")

    def test_infer_av_from_xv_prefix(self) -> None:
        self.assertEqual(infer_valve_type("35-24XV-669", "35-24XV-669 PLPR ISOL VLV"), "AV")

    def test_infer_drn_from_desc(self) -> None:
        self.assertEqual(infer_valve_type("35-24-137", "35-24-137 DRN"), "DRN")

    def test_infer_drn_nc_combined(self) -> None:
        # Rob review example 2: drain valve, normally closed
        result = infer_valve_type("35-24-137", "35-24-137 DRN NC")
        self.assertEqual(result, "DRN NC")

    def test_infer_nc_alone(self) -> None:
        self.assertEqual(infer_valve_type("35-24-030", "35-24-030 NC VLV"), "NC")

    def test_infer_hv_default_for_hv_tag(self) -> None:
        # HV prefix + "HAND VLV" description → no AV/DRN keywords → HV default
        self.assertEqual(infer_valve_type("35-24HV-623", "35-24HV-623 HYD VLV DN65"), "HV")

    def test_infer_av_immediate_beats_nc(self) -> None:
        # "AV NC" in description: AV is immediate → returns "AV", NC is ignored
        self.assertEqual(infer_valve_type("35-24FV-570", "35-24FV-570 AV NC VLV"), "AV")

    # --- format_valve_eqktx ---

    def test_format_rob_example_1_hv_tag_hand_vlv(self) -> None:
        # Rob example 1 — current description: "35-24HV-548 HAND VLV"
        # Rule-based gives HV (visual AV would need override); verify format is correct
        result = format_valve_eqktx("35-24HV-548", "35-24L005", "35-24HV-548 HV")
        self.assertTrue(result.startswith("HV "))
        self.assertIn("35-24-548", result)
        self.assertIn("35-24L005", result)

    def test_format_with_vision_av_type(self) -> None:
        # Vision cache classified HV-548 as AV despite its hand-valve tag prefix.
        result = format_valve_eqktx(
            "35-24HV-548", "35-24L005", "35-24HV-548 HV", valve_type_override="AV"
        )
        self.assertEqual(result, "HV 35-24-548 35-24L005 AV")

    def test_format_strips_nc_from_process_controlled(self) -> None:
        result = format_valve_eqktx(
            "35-24HV-548", "35-24L005", "35-24HV-548", valve_type_override="AV NC"
        )
        self.assertEqual(result, "HV 35-24-548 35-24L005 AV")

    def test_format_rob_example_2_drn_nc(self) -> None:
        # Rob example 2 — drain valve, normally closed
        result = format_valve_eqktx("35-24-137", "35-24L005", "35-24-137 DRN NC")
        self.assertEqual(result, "HV 35-24-137 35-24L005 DRN NC")

    def test_format_fv_auto_valve(self) -> None:
        result = format_valve_eqktx("35-24FV-570", "35-24L003", "35-24FV-570 FLOW VLV")
        self.assertEqual(result, "HV 35-24-570 35-24L003 AV")

    def test_format_plain_vlv_on_pipe(self) -> None:
        # Plain line tag described as valve on a pipe
        result = format_valve_eqktx("35-24-207", "35-24L001", "35-24-207 VLV ON 35-24-096")
        self.assertTrue(result.startswith("HV 35-24-207"))

    def test_format_max_length(self) -> None:
        result = format_valve_eqktx(
            "35-24HV-548", "35-24L005", "35-24HV-548 HV", valve_type_override="AV"
        )
        self.assertLessEqual(len(result), 40)

    def test_format_empty_parent_fn(self) -> None:
        # Empty parent_fn — still produces valid prefix + tag + type
        result = format_valve_eqktx("35-24HV-548", "", "35-24HV-548 DRN NC")
        self.assertEqual(result, "HV 35-24-548 DRN NC")

    # --- _valve_hint P-CVPOS layer ---

    def test_valve_hint_cvpos_sets_av_type(self) -> None:
        # P-CVPOS = control/automatic valve → SML type AV
        cache = {"35-24HV-548": {"layer": "P-CVPOS", "is_valve": True, "source": "cad_layer"}}
        hint = _valve_hint("35-24HV-548", cache=cache)
        self.assertEqual(hint["type"], "AV")
        self.assertTrue(hint["is_valve"])

    def test_valve_hint_cvpos_does_not_override_explicit_type(self) -> None:
        cache = {"35-24HV-548": {"layer": "P-CVPOS", "type": "AV", "is_valve": True}}
        hint = _valve_hint("35-24HV-548", cache=cache)
        self.assertEqual(hint["type"], "AV")

    def test_valve_hint_valvepos_layer_no_type(self) -> None:
        cache = {"35-24-137": {"layer": "P-VALVEPOS", "is_valve": True}}
        hint = _valve_hint("35-24-137", cache=cache)
        self.assertIsNone(hint["type"])
        self.assertTrue(hint["is_valve"])

    def test_valve_hint_missing_tag_returns_defaults(self) -> None:
        hint = _valve_hint("35-24HV-999", cache={})
        self.assertIsNone(hint["type"])
        self.assertFalse(hint["is_valve"])

    def test_build_rows_cvpos_layer_gives_av_type(self) -> None:
        """P-CVPOS layer in valve cache must produce AV type in EQKTX (control = automatic valve)."""
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 VLV"},
        ]
        cache = {"35-24HV-548": {"layer": "P-CVPOS", "is_valve": True, "source": "cad_layer"}}
        out = build_equipment_rows(rows, valve_cache=cache)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24HV-548", by_tag)
        eqktx = by_tag["35-24HV-548"]["EQKTX"]
        self.assertIn("AV", eqktx, msg=f"Expected AV type from P-CVPOS layer, got: {eqktx!r}")

    # --- is_valve_equipment ---

    def test_is_valve_equipment_by_tag(self) -> None:
        self.assertTrue(is_valve_equipment("35-24HV-548", "35-24HV-548 HV"))

    def test_is_valve_equipment_by_vlv_desc(self) -> None:
        # Plain line tag, description contains VLV → detected as valve
        self.assertTrue(is_valve_equipment("35-24-207", "35-24-207 VLV ON 35-24-096"))

    def test_is_valve_equipment_by_valve_word(self) -> None:
        self.assertTrue(is_valve_equipment("35-24-030", "35-24-030 CIRC VALVE 003-50"))

    def test_is_valve_equipment_pump_false(self) -> None:
        self.assertFalse(is_valve_equipment("35-24P519", "35-24P519 PMP"))

    def test_is_valve_equipment_line_false(self) -> None:
        self.assertFalse(is_valve_equipment("35-24-095", "LN 35-24-095 PRESS PLPR PP-200"))

    def test_is_valve_equipment_xs_switch_not_valve_despite_vlv_in_desc(self) -> None:
        # XS = switch; description incidentally says ISOL VLV (AI wrote XV tag in desc).
        # Must NOT be treated as a valve — tag prefix wins.
        self.assertFalse(is_valve_equipment("35-24XS-669", "35-24XV-669 PLPR ISOL VLV"))

    def test_is_valve_equipment_lc_instrument_not_valve(self) -> None:
        self.assertFalse(is_valve_equipment("35-24LC-576", "35-24LC-576 LVL CTRL VLV"))

    # --- integration: build_equipment_rows ---

    def test_build_rows_hv_tag_gets_hv_format(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24HV-548", by_tag)
        eqktx = by_tag["35-24HV-548"]["EQKTX"]
        self.assertTrue(eqktx.startswith("HV "), msg=f"Expected HV prefix, got: {eqktx!r}")
        self.assertIn("35-24-548", eqktx)
        self.assertIn("35-24L005", eqktx)

    def test_build_rows_fv_tag_gets_av_type(self) -> None:
        rows = [
            {"FUNCTION": "35-24L003", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24FV-570", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24FV-570 FLOW VLV"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24FV-570", by_tag)
        eqktx = by_tag["35-24FV-570"]["EQKTX"]
        self.assertIn("AV", eqktx, msg=f"Expected AV suffix, got: {eqktx!r}")

    def test_build_rows_plain_vlv_gets_hv_format(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-095 PRESS PLPR LINE"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-207", "DESCRIPTION": "35-24-207 VLV ON 35-24-096"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        # Line row still gets LN prefix
        self.assertTrue(by_tag["35-24-095"]["EQKTX"].startswith("LN "))
        # Valve sub-equipment gets HV format
        vlv_eqktx = by_tag["35-24-207"]["EQKTX"]
        self.assertTrue(vlv_eqktx.startswith("HV "), msg=f"Expected HV prefix, got: {vlv_eqktx!r}")

    def test_build_rows_drn_nc_from_desc(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-137", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-137 DRN NC"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24-137", by_tag)
        eqktx = by_tag["35-24-137"]["EQKTX"]
        self.assertIn("DRN", eqktx, msg=f"Expected DRN in eqktx, got: {eqktx!r}")
        self.assertIn("NC", eqktx, msg=f"Expected NC in eqktx, got: {eqktx!r}")

    def test_build_rows_non_valve_not_affected(self) -> None:
        # Pump and pure line equipment must not receive HV formatting
        rows = [
            {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24P519", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P519 PMP"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-095 PRESS PLPR LINE"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertFalse(by_tag["35-24P519"]["EQKTX"].startswith("HV "))
        self.assertFalse(by_tag["35-24-095"]["EQKTX"].startswith("HV "))

    def test_build_rows_reasoning_out_captures_valves_only(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV AV"},
            {"FUNCTION": "", "EQUIPMENT": "35-24P519", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P519 PMP"},
        ]
        reasoning: list[dict] = []
        build_equipment_rows(rows, reasoning_out=reasoning)
        self.assertEqual(len(reasoning), 1)
        r = reasoning[0]
        self.assertEqual(r["EQUNR"], "35-24HV-548")
        self.assertEqual(r["FUNCTION"], "35-24L005")
        self.assertIn("AV", r["TYPE"])
        self.assertEqual(r["SOURCE"], "AI_IMMEDIATE")
        self.assertIn("AV", r["AI_DESCRIPTION"])

    def test_build_rows_reasoning_vision_source(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV"},
        ]
        cache = {"35-24HV-548": {"type": "AV", "fn": "35-24L005", "is_valve": True, "source": "vision"}}
        reasoning: list[dict] = []
        build_equipment_rows(rows, valve_cache=cache, reasoning_out=reasoning)
        self.assertEqual(reasoning[0]["SOURCE"], "VISION")
        self.assertEqual(reasoning[0]["TYPE"], "AV")

    def test_valve_in_equipment_column_after_line_has_empty_hequi(self) -> None:
        """Issue 5: valves in EQUIPMENT column after a pipeline line tag must be standalone
        Equipment records (HEQUI=''), installed at the pipeline FLOC (TPLNR = pipeline FLOC).
        SAP PM standard: HEQUI is for physical components inside a machine, not valves on pipes.
        """
        rows = [
            {"FUNCTION": "35-24L002", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "BROKE THICKENER"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-150", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-150 WAA LN DN400"},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-626", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-626 HYD VLV"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-149", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-149 WAA LN DN300"},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-623", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-623 HYD VLV"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        # Pipeline lines are standalone FLOC children
        self.assertEqual(by_tag["35-24-150"]["HEQUI"], "")
        self.assertEqual(by_tag["35-24-149"]["HEQUI"], "")
        # Valves in EQUIPMENT column after a line: standalone Equipment, NOT sub-equipment of pipeline
        self.assertEqual(by_tag["35-24HV-626"]["HEQUI"], "",
                         "HV valve must not be sub-equipment of pipeline Equipment record")
        self.assertEqual(by_tag["35-24HV-623"]["HEQUI"], "",
                         "HV valve must not be sub-equipment of pipeline Equipment record")
        # Both valve and pipeline share the same FLOC TPLNR (installed at the pipeline FLOC)
        self.assertEqual(by_tag["35-24HV-626"]["TPLNR"], by_tag["35-24-150"]["TPLNR"],
                         "valve must be installed at the pipeline FLOC position")

    def test_valve_sub_equipment_in_csv_still_uses_hequi(self) -> None:
        """Valves explicitly listed in SUB-EQUIPMENT column retain HEQUI (AI placed them there).
        This preserves backward compat for CSVs where the AI used SUB-EQUIPMENT for valves.
        """
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "LINE"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-194", "DESCRIPTION": "HV DRN"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        # SUB-EQUIPMENT valves keep HEQUI = parent equipment
        self.assertEqual(by_tag["35-24-194"]["HEQUI"], "35-24-095")

    def test_build_rows_reasoning_columns_present(self) -> None:
        from dwg_reader.export_sap_equipment import REASONING_COLUMNS
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV"},
        ]
        reasoning: list[dict] = []
        build_equipment_rows(rows, reasoning_out=reasoning)
        self.assertEqual(set(reasoning[0].keys()), set(REASONING_COLUMNS))

    def test_write_valve_reasoning_csv_roundtrip(self) -> None:
        import csv
        import tempfile
        from dwg_reader.export_sap_equipment import write_valve_reasoning_csv, REASONING_COLUMNS
        data = [
            {k: f"val_{k}" for k in REASONING_COLUMNS},
        ]
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w") as tmp:
            p = Path(tmp.name)
        write_valve_reasoning_csv(p, data)
        with p.open() as f:
            reader = list(csv.DictReader(f))
        self.assertEqual(len(reader), 1)
        self.assertEqual(reader[0]["EQUNR"], "val_EQUNR")
        self.assertEqual(reader[0]["SOURCE"], "val_SOURCE")
        p.unlink(missing_ok=True)

    def test_motor_tag_reads_from_standard(self) -> None:
        """_motor_tag_for must use motor_from_equipment from ecosystem.standard."""
        from dwg_reader.dwg_ecosystem import Ecosystem
        valmet = Ecosystem(
            name="valmet",
            standard_id="valmet_ps21",
            standard={
                "motor_from_equipment": {
                    "mode": "strip_letter_append_dot_one",
                    "regex": r"^(\d{2}-\d{2})[A-Z]+(\d+)$",
                    "replace": r"\1-\2.1",
                }
            },
        )
        tissue = Ecosystem(
            name="gor",
            standard_id="tissue_ksdm160104",
            standard={
                "motor_from_equipment": {
                    "mode": "append_suffix",
                    "suffix": "-M1",
                }
            },
        )
        self.assertEqual(_motor_tag_for("35-24P518", ecosystem=valmet), "35-24-518.1")
        self.assertEqual(_motor_tag_for("124P-001", ecosystem=tissue), "124P-001-M1")

    def test_driven_reads_pump_from_standard(self) -> None:
        """_is_driven_equipment must use driven_patterns.pump from ecosystem.standard."""
        from dwg_reader.dwg_ecosystem import Ecosystem
        eco = Ecosystem(
            name="valmet",
            standard_id="valmet_ps21",
            standard={
                "driven_patterns": {
                    "pump": r"^\d{2}-\d{2}P\d+$",
                }
            },
        )
        self.assertTrue(_is_driven_equipment("35-24P518", eco))
        self.assertFalse(_is_driven_equipment("35-24L001", eco))

    def test_driven_reads_description_keywords_from_standard(self) -> None:
        """_is_driven_equipment must use description_keywords from ecosystem.standard."""
        from dwg_reader.dwg_ecosystem import Ecosystem
        eco = Ecosystem(
            name="valmet",
            standard_id="valmet_ps21",
            standard={
                "driven_patterns": {
                    "pump": r"^\d{2}-\d{2}P\d+$",
                    "description_keywords": {
                        "screen": ["SCRN", "SCREEN"],
                        "gearbox": ["GRBX", "GEARBOX"],
                    }
                }
            },
        )
        self.assertTrue(_is_driven_equipment("35-24L022", eco, desc="35-24L022 SCREEN"))
        self.assertFalse(_is_driven_equipment("35-24L022", eco, desc="35-24L022 PULPER"))


class PipelineVsVisionTests(unittest.TestCase):
    """Numeric pipe tags with LN/LINE in the AI text must not become HV from vision."""

    def test_095_stays_line_despite_vision_hv(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L001 PRESS PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24-095 PP LN DN200"},
        ]
        cache = {"35-24-095": {"is_valve": True, "type": "HV", "source": "vision", "fn": "35-24L001"}}
        out = build_equipment_rows(rows, valve_cache=cache)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertTrue(by_tag["35-24-095"]["EQKTX"].startswith("LN 35-24-095"))
        self.assertFalse(by_tag["35-24-095"]["EQKTX"].startswith("HV "))
        self.assertNotIn("DN", by_tag["35-24-095"]["EQKTX"])

    def test_096_line_not_drain_valve(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-096", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24-096 PRESS PLPR PP-900 LINE"},
        ]
        cache = {"35-24-096": {"is_valve": True, "type": "DRN NC", "source": "vision"}}
        out = build_equipment_rows(rows, valve_cache=cache)
        eqktx = next(r["EQKTX"] for r in out if r["EQUNR"] == "35-24-096")
        self.assertTrue(eqktx.startswith("LN 35-24-096"))
        self.assertNotIn("PP-900", eqktx)

    def test_189_overflow_line_not_valve(self) -> None:
        rows = [
            {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-189", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24-189 LN OVFL"},
        ]
        cache = {"35-24-189": {"is_valve": True, "type": "DRN NC", "source": "vision"}}
        out = build_equipment_rows(rows, valve_cache=cache)
        eqktx = next(r["EQKTX"] for r in out if r["EQUNR"] == "35-24-189")
        self.assertTrue(eqktx.startswith("LN 35-24-189"))

    def test_137_real_drain_valve_still_hv(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-137", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24-137 PP-250 CONVEYOR SPOOL"},
        ]
        cache = {
            "35-24-137": {
                "is_valve": True, "type": "DRN NC", "source": "vision", "fn": "35-24L005",
            }
        }
        out = build_equipment_rows(rows, valve_cache=cache)
        eqktx = next(r["EQKTX"] for r in out if r["EQUNR"] == "35-24-137")
        self.assertEqual(eqktx, "HV 35-24-137 35-24L005 DRN NC")


class MotorAndFunctionEquipmentTests(unittest.TestCase):
    def test_l008_motor_is_elec_with_mtr(self) -> None:
        rows = [
            {"FUNCTION": "35-24L008", "EQUIPMENT": "", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24L008 WINDER PLPR HP-33G2 800BDTPD"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-008.1", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24L008 WINDER PLPR HP-33G2"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24L008", by_tag)
        self.assertNotIn("HP-33G2", by_tag["35-24L008"]["EQKTX"])
        motor = by_tag["35-24-008.1"]
        self.assertEqual(motor["EQART"], "1101")
        self.assertEqual(motor["GEWRK"], "ELEC")
        self.assertIn("MTR", motor["EQKTX"])
        self.assertEqual(motor["HEQUI"], "35-24L008")

    def test_l001_emitted_as_equipment_with_clean_eqktx(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24L001 PRESS PLPR HP-50G2"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "",
             "DESCRIPTION": "35-24-095 PRESS PLPR PP-200 LINE"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("35-24L001", by_tag)
        self.assertEqual(by_tag["35-24L001"]["POSNR"], "0001")
        self.assertNotIn("HP-50G2", by_tag["35-24L001"]["EQKTX"])
        self.assertIn("35-24L001", by_tag["35-24L001"]["TPLNR"])


class CrossUnitEquipmentTests(unittest.TestCase):
    def test_cross_unit_pipeline_function_dropped(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "PRESS PLPR"},
            {"FUNCTION": "35-25-034", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-25-034 WAF LN DN600"},
            {"FUNCTION": "", "EQUIPMENT": "35-25-034", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-25-034 WAF LN DN600"},
            {"FUNCTION": "35-24-056", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "LN SUCT"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-056", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-056 LN SUCT"},
        ]
        out = build_equipment_rows(rows)
        tags = {r["EQUNR"] for r in out}
        self.assertNotIn("35-25-034", tags)
        self.assertIn("35-24-056", tags)
        self.assertIn("35-24L001", tags)


class GorFanMotorHequiTests(unittest.TestCase):
    def test_fan_motors_nested_under_parent(self) -> None:
        rows = [
            {"FUNCTION": "WU12", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "WU12 VENTIL UNIT"},
            {"FUNCTION": "", "EQUIPMENT": "168F-415-M1", "SUB-EQUIPMENT": "", "DESCRIPTION": "168F-415-M1 MTR"},
            {"FUNCTION": "", "EQUIPMENT": "168F-415-M2", "SUB-EQUIPMENT": "", "DESCRIPTION": "168F-415-M2 MTR"},
        ]
        ctx = {
            "ecosystem": "gor",
            "plant": "6001",
            "line_code": "TM01",
            "process_code": "WU",
            "sub_process": "WUC",
        }
        out = build_equipment_rows(rows, ctx=ctx)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertIn("168F-415", by_tag, "fan parent must be injected")
        self.assertEqual(by_tag["168F-415-M1"]["HEQUI"], "168F-415")
        self.assertEqual(by_tag["168F-415-M2"]["HEQUI"], "168F-415")
        self.assertEqual(by_tag["168F-415-M1"]["EQART"], "1101")
        self.assertEqual(by_tag["168F-415-M1"]["GEWRK"], "ELEC")
        self.assertEqual(by_tag["168F-415"]["EQART"], "801")


class GorObjectTypeTests(unittest.TestCase):
    def test_previously_uncategorized_gor_prefixes(self) -> None:
        cases = {
            "168TI2": "1202",
            "168TV": "202",
            "168FV1-416": "202",
            "168GSO1": "1210",
            "168GSC": "1210",
            "168F-415": "801",
            "168ST-1": "203",
            "168P-410-M1": "1101",
        }
        for tag, code in cases.items():
            got, _ = classify_equipment(tag, f"{tag} INST")
            self.assertEqual(got, code, msg=f"{tag} → {got} expected {code}")


if __name__ == "__main__":
    unittest.main()
