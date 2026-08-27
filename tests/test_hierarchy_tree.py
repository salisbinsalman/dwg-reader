#!/usr/bin/env python3
"""Colour-coded hierarchy tree export."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from dwg_reader.export_hierarchy_tree import (
    FILL_EQUIPMENT,
    FILL_FLOC,
    FILL_SUB,
    build_tree_rows,
    run_hierarchy_tree_export,
    write_hierarchy_html,
    write_hierarchy_xlsx,
)
from dwg_reader.io import write_csv_rows
from dwg_reader.models import HIERARCHY_COLUMNS


def _row(fn="", eq="", sub="", desc="", mask=""):
    return {
        "SUB-PROCESS": "BR1",
        "FUNCTION": fn,
        "EQUIPMENT": eq,
        "SUB-EQUIPMENT": sub,
        "MASK": mask,
        "DESCRIPTION": desc,
    }


class BuildTreeRowsTests(unittest.TestCase):
    def test_skeleton_plus_function_equipment_sub(self) -> None:
        rows = build_tree_rows(
            [
                _row(fn="35-24L009", desc="BROKE ROLL PLPR", mask="5001-PM03-BR-BR1-35-24L009"),
                _row(eq="35-24P519", desc="BROKE ROLL PLPR PMP"),
                _row(sub="35-24-519.1", desc="MTR"),
            ],
            ctx={"process_name": "BROKE SYSTEM", "sub_process": "BR1"},
        )
        kinds = [r["KIND"] for r in rows]
        self.assertEqual(kinds[:4], ["FLOC", "FLOC", "FLOC", "FLOC"])
        self.assertIn("FLOC", kinds)
        self.assertIn("EQUIPMENT", kinds)
        self.assertIn("SUB-EQUIPMENT", kinds)
        fn_row = next(r for r in rows if r["FUNCTION"] == "35-24L009" and not r["EQUIPMENT"])
        self.assertEqual(fn_row["KIND"], "FLOC")
        self.assertEqual(fn_row["DESCRIPTION"], "BROKE ROLL PLPR")
        eq_row = next(r for r in rows if r["EQUIPMENT"] == "35-24P519")
        self.assertEqual(eq_row["KIND"], "EQUIPMENT")
        self.assertEqual(eq_row["FUNCTION"], "35-24L009")
        sub_row = next(r for r in rows if r["SUB-EQUIPMENT"] == "35-24-519.1")
        self.assertEqual(sub_row["KIND"], "SUB-EQUIPMENT")
        plant = next(r for r in rows if r["MASK"] == "5001")
        self.assertEqual(plant["DESCRIPTION"], "SHOTTON MILL LTD")

    def test_process_and_subprocess_both_present(self) -> None:
        rows = build_tree_rows([], ctx={"process_name": "BROKE SYSTEM"})
        masks = [r["MASK"] for r in rows]
        self.assertIn("5001-PM03-BR", masks)
        self.assertIn("5001-PM03-BR-BR1", masks)


class WriteTreeTests(unittest.TestCase):
    def test_xlsx_fills_by_kind(self) -> None:
        from openpyxl import load_workbook

        rows = build_tree_rows(
            [
                _row(fn="35-24L009", desc="PLPR"),
                _row(eq="35-24-189", desc="OVFL"),
                _row(sub="35-24-194", desc="HV"),
            ]
        )
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "tree.xlsx"
            write_hierarchy_xlsx(path, rows)
            wb = load_workbook(path)
            ws = wb.active
            fills = []
            for r in range(2, ws.max_row + 1):
                fills.append(ws.cell(r, 1).fill.fgColor.rgb[-6:].upper())
            self.assertIn(FILL_FLOC, fills)
            self.assertIn(FILL_EQUIPMENT, fills)
            self.assertIn(FILL_SUB, fills)
            self.assertEqual(ws.cell(1, 1).value, "KIND")

    def test_html_contains_legend_and_rows(self) -> None:
        rows = build_tree_rows([_row(fn="35-24L009", desc="PLPR")])
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "tree.html"
            write_hierarchy_html(path, rows, title="Broke System")
            text = path.read_text(encoding="utf-8")
            self.assertIn("Broke System", text)
            self.assertIn("35-24L009", text)
            self.assertIn(FILL_FLOC, text)
            self.assertIn("FLOC", text)

    def test_run_export_writes_both_files(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            csv_path = out / "Broke System.hierarchy_orchestrator.csv"
            write_csv_rows(
                csv_path,
                [
                    _row(fn="35-24L009", desc="BROKE ROLL PLPR"),
                    _row(eq="35-24-189", desc="OVFL"),
                ],
                HIERARCHY_COLUMNS,
            )
            rc = run_hierarchy_tree_export(
                input_path=Path("inputs/Broke System.dwg"),
                out_dir=out,
                hierarchy_csv=csv_path,
            )
            self.assertEqual(rc, 0)
            self.assertTrue((out / "Broke System.hierarchy.xlsx").is_file())
            self.assertTrue((out / "Broke System.hierarchy.html").is_file())


if __name__ == "__main__":
    unittest.main()
