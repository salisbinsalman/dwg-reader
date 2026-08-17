#!/usr/bin/env python3
"""Build inputs/sml_abbreviations.json from the SML abbreviation workbook."""

from __future__ import annotations

import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "inputs/SML Naming Abbreviation Standard.xlsx"
OUT = ROOT / "inputs/sml_abbreviations.json"

# Broke System FLOC descriptions keep these full even when the workbook maps them.
PRESERVE_TOKENS = [
    "BROKE",
    "WINDER",
    "REEL",
    "SLABBING",
    "COUCH",
    "THICKENER",
    "SIZE",
    "PRESS",
    "ROLL",
    "AREA",
    "WHITE",
    "SHOWERS",
    "PIT",
    "START",
    "STOP",
    "SIGNAL",
    "RECIRC",
    "SYSTEM",
]

# Domain phrase compressions (longest first) applied before generic SML phrases.
DOMAIN_PHRASE_RULES = [
    {"pattern": "BROKE ROLL PULPER", "replacement": "BROKE ROLL PLPR"},
    {"pattern": "SIZE PRESS PULPER", "replacement": "SIZE PRESS PLPR"},
    {"pattern": "SLABBING PULPER", "replacement": "SLABBING PLPR"},
    {"pattern": "WINDER PULPER", "replacement": "WINDER PLPR"},
    {"pattern": "REEL PULPER", "replacement": "REEL PLPR"},
    {"pattern": "PRESS PULPER", "replacement": "PRESS PLPR"},
    {"pattern": "BROKE THICKENER", "replacement": "BROKE THICKENER"},
    {"pattern": "COUCH PIT", "replacement": "COUCH PIT"},
    {"pattern": "EMERGENCY STOP", "replacement": "E-STOP"},
    {"pattern": "PUSH BUTTON", "replacement": "PB"},
]


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name or "").strip()).upper()


def _normalize_abbr(abbr: str) -> str:
    return re.sub(r"\s+", " ", str(abbr or "").strip()).upper()


def load_workbook_rows(path: Path) -> tuple[dict[str, str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    words: dict[str, str] = {}
    phrases: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _normalize_name(row[0] if row else "")
        abbr = _normalize_abbr(row[1] if row and len(row) > 1 else "")
        if not name or not abbr or name == "NAN" or abbr == "NAN":
            continue
        if re.search(r"[ /]", name):
            phrases.append({"pattern": name, "replacement": abbr})
        else:
            words[name] = abbr
    phrases.sort(key=lambda r: len(r["pattern"]), reverse=True)
    return words, phrases


def build_payload(xlsx: Path = XLSX) -> dict:
    words, phrases = load_workbook_rows(xlsx)
    abbr_values = sorted(set(words.values()) | {r["replacement"] for r in phrases})
    return {
        "meta": {
            "source": str(xlsx.relative_to(ROOT)) if xlsx.is_relative_to(ROOT) else str(xlsx),
            "description": "SML naming abbreviations for PLTXT / EQKTX text normalization",
            "generated_by": "scripts/build_sml_abbreviations_json.py",
        },
        "abbreviations": dict(sorted(words.items())),
        "sml_phrase_rules": phrases,
        "domain_phrase_rules": DOMAIN_PHRASE_RULES,
        "preserve_tokens": PRESERVE_TOKENS,
        "abbreviation_values": abbr_values,
    }


def main() -> int:
    payload = build_payload(XLSX)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(
        f"Wrote {OUT} "
        f"({len(payload['abbreviations'])} words, "
        f"{len(payload['sml_phrase_rules'])} SML phrases, "
        f"{len(payload['domain_phrase_rules'])} domain phrases)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
