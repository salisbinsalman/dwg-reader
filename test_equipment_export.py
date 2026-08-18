#!/usr/bin/env python3
"""Unit tests for SAP Equipment export (no Bedrock)."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from dwg_floc_context import (
    format_line_eqktx,
    format_valve_eqktx,
    infer_valve_type,
    is_line_equipment_tag,
    is_valve_equipment,
    is_valve_tag,
    normalize_pltxt,
    strip_valve_prefix,
)
from dwg_object_type import classify_equipment, lookup
from export_sap_equipment import build_equipment_rows, write_equipment_workbook

ROOT = Path(__file__).resolve().parent
HIERARCHY_CSV = ROOT / "outputs/Broke System.hierarchy_orchestrator.csv"


class ObjectTypeClassifierTests(unittest.TestCase):
    """Tests for dwg_object_type.classify_equipment."""

    def _cls(self, tag, desc):
        return classify_equipment(tag, desc)

    # --- tag prefix rules ---
    def test_level_controller_tag(self):
        # LC = Level Controller → 1204 INSTRUMENT LEVEL (not generic 1200)
        code, wc = self._cls("35-24LC-576", "35-24LC-576 LVL CTRL")
        self.assertEqual(code, "1204")
        self.assertEqual(wc, "")

    def test_control_valve_hv_tag(self):
        # HV = Hand Valve → 201 VALVE HAND
        code, wc = self._cls("35-24HV-626", "35-24HV-626 DN300 HND VLV")
        self.assertEqual(code, "201")
        self.assertEqual(wc, "MECH")

    def test_flow_valve_fv_tag(self):
        code, wc = self._cls("35-24FV-570", "35-24FV-570 FLOW VLV")
        self.assertEqual(code, "202")
        self.assertEqual(wc, "MECH")

    def test_switch_xs_tag(self):
        code, wc = self._cls("35-24XS-588", "35-24XS-588 MCS SIG")
        self.assertEqual(code, "1108")

    # --- description keyword rules ---
    def test_motor_keyword(self):
        code, wc = self._cls("35-24-001.1", "35-24-001.1 PRESS PLPR ROTOR 1 MTR")
        self.assertEqual(code, "1101")
        self.assertEqual(wc, "ELEC")

    def test_pump_keyword(self):
        code, wc = self._cls("35-24P519", "35-24P519 PMP")
        self.assertEqual(code, "701")
        self.assertEqual(wc, "MECH")

    def test_pulper_keyword(self):
        code, wc = self._cls("35-24L009", "35-24L009 BROKE ROLL PLPR")
        self.assertEqual(code, "2005")
        self.assertEqual(wc, "MECH")

    def test_hand_valve_keyword(self):
        code, wc = self._cls("35-24-207", "35-24-207 CIRC VLV 003-50")
        self.assertEqual(code, "201")
        self.assertEqual(wc, "MECH")

    def test_pipe_line_keyword(self):
        code, wc = self._cls("35-24-149", "35-24-149 WAA DN300 INLET LINE")
        self.assertEqual(code, "2100")
        self.assertEqual(wc, "MECH")

    def test_thickener_keyword(self):
        code, wc = self._cls("35-24L002", "35-24L002 BROKE THICKENER")
        self.assertEqual(code, "2009")
        self.assertEqual(wc, "MECH")

    # --- lookup helper ---
    def test_lookup(self):
        name, wc = lookup("701")
        self.assertEqual(name, "PUMP CENTRIFUGAL")
        self.assertEqual(wc, "MECH")

    def test_fallback(self):
        code, wc = classify_equipment("35-24-999", "35-24-999 UNKNOWN THING")
        self.assertEqual(code, "9999")
        self.assertEqual(wc, "")

    def test_pump_single_letter_tag(self):
        # "35-24P519" — single-letter no-dash format → P → 701 without needing PMP in description
        code, wc = self._cls("35-24P519", "35-24P519 OUTLET")
        self.assertEqual(code, "701")
        self.assertEqual(wc, "MECH")

    def test_motor_suffix_tag(self):
        # ".N" suffix convention marks motor sub-equipment even with a thin description
        code, wc = self._cls("35-24-001.1", "35-24-001.1 ROTOR 1")
        self.assertEqual(code, "1101")
        self.assertEqual(wc, "ELEC")

    def test_plpr_context_disambiguation(self):
        # PLPR in a child description is context (parent system name), not the equipment type
        code, wc = classify_equipment("35-24-095", "35-24-095 PRESS PLPR OUTLET PP-200")
        self.assertEqual(code, "2100")   # pipe line, not pulper
        self.assertEqual(wc, "MECH")
        code, _ = classify_equipment("35-24-101", "35-24-101 PRESS PLPR SUCT PP-150")
        self.assertEqual(code, "2100")
        # Motor on a pulper still classifies correctly (MTR rule fires first)
        code, wc = classify_equipment("35-24-001.1", "35-24-001.1 PRESS PLPR ROTOR 1 MTR")
        self.assertEqual(code, "1101")
        self.assertEqual(wc, "ELEC")


class FormatLineEqktxTests(unittest.TestCase):
    def test_review_examples(self) -> None:
        self.assertEqual(
            format_line_eqktx(
                "35-24-095",
                normalize_pltxt("35-24-095 PRESS PLPR PP-200 LINE"),
            ),
            "LN 35-24-095 PRESS PLPR PP-200",
        )
        self.assertEqual(
            format_line_eqktx(
                "35-24-096",
                normalize_pltxt("35-24-096 PRESS PLPR PP-900 LINE"),
            ),
            "LN 35-24-096 PRESS PLPR PP-900",
        )

    def test_is_line_equipment_tag(self) -> None:
        self.assertTrue(is_line_equipment_tag("35-24-095"))
        self.assertFalse(is_line_equipment_tag("35-24-001.1"))
        self.assertFalse(is_line_equipment_tag("35-24LC-576"))
        self.assertFalse(is_line_equipment_tag("35-24HV-548"))
        self.assertFalse(is_line_equipment_tag("35-24P519"))

    def test_idempotent_when_already_prefixed(self) -> None:
        text = "LN 35-24-095 PRESS PLPR PP-200"
        self.assertEqual(format_line_eqktx("35-24-095", text), text)

    def test_overflow_line_with_embedded_ln(self) -> None:
        self.assertEqual(
            format_line_eqktx("35-24-189", normalize_pltxt("35-24-189 LN OVFL")),
            "LN 35-24-189 OVFL",
        )

    def test_skips_valve_sub_equipment(self) -> None:
        text = normalize_pltxt("35-24-207 VLV ON 35-24-096")
        self.assertEqual(
            format_line_eqktx("35-24-207", text, hequi="35-24-096"),
            text,
        )

    def test_skips_valve_description_without_hequi(self) -> None:
        text = normalize_pltxt("35-24-234 VLV ON 35-24-093")
        self.assertEqual(format_line_eqktx("35-24-234", text), text)

    def test_applies_without_trailing_line_word(self) -> None:
        self.assertEqual(
            format_line_eqktx("35-24-149", "35-24-149 WAA DN300 INLET"),
            "LN 35-24-149 WAA DN300 INLET",
        )

    def test_max_length_40(self) -> None:
        long_desc = "35-24-095 " + "PRESS PLPR PP-200 EXTRA WORDS " * 3
        out = format_line_eqktx("35-24-095", normalize_pltxt(long_desc))
        self.assertLessEqual(len(out), 40)
        self.assertTrue(out.startswith("LN 35-24-095"))


class EquipmentExportTests(unittest.TestCase):
    def test_build_rows_hequi_and_posnr(self) -> None:
        rows = [
            {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24L009 BROKE ROLL PLPR"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-189", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-189 LN OVFL"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-194", "DESCRIPTION": "35-24-194 HV DRN"},
            {"FUNCTION": "", "EQUIPMENT": "35-24P519", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P519 PMP"},
        ]
        out = build_equipment_rows(rows)
        self.assertEqual(len(out), 3)
        self.assertEqual(out[0]["EQUNR"], "35-24-189")
        self.assertEqual(out[0]["HEQUI"], "")
        self.assertEqual(out[0]["POSNR"], "0010")
        self.assertEqual(out[0]["TPLNR"], "5001-PM03-BR-BR1-35-24L009")
        self.assertEqual(out[0]["EQTYP"], "P")
        self.assertEqual(out[0]["EQART"], "9999")   # no keyword match → NOT CATEGORIZED
        self.assertEqual(out[0]["GEWRK"], "")
        self.assertEqual(out[0]["SWERK"], "5001")
        self.assertEqual(out[0]["ABCKZ"], "D")
        self.assertEqual(out[0]["INGRP"], "P01")
        self.assertLessEqual(len(out[0]["EQKTX"]), 40)
        # pump tag (35-24P519 PMP) → PUMP CENTRIFUGAL
        self.assertEqual(out[2]["EQART"], "701")
        self.assertEqual(out[2]["GEWRK"], "MECH")

        self.assertEqual(out[1]["EQUNR"], "35-24-194")
        self.assertEqual(out[1]["HEQUI"], "35-24-189")
        self.assertEqual(out[1]["POSNR"], "0010")

        self.assertEqual(out[2]["EQUNR"], "35-24P519")
        self.assertEqual(out[2]["HEQUI"], "")
        self.assertEqual(out[2]["POSNR"], "0020")

    def test_sub_equipment_posnr_resets_per_parent(self) -> None:
        """Review issue 3: subs reset POSNR; top-level continues separately."""
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "LINE A"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24LV2-576", "DESCRIPTION": "LVL VLV"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-096", "SUB-EQUIPMENT": "", "DESCRIPTION": "LINE B"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-207", "DESCRIPTION": "VLV 1"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-217", "DESCRIPTION": "VLV 2"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-1105", "DESCRIPTION": "VLV 3"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-093", "SUB-EQUIPMENT": "", "DESCRIPTION": "LINE C"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertEqual(by_tag["35-24-095"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24LV2-576"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24-096"]["POSNR"], "0020")
        self.assertEqual(by_tag["35-24-207"]["POSNR"], "0010")
        self.assertEqual(by_tag["35-24-217"]["POSNR"], "0020")
        self.assertEqual(by_tag["35-24-1105"]["POSNR"], "0030")
        self.assertEqual(by_tag["35-24-093"]["POSNR"], "0030")

    def test_line_eqktx_in_export(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-095 PRESS PLPR PP-200 LINE"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-096", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-096 PRESS PLPR PP-900 LINE"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-207", "DESCRIPTION": "35-24-207 VLV ON 35-24-096"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertEqual(by_tag["35-24-095"]["EQKTX"], "LN 35-24-095 PRESS PLPR PP-200")
        self.assertEqual(by_tag["35-24-096"]["EQKTX"], "LN 35-24-096 PRESS PLPR PP-900")
        self.assertFalse(by_tag["35-24-207"]["EQKTX"].startswith("LN "))

    def test_limit_functions(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-001.1", "SUB-EQUIPMENT": "", "DESCRIPTION": "MTR"},
            {"FUNCTION": "35-24L002", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-002.1", "SUB-EQUIPMENT": "", "DESCRIPTION": "MTR"},
        ]
        out = build_equipment_rows(rows, limit_functions=1)
        self.assertEqual(len(out), 1)
        self.assertEqual(out[0]["EQUNR"], "35-24-001.1")
        self.assertIn("35-24L001", out[0]["TPLNR"])

    def test_write_workbook(self) -> None:
        template = ROOT / "docs/examples/SML-Equipment Template RW.xlsx"
        self.assertTrue(template.exists())
        rows = build_equipment_rows(
            [
                {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
                {"FUNCTION": "", "EQUIPMENT": "35-24-189", "SUB-EQUIPMENT": "", "DESCRIPTION": "OVFL"},
            ]
        )
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "eq.xlsx"
            write_equipment_workbook(template, out, rows)
            self.assertTrue(out.exists())
            from openpyxl import load_workbook

            wb = load_workbook(out)
            self.assertIn("Equipment", wb.sheetnames)
            ws = wb["Equipment"]
            self.assertEqual(ws.cell(7, 2).value, "5001-PM03-BR-BR1-35-24L009")
            self.assertEqual(ws.cell(7, 3).value, "35-24-189")
            self.assertEqual(ws.cell(7, 7).value, "P")
            self.assertEqual(ws.cell(7, 8).value, "9999")  # "35-24-189 OVFL" → no match


class RealHierarchyLineEqktxTests(unittest.TestCase):
    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_all_numeric_line_equipment_prefixed_with_ln(self) -> None:
        import re

        from export_sap_equipment import read_hierarchy_csv

        hierarchy = read_hierarchy_csv(HIERARCHY_CSV)
        out = build_equipment_rows(hierarchy, limit_functions=5)
        line_pat = re.compile(r"^35-24-\d+$")
        valve_markers = ("VLV", "VALVE", " HV", " FV", " LV", " ON 35-24-")

        missing_ln = []
        has_line_word = []
        for row in out:
            tag = row["EQUNR"]
            eqktx = row["EQKTX"]
            if not line_pat.match(tag):
                continue
            if eqktx.startswith("HV "):
                continue
            if any(m in eqktx for m in valve_markers):
                continue
            if re.search(r"\b(NC|DRN|CHK|FLS|SMP|PRV|SV)\b", eqktx):
                continue
            if not eqktx.startswith("LN "):
                missing_ln.append((tag, eqktx))
            if eqktx.endswith(" LINE") or eqktx.endswith(" LN"):
                has_line_word.append((tag, eqktx))

        self.assertEqual(missing_ln, [], msg=f"missing LN prefix: {missing_ln[:10]}")
        self.assertEqual(has_line_word, [], msg=f"trailing LINE/LN: {has_line_word[:10]}")

    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_l001_review_line_examples(self) -> None:
        out = build_equipment_rows(
            __import__("export_sap_equipment").read_hierarchy_csv(HIERARCHY_CSV),
            limit_functions=1,
        )
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertTrue(by_tag["35-24-095"]["EQKTX"].startswith("LN 35-24-095"))
        self.assertTrue(by_tag["35-24-096"]["EQKTX"].startswith("LN 35-24-096"))
        self.assertFalse(by_tag["35-24-207"]["EQKTX"].startswith("LN "))


class ValveFormattingTests(unittest.TestCase):
    """Tests for Issue 5 — valve HV-scheme equipment text."""

    # --- is_valve_tag ---

    def test_is_valve_tag_hv(self) -> None:
        self.assertTrue(is_valve_tag("35-24HV-548"))

    def test_is_valve_tag_fv(self) -> None:
        self.assertTrue(is_valve_tag("35-24FV-570"))

    def test_is_valve_tag_lv(self) -> None:
        self.assertTrue(is_valve_tag("35-24LV-622"))

    def test_is_valve_tag_lv_with_digit(self) -> None:
        # 35-24LV2-576 — LV2 is the prefix (second level valve on same line)
        self.assertTrue(is_valve_tag("35-24LV2-576"))

    def test_is_valve_tag_xv(self) -> None:
        self.assertTrue(is_valve_tag("35-24XV-669"))

    def test_is_valve_tag_kv(self) -> None:
        self.assertTrue(is_valve_tag("35-24KV-573"))

    def test_is_valve_tag_plain_line_false(self) -> None:
        self.assertFalse(is_valve_tag("35-24-207"))

    def test_is_valve_tag_pump_false(self) -> None:
        self.assertFalse(is_valve_tag("35-24P519"))

    def test_is_valve_tag_switch_false(self) -> None:
        # XS = switch, not a valve prefix
        self.assertFalse(is_valve_tag("35-24XS-588"))

    def test_is_valve_tag_level_controller_false(self) -> None:
        self.assertFalse(is_valve_tag("35-24LC-576"))

    # --- strip_valve_prefix ---

    def test_strip_hv(self) -> None:
        # Rob review example 1
        self.assertEqual(strip_valve_prefix("35-24HV-548"), "35-24-548")

    def test_strip_fv(self) -> None:
        self.assertEqual(strip_valve_prefix("35-24FV-570"), "35-24-570")

    def test_strip_lv_with_digit(self) -> None:
        # Position digit preserved so LV2-576 and a hypothetical LV1-576 stay distinct.
        self.assertEqual(strip_valve_prefix("35-24LV2-576"), "35-24-2-576")

    def test_strip_lv1(self) -> None:
        self.assertEqual(strip_valve_prefix("35-24LV1-560"), "35-24-1-560")

    def test_strip_lv2_distinct_from_lv1(self) -> None:
        # LV1-560 and LV2-560 must not both collapse to 35-24-560.
        self.assertNotEqual(
            strip_valve_prefix("35-24LV1-560"),
            strip_valve_prefix("35-24LV2-560"),
        )

    def test_strip_plain_tag_unchanged(self) -> None:
        # Plain line tag — no embedded letters to strip
        self.assertEqual(strip_valve_prefix("35-24-137"), "35-24-137")

    def test_strip_plain_207_unchanged(self) -> None:
        self.assertEqual(strip_valve_prefix("35-24-207"), "35-24-207")

    # --- infer_valve_type ---

    def test_infer_av_from_desc_keyword(self) -> None:
        self.assertEqual(infer_valve_type("35-24HV-548", "35-24HV-548 AV"), "AV")

    def test_infer_av_from_auto_keyword(self) -> None:
        self.assertEqual(infer_valve_type("35-24HV-548", "35-24HV-548 AUTO VLV"), "AV")

    def test_infer_av_from_fv_prefix(self) -> None:
        # No AV/AUTO in description; FV prefix → AV
        self.assertEqual(infer_valve_type("35-24FV-570", "35-24FV-570 PLPR FLOW VLV"), "AV")

    def test_infer_av_from_xv_prefix(self) -> None:
        self.assertEqual(infer_valve_type("35-24XV-669", "35-24XV-669 PLPR ISOL VLV"), "AV")

    def test_infer_drn_from_desc(self) -> None:
        self.assertEqual(infer_valve_type("35-24-137", "35-24-137 DRN"), "DRN")

    def test_infer_drn_nc_combined(self) -> None:
        # Rob review example 2: drain valve, normally closed
        result = infer_valve_type("35-24-137", "35-24-137 DRN NC")
        self.assertEqual(result, "DRN NC")

    def test_infer_nc_alone(self) -> None:
        self.assertEqual(infer_valve_type("35-24-030", "35-24-030 NC VLV"), "NC")

    def test_infer_hv_default_for_hv_tag(self) -> None:
        # HV prefix + "HAND VLV" description → no AV/DRN keywords → HV default
        self.assertEqual(infer_valve_type("35-24HV-623", "35-24HV-623 HYD VLV DN65"), "HV")

    def test_infer_av_immediate_beats_nc(self) -> None:
        # "AV NC" in description: AV is immediate → returns "AV", NC is ignored
        self.assertEqual(infer_valve_type("35-24FV-570", "35-24FV-570 AV NC VLV"), "AV")

    # --- format_valve_eqktx ---

    def test_format_rob_example_1_hv_tag_hand_vlv(self) -> None:
        # Rob example 1 — current description: "35-24HV-548 HAND VLV"
        # Rule-based gives HV (visual AV would need override); verify format is correct
        result = format_valve_eqktx("35-24HV-548", "35-24L005", "35-24HV-548 HV")
        self.assertTrue(result.startswith("HV "))
        self.assertIn("35-24-548", result)
        self.assertIn("35-24L005", result)

    def test_format_with_vision_av_type(self) -> None:
        # Vision cache classified HV-548 as AV despite its hand-valve tag prefix.
        result = format_valve_eqktx(
            "35-24HV-548", "35-24L005", "35-24HV-548 HV", valve_type_override="AV"
        )
        self.assertEqual(result, "HV 35-24-548 35-24L005 AV")

    def test_format_strips_nc_from_process_controlled(self) -> None:
        result = format_valve_eqktx(
            "35-24HV-548", "35-24L005", "35-24HV-548", valve_type_override="AV NC"
        )
        self.assertEqual(result, "HV 35-24-548 35-24L005 AV")

    def test_format_rob_example_2_drn_nc(self) -> None:
        # Rob example 2 — drain valve, normally closed
        result = format_valve_eqktx("35-24-137", "35-24L005", "35-24-137 DRN NC")
        self.assertEqual(result, "HV 35-24-137 35-24L005 DRN NC")

    def test_format_fv_auto_valve(self) -> None:
        result = format_valve_eqktx("35-24FV-570", "35-24L003", "35-24FV-570 FLOW VLV")
        self.assertEqual(result, "HV 35-24-570 35-24L003 AV")

    def test_format_plain_vlv_on_pipe(self) -> None:
        # Plain line tag described as valve on a pipe
        result = format_valve_eqktx("35-24-207", "35-24L001", "35-24-207 VLV ON 35-24-096")
        self.assertTrue(result.startswith("HV 35-24-207"))

    def test_format_max_length(self) -> None:
        result = format_valve_eqktx(
            "35-24HV-548", "35-24L005", "35-24HV-548 HV", valve_type_override="AV"
        )
        self.assertLessEqual(len(result), 40)

    def test_format_empty_parent_fn(self) -> None:
        # Empty parent_fn — still produces valid prefix + tag + type
        result = format_valve_eqktx("35-24HV-548", "", "35-24HV-548 DRN NC")
        self.assertEqual(result, "HV 35-24-548 DRN NC")

    # --- is_valve_equipment ---

    def test_is_valve_equipment_by_tag(self) -> None:
        self.assertTrue(is_valve_equipment("35-24HV-548", "35-24HV-548 HV"))

    def test_is_valve_equipment_by_vlv_desc(self) -> None:
        # Plain line tag, description contains VLV → detected as valve
        self.assertTrue(is_valve_equipment("35-24-207", "35-24-207 VLV ON 35-24-096"))

    def test_is_valve_equipment_by_valve_word(self) -> None:
        self.assertTrue(is_valve_equipment("35-24-030", "35-24-030 CIRC VALVE 003-50"))

    def test_is_valve_equipment_pump_false(self) -> None:
        self.assertFalse(is_valve_equipment("35-24P519", "35-24P519 PMP"))

    def test_is_valve_equipment_line_false(self) -> None:
        self.assertFalse(is_valve_equipment("35-24-095", "LN 35-24-095 PRESS PLPR PP-200"))

    def test_is_valve_equipment_xs_switch_not_valve_despite_vlv_in_desc(self) -> None:
        # XS = switch; description incidentally says ISOL VLV (AI wrote XV tag in desc).
        # Must NOT be treated as a valve — tag prefix wins.
        self.assertFalse(is_valve_equipment("35-24XS-669", "35-24XV-669 PLPR ISOL VLV"))

    def test_is_valve_equipment_lc_instrument_not_valve(self) -> None:
        self.assertFalse(is_valve_equipment("35-24LC-576", "35-24LC-576 LVL CTRL VLV"))

    # --- integration: build_equipment_rows ---

    def test_build_rows_hv_tag_gets_hv_format(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV"},
        ]
        out = build_equipment_rows(rows)
        self.assertEqual(len(out), 1)
        eqktx = out[0]["EQKTX"]
        self.assertTrue(eqktx.startswith("HV "), msg=f"Expected HV prefix, got: {eqktx!r}")
        self.assertIn("35-24-548", eqktx)
        self.assertIn("35-24L005", eqktx)

    def test_build_rows_fv_tag_gets_av_type(self) -> None:
        rows = [
            {"FUNCTION": "35-24L003", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24FV-570", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24FV-570 FLOW VLV"},
        ]
        out = build_equipment_rows(rows)
        eqktx = out[0]["EQKTX"]
        self.assertIn("AV", eqktx, msg=f"Expected AV suffix, got: {eqktx!r}")

    def test_build_rows_plain_vlv_gets_hv_format(self) -> None:
        rows = [
            {"FUNCTION": "35-24L001", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-095 PRESS PLPR LINE"},
            {"FUNCTION": "", "EQUIPMENT": "", "SUB-EQUIPMENT": "35-24-207", "DESCRIPTION": "35-24-207 VLV ON 35-24-096"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        # Line row still gets LN prefix
        self.assertTrue(by_tag["35-24-095"]["EQKTX"].startswith("LN "))
        # Valve sub-equipment gets HV format
        vlv_eqktx = by_tag["35-24-207"]["EQKTX"]
        self.assertTrue(vlv_eqktx.startswith("HV "), msg=f"Expected HV prefix, got: {vlv_eqktx!r}")

    def test_build_rows_drn_nc_from_desc(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24-137", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-137 DRN NC"},
        ]
        out = build_equipment_rows(rows)
        eqktx = out[0]["EQKTX"]
        self.assertIn("DRN", eqktx, msg=f"Expected DRN in eqktx, got: {eqktx!r}")
        self.assertIn("NC", eqktx, msg=f"Expected NC in eqktx, got: {eqktx!r}")

    def test_build_rows_non_valve_not_affected(self) -> None:
        # Pump and pure line equipment must not receive HV formatting
        rows = [
            {"FUNCTION": "35-24L009", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24P519", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P519 PMP"},
            {"FUNCTION": "", "EQUIPMENT": "35-24-095", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24-095 PRESS PLPR LINE"},
        ]
        out = build_equipment_rows(rows)
        by_tag = {r["EQUNR"]: r for r in out}
        self.assertFalse(by_tag["35-24P519"]["EQKTX"].startswith("HV "))
        self.assertFalse(by_tag["35-24-095"]["EQKTX"].startswith("HV "))

    def test_build_rows_reasoning_out_captures_valves_only(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV AV"},
            {"FUNCTION": "", "EQUIPMENT": "35-24P519", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24P519 PMP"},
        ]
        reasoning: list[dict] = []
        build_equipment_rows(rows, reasoning_out=reasoning)
        self.assertEqual(len(reasoning), 1)
        r = reasoning[0]
        self.assertEqual(r["EQUNR"], "35-24HV-548")
        self.assertEqual(r["FUNCTION"], "35-24L005")
        self.assertIn("AV", r["TYPE"])
        self.assertEqual(r["SOURCE"], "AI_IMMEDIATE")
        self.assertIn("AV", r["AI_DESCRIPTION"])

    def test_build_rows_reasoning_vision_source(self) -> None:
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV"},
        ]
        cache = {"35-24HV-548": {"type": "AV", "fn": "35-24L005", "is_valve": True, "source": "vision"}}
        reasoning: list[dict] = []
        build_equipment_rows(rows, valve_cache=cache, reasoning_out=reasoning)
        self.assertEqual(reasoning[0]["SOURCE"], "VISION")
        self.assertEqual(reasoning[0]["TYPE"], "AV")

    def test_build_rows_reasoning_columns_present(self) -> None:
        from export_sap_equipment import REASONING_COLUMNS
        rows = [
            {"FUNCTION": "35-24L005", "EQUIPMENT": "", "SUB-EQUIPMENT": "", "DESCRIPTION": ""},
            {"FUNCTION": "", "EQUIPMENT": "35-24HV-548", "SUB-EQUIPMENT": "", "DESCRIPTION": "35-24HV-548 HAND VLV"},
        ]
        reasoning: list[dict] = []
        build_equipment_rows(rows, reasoning_out=reasoning)
        self.assertEqual(set(reasoning[0].keys()), set(REASONING_COLUMNS))

    def test_write_valve_reasoning_csv_roundtrip(self) -> None:
        import csv
        import tempfile
        from export_sap_equipment import write_valve_reasoning_csv, REASONING_COLUMNS
        data = [
            {k: f"val_{k}" for k in REASONING_COLUMNS},
        ]
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w") as tmp:
            p = Path(tmp.name)
        write_valve_reasoning_csv(p, data)
        with p.open() as f:
            reader = list(csv.DictReader(f))
        self.assertEqual(len(reader), 1)
        self.assertEqual(reader[0]["EQUNR"], "val_EQUNR")
        self.assertEqual(reader[0]["SOURCE"], "val_SOURCE")
        p.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
