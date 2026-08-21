#!/usr/bin/env python3
"""
Export hierarchy FUNCTION rows into the SAP Functional Location load workbook.

Reads hierarchy CSV (orchestrator) and writes a workbook shaped
like docs/examples/final-output-template.xlsx.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dwg_floc_context import (
    build_tplnr,
    floc_paths_for_function,
    format_line_eqktx,
    is_line_equipment_tag,
    load_floc_context_for_input,
    merge_floc_context,
    normalize_pltxt,
)
from dwg_floc_standards import lookup_line, lookup_process, lookup_sub_process
from dwg_object_type import classify_equipment
from dwg_pure_dump import find_json, json_path, safe_name, write_json

TEMPLATE_DEFAULT = Path("docs/examples/final-output-template.xlsx")

# Strips trailing model/spec codes like "HP-33G2", "GT-2540", "PROFS-250HC", "141BDTPD"
# from FUNCTION-level PLTXT descriptions.  Plain integers ("CONVEYOR 3") are left intact.
_TRAILING_SPEC_RE = re.compile(
    r"\s+(?:[A-Z]{1,8}-\d+[A-Z0-9]*|\d+[A-Z]{2,}[A-Z0-9]*)$"
)


def _strip_trailing_spec(text: str) -> str:
    return _TRAILING_SPEC_RE.sub("", text).strip()
SHEET_NAME = "Functional Location"

# Column order matches template row 0 (Technical Area codes).
SAP_COLUMNS = [
    "TPLNR",
    "TPLMA",
    "POSNR",
    "TPLKZ",
    "PLTXT",
    "EQUART",
    "ABCKZ",
    "STORT",
    "HERST",
    "TYPBZ",
    "SERGE",
    "FLTYP",
    "EQART",
    "SWERK",
    "KOSTL",
    "IWERK",
    "INGRP",
    "GEWRK",
]

# Codes Excel often coerces to numbers (losing leading zeros / type). Force text.
_TEXT_FORMAT_COLUMNS = frozenset({"POSNR", "FLTYP", "EQART", "EQUART", "ABCKZ"})


def _norm(value: object) -> str:
    s = str(value or "").strip()
    return "" if not s or s.lower() == "nan" else s


def read_hierarchy_csv(path: Path) -> List[Dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as f:
        return [{k: _norm(v) for k, v in row.items()} for row in csv.DictReader(f)]


def collect_functions(
    rows: Iterable[Dict[str, str]],
    *,
    gt_descriptions: Optional[Dict[str, str]] = None,
) -> List[Tuple[str, str, str]]:
    """
    Return ordered unique FUNCTIONs as (tag, mask, description).
    Prefers first non-empty MASK/DESCRIPTION seen for that function.
    """
    out: List[Tuple[str, str, str]] = []
    seen = set()
    for row in rows:
        fn = _norm(row.get("FUNCTION")).upper().replace(" ", "")
        if not fn or fn in seen:
            continue
        seen.add(fn)
        mask = _norm(row.get("MASK"))
        desc = _strip_trailing_spec(normalize_pltxt(_norm(row.get("DESCRIPTION"))))
        if not desc and gt_descriptions:
            desc = gt_descriptions.get(fn, "")
        out.append((fn, mask, desc))
    return out


def load_gt_function_descriptions(gt_path: Path) -> Dict[str, str]:
    import pandas as pd

    gt = pd.read_excel(gt_path)
    desc_col = "DESCRIPTION (max 40)" if "DESCRIPTION (max 40)" in gt.columns else "DESCRIPTION"
    out: Dict[str, str] = {}
    for _, raw in gt.iterrows():
        fn = _norm(raw.get("FUNCTION")).upper().replace(" ", "")
        if not fn or fn in out:
            continue
        out[fn] = normalize_pltxt(_norm(raw.get(desc_col)))
    return out



def build_floc_rows(
    functions: Sequence[Tuple[str, str, str]],
    ctx: Optional[Dict[str, str]] = None,
) -> List[Dict[str, str]]:
    """Build plant→line→process→subprocess→function FL rows."""
    c = merge_floc_context(ctx)
    # raw_ctx holds only what the caller explicitly passed — used below so that
    # DEFAULT_FLOC_CONTEXT values don't shadow FLOC standards lookups.
    raw_ctx: Dict[str, str] = dict(ctx or {})
    plant = c["plant"]
    line = build_tplnr(plant, c["line_code"])
    process = build_tplnr(plant, c["line_code"], c["process_code"])
    subprocess = build_tplnr(plant, c["line_code"], c["process_code"], c["sub_process"])

    # Resolve display names: explicit raw ctx wins, then FLOC standards lookup, then code.
    # We check raw_ctx (not merged c) so DEFAULT_FLOC_CONTEXT values don't suppress
    # the standards lookup when a caller passes only line_code / process_code.
    line_name = (
        raw_ctx.get("line_name")
        or lookup_line(c["line_code"])
        or c["line_code"]
    )
    process_name = (
        raw_ctx.get("process_name")
        or lookup_process(c["process_code"])
        or c["process_code"]
    )
    sub_process_name = (
        raw_ctx.get("sub_process_name")
        or lookup_sub_process(c["sub_process"])
        or process_name
    )

    def blank_row(**kwargs: str) -> Dict[str, str]:
        row = {k: "" for k in SAP_COLUMNS}
        row["TPLKZ"] = c["structure_indicator"]
        row["FLTYP"] = c["fl_category"]
        row["SWERK"] = c["maintenance_plant"]
        row["ABCKZ"] = "D"
        row.update(kwargs)
        return row

    rows: List[Dict[str, str]] = []
    # Plant
    rows.append(
        blank_row(
            TPLNR=plant,
            PLTXT=normalize_pltxt(c.get("site_name") or "SHOTTON MILL LTD"),
            SWERK=c["maintenance_plant"],
        )
    )
    # Line
    rows.append(
        blank_row(
            TPLNR=line,
            TPLMA=plant,
            POSNR="0010",
            PLTXT=normalize_pltxt(line_name),
            EQART=c.get("fl_type_line") or "0100",
            IWERK=c.get("planning_plant") or plant,
            INGRP=c.get("planning_group", "P01"),
        )
    )
    # Process
    rows.append(
        blank_row(
            TPLNR=process,
            TPLMA=line,
            POSNR="0010",
            PLTXT=normalize_pltxt(process_name),
            IWERK=c.get("planning_plant") or plant,
            INGRP=c.get("planning_group", "P01"),
        )
    )
    # Sub-process
    rows.append(
        blank_row(
            TPLNR=subprocess,
            TPLMA=process,
            POSNR="0010",
            PLTXT=normalize_pltxt(sub_process_name),
            IWERK=c.get("planning_plant") or plant,
            INGRP=c.get("planning_group", "P01"),
        )
    )

    for i, (tag, mask, desc) in enumerate(functions):
        paths = floc_paths_for_function(tag, c)
        tplnr = mask if mask.startswith(subprocess) else paths["function"]
        if len(tplnr) > 30:
            tplnr = tplnr[:30]
        # Always normalize; then apply LN prefix for pipe-line FUNCTION tags
        # (same rule as Equipment EQKTX — Rob/SML feedback).
        pltxt = desc or tag
        if pltxt and not pltxt.startswith(tag):
            pltxt = normalize_pltxt(f"{tag} {pltxt}")
        else:
            pltxt = normalize_pltxt(pltxt)
        if is_line_equipment_tag(tag):
            pltxt = format_line_eqktx(tag, pltxt)
        _eqart, gewrk = classify_equipment(tag, pltxt)
        rows.append(
            blank_row(
                TPLNR=tplnr,
                TPLMA=subprocess,
                POSNR=f"{(i + 1) * 10:04d}",
                PLTXT=pltxt[:40],
                EQART=_eqart if _eqart != "9999" else "",
                GEWRK=gewrk,
                IWERK=c.get("planning_plant") or plant,
                INGRP=c.get("planning_group", "P01"),
            )
        )
    return rows


def write_floc_workbook(
    template_path: Path,
    out_path: Path,
    floc_rows: List[Dict[str, str]],
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # Strip template comments (Turkish column metadata) — openpyxl restructures
    # them to a different internal XML path, which breaks some xlsx readers.
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None:
                cell.comment = None

    # Template data starts at Excel row 8 (1-based) = index 7 in examples.
    # Clear any previous data rows below the header block (rows 1-7).
    if ws.max_row > 7:
        ws.delete_rows(8, ws.max_row - 7)

    # Columns B..S map to SAP_COLUMNS (A is ID / blank in examples).
    for r_i, row in enumerate(floc_rows, start=8):
        ws.cell(r_i, 1, value=None)  # ID column unused in examples
        for c_i, key in enumerate(SAP_COLUMNS, start=2):
            val = row.get(key) or None
            cell = ws.cell(r_i, c_i, value=val)
            if key in _TEXT_FORMAT_COLUMNS and val is not None:
                cell.number_format = "@"
                # Re-set as string so Excel/openpyxl keep leading zeros (e.g. 0100).
                cell.value = str(val)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def evaluate_against_gt(
    floc_rows: List[Dict[str, str]],
    gt_path: Path,
) -> Dict[str, Any]:
    """Score FUNCTION-level TPLNR/PLTXT vs GT MASK + DESCRIPTION."""
    import pandas as pd

    gt = pd.read_excel(gt_path)
    mask_col = "MASK (max 30)" if "MASK (max 30)" in gt.columns else "MASK"
    desc_col = "DESCRIPTION (max 40)" if "DESCRIPTION (max 40)" in gt.columns else "DESCRIPTION"

    gt_fn: Dict[str, Dict[str, str]] = {}
    for _, raw in gt.iterrows():
        fn = _norm(raw.get("FUNCTION")).upper().replace(" ", "")
        if not fn:
            continue
        if fn in gt_fn:
            continue
        gt_fn[fn] = {
            "mask": _norm(raw.get(mask_col)).upper().replace(" ", ""),
            "description": normalize_pltxt(_norm(raw.get(desc_col))),
        }

    pred_fn = {
        re.sub(r"^.*?(35-\d{2}[A-Z0-9./-]+)$", r"\1", r["TPLNR"]): r
        for r in floc_rows
        if re.search(r"35-\d{2}", r.get("TPLNR") or "")
    }
    # Better: extract tag after BR1-
    pred_by_tag: Dict[str, Dict[str, str]] = {}
    for r in floc_rows:
        tplnr = r.get("TPLNR") or ""
        m = re.search(r"(35-\d{2}[A-Z0-9./-]+)$", tplnr)
        if m:
            pred_by_tag[m.group(1)] = r

    mask_hit = mask_miss = desc_exact = desc_total = 0
    details = []
    for tag, gold in gt_fn.items():
        if tag not in pred_by_tag:
            continue
        pred = pred_by_tag[tag]
        desc_total += 1
        mask_ok = pred.get("TPLNR", "").upper() == gold["mask"]
        if gold["mask"]:
            if mask_ok:
                mask_hit += 1
            else:
                mask_miss += 1
        desc_ok = pred.get("PLTXT", "") == gold["description"]
        if desc_ok:
            desc_exact += 1
        details.append(
            {
                "function": tag,
                "tplnr_pred": pred.get("TPLNR"),
                "tplnr_gt": gold["mask"],
                "mask_match": mask_ok,
                "pltxt_pred": pred.get("PLTXT"),
                "pltxt_gt": gold["description"],
                "pltxt_exact": desc_ok,
            }
        )

    return {
        "functions_compared": len(details),
        "mask_hit": mask_hit,
        "mask_miss": mask_miss,
        "mask_accuracy": (mask_hit / (mask_hit + mask_miss)) if (mask_hit + mask_miss) else 0.0,
        "description_exact": desc_exact,
        "description_total": desc_total,
        "description_accuracy": (desc_exact / desc_total) if desc_total else 0.0,
        "details": details,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Export SAP Functional Location workbook")
    parser.add_argument(
        "--hierarchy-csv",
        default="",
        help="Hierarchy CSV (default: outputs/<stem>.hierarchy_orchestrator.csv)",
    )
    parser.add_argument("--input", default="inputs/Broke System.dwg", help="Used for output stem")
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--template", default=str(TEMPLATE_DEFAULT))
    parser.add_argument("--gt", default="resources/gt_hierarchy_broke_system.xlsx")
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional max FUNCTIONs to export (0 = all)",
    )
    parser.add_argument(
        "--enrich-descriptions-from-gt",
        action="store_true",
        help="If hierarchy CSV lacks DESCRIPTION, fill PLTXT from GT (offline/testing)",
    )
    args = parser.parse_args()

    out_dir = Path(args.output_dir).expanduser().resolve()
    base = safe_name(Path(args.input))
    hier_path = (
        Path(args.hierarchy_csv).expanduser().resolve()
        if args.hierarchy_csv
        else out_dir / f"{base}.hierarchy_orchestrator.csv"
    )
    if not hier_path.exists():
        print(f"[error] Missing hierarchy CSV: {hier_path}", flush=True)
        return 2

    template = Path(args.template).expanduser().resolve()
    if not template.exists():
        print(f"[error] Missing template: {template}", flush=True)
        return 2

    rows = read_hierarchy_csv(hier_path)
    gt_path = Path(args.gt).expanduser()
    gt_descriptions = None
    if args.enrich_descriptions_from_gt and gt_path.exists():
        gt_descriptions = load_gt_function_descriptions(gt_path)
    functions = collect_functions(rows, gt_descriptions=gt_descriptions)
    if args.limit > 0:
        functions = functions[: args.limit]

    ctx = load_floc_context_for_input(Path(args.input))
    floc_rows = build_floc_rows(functions, ctx=ctx)
    out_xlsx = out_dir / f"{base}.functional_locations.xlsx"
    write_floc_workbook(template, out_xlsx, floc_rows)

    report: Dict[str, Any] = {
        "hierarchy_csv": str(hier_path),
        "output": str(out_xlsx),
        "function_count": len(functions),
        "floc_row_count": len(floc_rows),
        "functions": [f[0] for f in functions],
    }
    if gt_path.exists():
        report["gt_eval"] = evaluate_against_gt(floc_rows, gt_path)
        ge = report["gt_eval"]
        print(
            f"GT MASK accuracy: {ge['mask_accuracy']*100:.1f}% "
            f"({ge['mask_hit']}/{ge['mask_hit']+ge['mask_miss']})"
        )
        print(
            f"GT DESCRIPTION exact: {ge['description_accuracy']*100:.1f}% "
            f"({ge['description_exact']}/{ge['description_total']})"
        )

    report_path = json_path(out_dir, f"{base}.functional_locations_report.json")
    write_json(report_path, report)
    print(f"Wrote {out_xlsx} ({len(floc_rows)} FL rows, {len(functions)} functions)")
    print(f"Report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
