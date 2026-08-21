#!/usr/bin/env python3
"""
Export hierarchy EQUIPMENT / SUB-EQUIPMENT rows into the SML Equipment load workbook.

Reads hierarchy CSV (orchestrator) and writes a workbook shaped
like docs/examples/SML-Equipment Template RW.xlsx.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set

from dwg_floc_context import (
    combine_valve_type,
    explain_valve_type,
    floc_paths_for_function,
    format_line_eqktx,
    format_valve_eqktx,
    is_line_equipment_tag,
    is_pump_equipment,
    is_valve_equipment,
    load_floc_context_for_input,
    merge_floc_context,
    normalize_pltxt,
)
from dwg_object_type import classify_equipment
from dwg_pure_dump import json_path, safe_name, write_json

TEMPLATE_DEFAULT = Path("docs/examples/SML-Equipment Template RW.xlsx")
SHEET_NAME = "Equipment"
DATA_START_ROW = 7  # 1-based; row 7 is the sample Boiler in the template
VALVE_LAYERS = {"P-VALVEPOS", "P-CVPOS", "1-VALVE TEXT GOR"}

_INVALID_EQUNR_RE = re.compile(r"^(CHAR\s*\d+|FUNCTIONAL\s+LOCATION)$", re.I)

# Driven equipment patterns: pumps (35-24P518) and agitators by L401–L499 range
_PUMP_TAG_NODASH_RE = re.compile(r"^(\d{2}-\d{2})P(\d+)$", re.I)
_AGITATOR_L_NODASH_RE = re.compile(r"^(\d{2}-\d{2})L(4\d{2})$", re.I)


def _motor_tag_for(tag: str, *, tissue_standard: bool = False) -> str:
    """Derive the motor tag for a driven equipment tag.

    SML PS-21 / Valmet PM3: strip letter, append .1  →  35-24P518 → 35-24-518.1
    Valmet Tissue:           append -M1              →  124P-001  → 124P-001-M1
    Returns "" if the tag format is not recognised.
    """
    t = re.sub(r"\s+", "", str(tag or "").strip()).upper()
    if tissue_standard:
        return f"{t}-M1"
    derived = re.sub(r"^(\d{2}-\d{2})[A-Z]+(\d+)$", r"\1-\2.1", t)
    return derived if derived != t else ""


def _is_driven_equipment(tag: str) -> bool:
    """True for pumps and agitators (L401–L499) that must have a motor."""
    t = re.sub(r"\s+", "", str(tag or "").strip()).upper()
    if _PUMP_TAG_NODASH_RE.match(t):
        return True
    m = _AGITATOR_L_NODASH_RE.match(t)
    if m and 401 <= int(m.group(2)) <= 499:
        return True
    return False


def _is_valid_equipment_tag(tag: str) -> bool:
    t = str(tag or "").strip().upper()
    if not t or _INVALID_EQUNR_RE.match(t):
        return False
    return bool(re.match(r"^[\dA-Z][\dA-Z./-]{2,}$", t))

SAP_COLUMNS = [
    "TPLNR",
    "EQUNR",
    "HEQUI",
    "POSNR",
    "EQKTX",
    "EQTYP",
    "EQART",
    "INGRP",
    "GEWRK",
    "IWERK",
    "SWERK",
    "KOSTL",
    "GWLDT",
    "GWLEN",
    "ABCKZ",
    "STORT",
    "BEGRU",
]


_CACHE_META_KEYS = {"input", "model_id", "region", "legend", "tags", "count"}


def _load_valve_cache(path: Path) -> dict[str, dict]:
    """Load per-tag CAD/vision cache written by dwg_valve_classify.py."""
    p = path.expanduser().resolve()
    if not p.exists():
        return {}
    try:
        from dwg_valve_classify import load_valve_cache

        return load_valve_cache(p)
    except Exception:
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as exc:
            print(f"[warn] Failed to read valve cache {p}: {exc}", file=sys.stderr, flush=True)
            return {}
        blob = data.get("tags") if isinstance(data.get("tags"), dict) else data
        return {
            str(k).strip().upper(): v
            for k, v in blob.items()
            if k not in _CACHE_META_KEYS and isinstance(v, dict)
        }


def _valve_hint(tag: str, *, cache: dict[str, dict]) -> dict:
    """Return CAD/vision classification cache entry for a tag."""
    ca = cache.get(tag) or {}
    layer = str(ca.get("layer") or "")
    stored_is_valve = ca.get("is_valve")
    # Explicit False in cache is authoritative (e.g. instruments in GOR drawings).
    if stored_is_valve is False:
        is_valve = False
    else:
        is_valve = bool(stored_is_valve or ca.get("type") or layer in VALVE_LAYERS)
    return {
        "type": ca.get("type") or None,
        "fn": ca.get("fn") or None,
        "is_valve": is_valve,
        "source": ca.get("source"),
    }


def _norm(value: object) -> str:
    s = str(value or "").strip()
    return "" if not s or s.lower() == "nan" else s


def read_hierarchy_csv(path: Path) -> List[Dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as f:
        return [{k: _norm(v) for k, v in row.items()} for row in csv.DictReader(f)]


REASONING_COLUMNS = [
    "EQUNR",
    "FUNCTION",
    "EQKTX",
    "TYPE",
    "SOURCE",
    "AI_DESCRIPTION",
    "REASONING",
]


def write_valve_reasoning_csv(path: Path, reasoning_rows: List[Dict[str, str]]) -> None:
    """Write the valve classification reasoning CSV alongside the equipment workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=REASONING_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(reasoning_rows)


def build_equipment_rows(
    hierarchy_rows: Sequence[Dict[str, str]],
    *,
    limit_functions: int = 0,
    ctx: Optional[Dict[str, str]] = None,
    valve_cache: dict[str, dict] | None = None,
    reasoning_out: Optional[List[Dict[str, str]]] = None,
) -> List[Dict[str, str]]:
    """
    Walk hierarchy stream and emit one Equipment row per EQUIPMENT / SUB-EQUIPMENT.

    FUNCTION headers set the install TPLNR. SUB-EQUIPMENT gets HEQUI = last EQUIPMENT.
    reasoning_out: if a list is passed, one reasoning dict per valve is appended to it.
    Columns: EQUNR, FUNCTION, EQKTX, TYPE, SOURCE, AI_DESCRIPTION, REASONING.
    """
    c = merge_floc_context(ctx)
    plant = c["plant"]
    cache: dict[str, dict] = valve_cache if valve_cache is not None else {}
    allowed: Optional[Set[str]] = None
    if limit_functions > 0:
        seen_fns: List[str] = []
        for row in hierarchy_rows:
            fn = _norm(row.get("FUNCTION")).upper().replace(" ", "")
            eq = _norm(row.get("EQUIPMENT"))
            sub = _norm(row.get("SUB-EQUIPMENT"))
            if fn and not eq and not sub and fn not in seen_fns:
                seen_fns.append(fn)
                if len(seen_fns) >= limit_functions:
                    break
        allowed = set(seen_fns)

    def blank_row(**kwargs: str) -> Dict[str, str]:
        row = {k: "" for k in SAP_COLUMNS}
        row["EQTYP"] = "P"
        row["INGRP"] = c.get("planning_group", "P01")
        row["IWERK"] = plant
        row["SWERK"] = plant
        row["ABCKZ"] = "D"
        row["BEGRU"] = "001"
        row.update(kwargs)
        return row

    out: List[Dict[str, str]] = []
    current_fn = ""
    current_tplnr = ""
    current_equipment = ""
    current_line = ""  # last line-equipment tag in this FUNCTION; valves after it become sub-equipment
    top_pos_by_tplnr: Dict[str, int] = {}
    sub_pos_by_parent: Dict[str, int] = {}
    emitted: Set[str] = set()

    for row in hierarchy_rows:
        fn = _norm(row.get("FUNCTION")).upper().replace(" ", "")
        eq = _norm(row.get("EQUIPMENT")).upper().replace(" ", "")
        sub = _norm(row.get("SUB-EQUIPMENT")).upper().replace(" ", "")
        desc = normalize_pltxt(_norm(row.get("DESCRIPTION")))

        if fn and not eq and not sub:
            if allowed is not None and fn not in allowed:
                current_fn = ""
                current_tplnr = ""
                current_equipment = ""
                current_line = ""
                continue
            current_fn = fn
            current_tplnr = floc_paths_for_function(fn, c)["function"]
            current_equipment = ""
            current_line = ""
            continue

        if allowed is not None and current_fn and current_fn not in allowed:
            continue
        if not current_tplnr:
            continue

        if eq:
            tag = eq
            if is_line_equipment_tag(tag):
                # This is a pipe/line tag — it becomes the parent for subsequent valves.
                hequi = ""
                current_line = tag
                current_equipment = tag
            elif current_line and is_valve_equipment(tag, _norm(row.get("DESCRIPTION"))):
                # SOP rule: valve/instrument appearing after a line in the same FUNCTION
                # is sub-equipment of that line (line = equipment, valve = sub-equipment).
                hequi = current_line
                current_equipment = tag
            else:
                # Non-line, non-valve equipment (motor, instrument, etc.) — direct FUNCTION child.
                hequi = ""
                current_line = ""
                current_equipment = tag
        elif sub:
            tag = sub
            hequi = current_equipment
        else:
            continue

        if not _is_valid_equipment_tag(tag):
            continue

        if tag in emitted:
            continue
        emitted.add(tag)

        if hequi:
            parent_key = f"{current_tplnr}|{hequi}"
            sub_pos_by_parent[parent_key] = sub_pos_by_parent.get(parent_key, 0) + 10
            posnr = f"{sub_pos_by_parent[parent_key]:04d}"
        else:
            top_pos_by_tplnr[current_tplnr] = top_pos_by_tplnr.get(current_tplnr, 0) + 10
            posnr = f"{top_pos_by_tplnr[current_tplnr]:04d}"
        raw_desc = _norm(row.get("DESCRIPTION"))  # original AI text, before abbreviation
        eqktx = desc if desc else normalize_pltxt(tag)
        if eqktx and not eqktx.startswith(tag):
            eqktx = normalize_pltxt(f"{tag} {eqktx}")
        eqart, gewrk = classify_equipment(tag, eqktx)
        tag_upper = tag.upper()
        hint = _valve_hint(tag_upper, cache=cache)
        cache_fn = hint.get("fn")
        cache_type = hint.get("type")
        if cache_type:
            cache_type = combine_valve_type(cache_type, desc) or cache_type
        # Vision/CAD cache parent fn reassigns TPLNR so SAP placement matches EQKTX.
        effective_fn = cache_fn or current_fn
        effective_tplnr = (
            floc_paths_for_function(cache_fn, c)["function"] if cache_fn else current_tplnr
        )
        is_valve = is_valve_equipment(tag, eqktx) or bool(hint.get("is_valve"))
        if is_valve:
            eqktx = format_valve_eqktx(tag, effective_fn, eqktx, valve_type_override=cache_type)
            if reasoning_out is not None:
                source = hint.get("source")
                if source == "tipo_code" and cache_type:
                    vtype = cache_type
                    vsource = "TIPO_CODE"
                    tipo = str(cache.get(tag_upper, {}).get("tipo") or "").strip()
                    vreason = (
                        f"GOR TIPO_VALVOLA '{tipo}' → {vtype}"
                        if tipo
                        else f"GOR TIPO code → {vtype}"
                    )
                elif source == "gor_tag" and cache_type:
                    vtype = cache_type
                    vsource = "GOR_TAG"
                    vreason = f"GOR tag pattern → {vtype}"
                elif source == "vision" and cache_type:
                    vtype = cache_type
                    vsource = "VISION"
                    vreason = f"Tight-crop + legend classification → {vtype}"
                elif source == "cad_layer":
                    layer_name = cache.get(tag_upper, {}).get("layer") or "valve layer"
                    vtype, vsource, vreason = explain_valve_type(tag, desc)
                    vsource = "CAD_LAYER"
                    vreason = f"Tag text on {layer_name} → valve"
                else:
                    vtype, vsource, vreason = explain_valve_type(tag, desc)
                    if cache_type:
                        vtype = cache_type
                reasoning_out.append({
                    "EQUNR": tag[:18],
                    "FUNCTION": effective_fn,
                    "EQKTX": eqktx[:40],
                    "TYPE": vtype,
                    "SOURCE": vsource,
                    "AI_DESCRIPTION": raw_desc,
                    "REASONING": vreason,
                })
        else:
            eqktx = format_line_eqktx(tag, eqktx, hequi=hequi)
        out.append(
            blank_row(
                TPLNR=effective_tplnr[:30],
                EQUNR=tag[:18],
                HEQUI=hequi[:18],
                POSNR=posnr,
                EQKTX=eqktx[:40],
                EQART=eqart,
                GEWRK=gewrk,
            )
        )

    # Second pass: inject implicit motor rows for driven equipment with no motor emitted
    for eq_row in list(out):
        eq_tag = eq_row["EQUNR"]
        if eq_row["HEQUI"] or not _is_driven_equipment(eq_tag):
            continue
        motor_tag = _motor_tag_for(eq_tag)
        if not motor_tag or motor_tag in emitted:
            continue
        emitted.add(motor_tag)
        parent_tplnr = eq_row["TPLNR"]
        parent_key = f"{parent_tplnr}|{eq_tag}"
        sub_pos_by_parent[parent_key] = sub_pos_by_parent.get(parent_key, 0) + 10
        motor_eqktx = normalize_pltxt(f"{motor_tag} MOTOR")[:40]
        m_eqart, m_gewrk = classify_equipment(motor_tag, motor_eqktx)
        out.append(
            blank_row(
                TPLNR=parent_tplnr,
                EQUNR=motor_tag[:18],
                HEQUI=eq_tag[:18],
                POSNR=f"{sub_pos_by_parent[parent_key]:04d}",
                EQKTX=motor_eqktx,
                EQART=m_eqart,
                GEWRK=m_gewrk,
            )
        )

    return out


def write_equipment_workbook(
    template_path: Path,
    out_path: Path,
    equipment_rows: List[Dict[str, str]],
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # Strip template comments (Turkish column metadata) — they get restructured
    # to a different internal path by openpyxl, which breaks some xlsx readers.
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None:
                cell.comment = None

    # Clear prior data from the sample row downward (keep header block rows 1-6).
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)

    for r_i, row in enumerate(equipment_rows, start=DATA_START_ROW):
        ws.cell(r_i, 1, value=None)  # ID column unused
        for c_i, key in enumerate(SAP_COLUMNS, start=2):
            ws.cell(r_i, c_i, value=row.get(key) or None)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def resolve_hierarchy_csv(out_dir: Path, base: str, explicit: str = "") -> Optional[Path]:
    if explicit:
        p = Path(explicit).expanduser().resolve()
        return p if p.exists() else None
    primary = out_dir / f"{base}.hierarchy_orchestrator.csv"
    return primary if primary.exists() else None


def main() -> int:
    parser = argparse.ArgumentParser(description="Export SAP Equipment workbook")
    parser.add_argument("--hierarchy-csv", default="")
    parser.add_argument("--input", default="inputs/Broke System.dwg", help="Used for output stem")
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--template", default=str(TEMPLATE_DEFAULT))
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional max FUNCTIONs whose children to export (0 = all)",
    )
    args = parser.parse_args()

    out_dir = Path(args.output_dir).expanduser().resolve()
    base = safe_name(Path(args.input))
    hier_path = resolve_hierarchy_csv(out_dir, base, args.hierarchy_csv)
    if not hier_path:
        print(f"[error] Missing hierarchy CSV under {out_dir}", flush=True)
        return 2

    template = Path(args.template).expanduser().resolve()
    if not template.exists():
        print(f"[error] Missing template: {template}", flush=True)
        return 2

    ctx = load_floc_context_for_input(Path(args.input))
    hierarchy_rows = read_hierarchy_csv(hier_path)
    reasoning_rows: List[Dict[str, str]] = []
    cache_path = json_path(out_dir, f"{base}.valve_types.json")
    valve_cache = _load_valve_cache(cache_path)
    equipment_rows = build_equipment_rows(
        hierarchy_rows,
        limit_functions=args.limit,
        ctx=ctx,
        reasoning_out=reasoning_rows,
        valve_cache=valve_cache,
    )
    out_xlsx = out_dir / f"{base}.equipment.xlsx"
    write_equipment_workbook(template, out_xlsx, equipment_rows)

    reasoning_csv = out_dir / f"{base}.valve_reasoning.csv"
    write_valve_reasoning_csv(reasoning_csv, reasoning_rows)

    tplnrs = sorted({r["TPLNR"] for r in equipment_rows})
    report: Dict[str, Any] = {
        "hierarchy_csv": str(hier_path),
        "output": str(out_xlsx),
        "valve_reasoning_csv": str(reasoning_csv),
        "equipment_row_count": len(equipment_rows),
        "valve_row_count": len(reasoning_rows),
        "tplnr_count": len(tplnrs),
        "tplnrs": tplnrs,
        "with_hequi": sum(1 for r in equipment_rows if r.get("HEQUI")),
    }
    report_path = json_path(out_dir, f"{base}.equipment_report.json")
    write_json(report_path, report)
    print(
        f"Wrote {out_xlsx} ({len(equipment_rows)} equipment rows, "
        f"{len(tplnrs)} FLOCs, {report['with_hequi']} with HEQUI)"
    )
    print(f"Valve reasoning: {reasoning_csv} ({len(reasoning_rows)} valves)")
    print(f"Report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
