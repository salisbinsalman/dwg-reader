#!/usr/bin/env python3
"""Unit tests for FLOC path/description + SAP export (no Bedrock)."""

from __future__ import annotations

import csv
import json
import tempfile
import unittest
from pathlib import Path

from dwg_floc_context import (
    abbrev_data,
    abbreviate_pltxt,
    build_tplnr,
    description_from_nearby,
    floc_paths_for_function,
    load_sml_abbreviations,
    normalize_pltxt,
)
from dwg_pid_hierarchy_ai import rows_from_ai, title_context
from export_sap_floc import _strip_trailing_spec, build_floc_rows, collect_functions, evaluate_against_gt, write_floc_workbook


ROOT = Path(__file__).resolve().parent
ABBREV_JSON = ROOT / "inputs/sml_abbreviations.json"
HIERARCHY_CSV = ROOT / "outputs/Broke System.hierarchy_orchestrator.csv"


class FlocContextTests(unittest.TestCase):
    def test_tplnr_function_path(self) -> None:
        paths = floc_paths_for_function("35-24L009")
        self.assertEqual(paths["plant"], "5001")
        self.assertEqual(paths["line"], "5001-PM03")
        self.assertEqual(paths["process"], "5001-PM03-BR")
        self.assertEqual(paths["subprocess"], "5001-PM03-BR-BR1")
        self.assertEqual(paths["function"], "5001-PM03-BR-BR1-35-24L009")
        self.assertLessEqual(len(paths["function"]), 30)

    def test_build_tplnr_truncates(self) -> None:
        s = build_tplnr("5001", "PM03", "BR", "BR1", "35-24L009EXTRA")
        self.assertEqual(len(s), 30)

    def test_normalize_pltxt(self) -> None:
        d = normalize_pltxt("35-24L009 Broke Roll Pulper")
        self.assertTrue(d.startswith("35-24L009"))
        self.assertIn("PLPR", d)
        self.assertLessEqual(len(d), 40)

    def test_description_fallback(self) -> None:
        d = description_from_nearby("35-24P519", ["BROKE ROLL PULPER PUMP", "35-24P519"])
        self.assertTrue(d.startswith("35-24P519"))
        self.assertIn("PMP", d)
        self.assertLessEqual(len(d), 40)


class SmlAbbreviationJsonTests(unittest.TestCase):
    def test_json_exists_and_has_conveyor(self) -> None:
        self.assertTrue(ABBREV_JSON.exists())
        data = abbrev_data(ABBREV_JSON)
        self.assertIn("abbreviations", data)
        self.assertEqual(data["abbreviations"]["CONVEYOR"], "CVYR")
        self.assertGreaterEqual(len(data["abbreviations"]), 300)

    def test_load_sml_abbreviations_from_json(self) -> None:
        mapping = load_sml_abbreviations(ABBREV_JSON)
        self.assertEqual(mapping["PUMP"], "PMP")
        self.assertEqual(mapping["CONVEYOR"], "CVYR")


class AbbreviatePltxtTests(unittest.TestCase):
    def test_review_case_conveyor(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L004 BROKE CONVEYOR 2"),
            "35-24L004 BROKE CVYR 2",
        )

    def test_conveyor_plural(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L006 BROKE CONVEYORS 3"),
            "35-24L006 BROKE CVYR 3",
        )

    def test_broke_stays_full(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L004 BROKE CONVEYOR 2"),
            "35-24L004 BROKE CVYR 2",
        )
        self.assertIn("BROKE", abbreviate_pltxt("35-24L009 BROKE ROLL PULPER"))

    def test_domain_phrase_broke_roll_pulper(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L009 BROKE ROLL PULPER"),
            "35-24L009 BROKE ROLL PLPR",
        )

    def test_domain_phrase_size_press_pulper(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L003 SIZE PRESS PULPER"),
            "35-24L003 SIZE PRESS PLPR",
        )

    def test_pump_abbreviated_after_phrase(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24P519 BROKE ROLL PLPR PUMP"),
            "35-24P519 BROKE ROLL PLPR PMP",
        )

    def test_sml_phrase_hand_valve(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24HV-548 HAND VALVE"),
            "35-24HV-548 HV",
        )

    def test_sml_phrase_normally_closed(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("DRAIN VALVE NORMALLY CLOSED"),
            "DRN VLV NC",
        )

    def test_preserves_tag_and_number(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24L006 BROKE CONVEYOR 3"),
            "35-24L006 BROKE CVYR 3",
        )

    def test_already_abbreviated_tokens_unchanged(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24P519 BROKE ROLL PLPR PMP"),
            "35-24P519 BROKE ROLL PLPR PMP",
        )

    def test_motor_tank_valve_shower(self) -> None:
        self.assertEqual(abbreviate_pltxt("PUMP MOTOR"), "PMP MTR")
        self.assertEqual(abbreviate_pltxt("STORAGE TANK"), "STOR TNK")
        self.assertEqual(abbreviate_pltxt("CONTROL VALVE"), "CTRL VLV")
        self.assertEqual(abbreviate_pltxt("SHOWER S1"), "SHW S1")

    def test_overflow_uses_sml_standard(self) -> None:
        self.assertEqual(abbreviate_pltxt("LINE OVERFLOW"), "LN OVRFL")

    def test_max_length_40(self) -> None:
        long_text = "35-24L001 BROKE ROLL PULPER PUMP MOTOR AGITATOR VALVE SHOWER"
        self.assertLessEqual(len(abbreviate_pltxt(long_text)), 40)

    def test_emergency_stop_phrase(self) -> None:
        self.assertEqual(
            abbreviate_pltxt("35-24ES-508 WINDER AREA EMERGENCY STOP"),
            "35-24ES-508 WINDER AREA E-STOP",
        )

    def test_idempotent_on_normalized_text(self) -> None:
        once = abbreviate_pltxt("35-24L004 BROKE CONVEYOR 2")
        twice = abbreviate_pltxt(once)
        self.assertEqual(once, twice)


class RowsFromAiTests(unittest.TestCase):
    def test_deterministic_mask_and_description(self) -> None:
        ctx = title_context({"title_block": []}, "Broke System")
        parsed = {
            "sub_process": "BR1",
            "function": "35-24L009",
            "description": "35-24L009 BROKE ROLL PLPR",
            "rows": [
                {"equipment": "35-24-189", "subequipment": "", "description": "LN OVFL"},
                {"equipment": "", "subequipment": "35-24-194", "description": "HV DRN"},
            ],
        }
        rows, order = rows_from_ai(1, ctx, "35-24L009", parsed, parent={"nearby_descriptions": ["BROKE ROLL PULPER"]})
        self.assertGreater(order, 1)
        masks = [r["MASK"] for r in rows if r["MASK"]]
        self.assertIn("5001-PM03-BR", masks)
        self.assertIn("5001-PM03-BR-BR1", masks)
        self.assertIn("5001-PM03-BR-BR1-35-24L009", masks)
        fn_row = next(r for r in rows if r["FUNCTION"] == "35-24L009")
        self.assertEqual(fn_row["MASK"], "5001-PM03-BR-BR1-35-24L009")
        self.assertTrue(fn_row["DESCRIPTION"].startswith("35-24L009"))
        self.assertLessEqual(len(fn_row["DESCRIPTION"]), 40)
        child = next(r for r in rows if r["EQUIPMENT"] == "35-24-189")
        self.assertEqual(child["MASK"], "")


class StripTrailingSpecTests(unittest.TestCase):
    def test_strips_hyphenated_model(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L008 WINDER PLPR HP-33G2"), "35-24L008 WINDER PLPR")
        self.assertEqual(_strip_trailing_spec("35-24L002 BROKE THICKENER GT-2540"), "35-24L002 BROKE THICKENER")
        self.assertEqual(_strip_trailing_spec("35-24L010 BROKE SCREEN PROFS-250HC"), "35-24L010 BROKE SCREEN")

    def test_strips_mixed_alphanumeric(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L009 BROKE ROLL PLPR 141BDTPD"), "35-24L009 BROKE ROLL PLPR")

    def test_keeps_plain_integer(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L006 BROKE CONVEYOR 3"), "35-24L006 BROKE CONVEYOR 3")
        self.assertEqual(_strip_trailing_spec("35-24L004 BROKE CONVEYOR 2"), "35-24L004 BROKE CONVEYOR 2")

    def test_keeps_clean_description(self) -> None:
        self.assertEqual(_strip_trailing_spec("35-24L003 SIZE PRESS PLPR"), "35-24L003 SIZE PRESS PLPR")


class ExportFlocTests(unittest.TestCase):
    def test_build_and_write_workbook(self) -> None:
        functions = [
            ("35-24L009", "5001-PM03-BR-BR1-35-24L009", "35-24L009 BROKE ROLL PLPR"),
            ("35-24P519", "5001-PM03-BR-BR1-35-24P519", "35-24P519 BROKE ROLL PLPR PMP"),
        ]
        floc_rows = build_floc_rows(functions)
        self.assertEqual(len(floc_rows), 6)
        self.assertEqual(floc_rows[0]["TPLNR"], "5001")
        self.assertEqual(floc_rows[3]["TPLNR"], "5001-PM03-BR-BR1")
        self.assertEqual(floc_rows[4]["TPLMA"], "5001-PM03-BR-BR1")
        self.assertEqual(floc_rows[4]["TPLNR"], "5001-PM03-BR-BR1-35-24L009")
        self.assertEqual(floc_rows[4]["FLTYP"], "M")
        self.assertEqual(floc_rows[4]["SWERK"], "5001")
        self.assertEqual(floc_rows[4]["ABCKZ"], "D")
        self.assertEqual(floc_rows[4]["INGRP"], "P01")
        self.assertLessEqual(len(floc_rows[4]["PLTXT"]), 40)

        template = ROOT / "docs/examples/final-output-template.xlsx"
        self.assertTrue(template.exists())
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "out.xlsx"
            write_floc_workbook(template, out, floc_rows)
            self.assertTrue(out.exists())
            from openpyxl import load_workbook

            wb = load_workbook(out)
            ws = wb["Functional Location"]
            self.assertEqual(ws.cell(8, 2).value, "5001")
            self.assertEqual(ws.cell(12, 2).value, "5001-PM03-BR-BR1-35-24L009")

    def test_function_rows_populate_gewrk(self) -> None:
        functions = [
            ("35-24L001", "5001-PM03-BR-BR1-35-24L001", "35-24L001 PRESS PLPR"),
            ("35-24L004", "5001-PM03-BR-BR1-35-24L004", "35-24L004 BROKE CVYR 2"),
            ("35-24P519", "5001-PM03-BR-BR1-35-24P519", "35-24P519 BROKE ROLL PLPR PMP"),
            ("35-24T602", "5001-PM03-BR-BR1-35-24T602", "35-24T602 BROKE STOR TWR 4000M3"),
            ("35-24ES-508", "5001-PM03-BR-BR1-35-24ES-508", "35-24ES-508 WINDER AREA E-STOP"),
        ]
        floc_rows = build_floc_rows(functions)
        fn_rows = floc_rows[4:]
        self.assertEqual(fn_rows[0]["GEWRK"], "MECH")
        self.assertEqual(fn_rows[0]["EQART"], "2005")
        self.assertEqual(fn_rows[1]["GEWRK"], "MECH")  # CVYR → conveyors
        self.assertEqual(fn_rows[2]["GEWRK"], "MECH")  # pump tag prefix
        self.assertEqual(fn_rows[3]["GEWRK"], "MECH")  # TWR → tank
        self.assertEqual(fn_rows[4]["GEWRK"], "ELEC")  # ES → switch
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "h.csv"
            with p.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(
                    f,
                    fieldnames=["SUB-PROCESS", "FUNCTION", "EQUIPMENT", "SUB-EQUIPMENT", "MASK", "DESCRIPTION"],
                )
                w.writeheader()
                w.writerow(
                    {
                        "SUB-PROCESS": "BR1",
                        "FUNCTION": "35-24L004",
                        "EQUIPMENT": "",
                        "SUB-EQUIPMENT": "",
                        "MASK": "5001-PM03-BR-BR1-35-24L004",
                        "DESCRIPTION": "35-24L004 BROKE CONVEYOR 2",
                    }
                )
            rows = list(csv.DictReader(p.open(encoding="utf-8")))
            fns = collect_functions(rows)
            self.assertEqual(fns[0][2], "35-24L004 BROKE CVYR 2")

    def test_collect_functions_and_gt_eval(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "h.csv"
            with p.open("w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(
                    f,
                    fieldnames=["SUB-PROCESS", "FUNCTION", "EQUIPMENT", "SUB-EQUIPMENT", "MASK", "DESCRIPTION"],
                )
                w.writeheader()
                w.writerow(
                    {
                        "SUB-PROCESS": "BR1",
                        "FUNCTION": "35-24L009",
                        "EQUIPMENT": "",
                        "SUB-EQUIPMENT": "",
                        "MASK": "5001-PM03-BR-BR1-35-24L009",
                        "DESCRIPTION": "35-24L009 BROKE ROLL PLPR",
                    }
                )
            rows = list(csv.DictReader(p.open(encoding="utf-8")))
            fns = collect_functions(rows)
            self.assertEqual(fns[0][0], "35-24L009")
            floc_rows = build_floc_rows(fns)
            gt = ROOT / "inputs/gt_hierarchy_broke_system.xlsx"
            if gt.exists():
                report = evaluate_against_gt(floc_rows, gt)
                self.assertGreaterEqual(report["functions_compared"], 1)
                self.assertEqual(report["mask_hit"], 1)


class RealHierarchyAbbreviationTests(unittest.TestCase):
    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_conveyor_descriptions_abbreviated_in_export(self) -> None:
        rows = list(csv.DictReader(HIERARCHY_CSV.open(encoding="utf-8")))
        fns = collect_functions(rows)
        by_tag = {tag: desc for tag, _mask, desc in fns}
        for tag in ("35-24L004", "35-24L006"):
            if tag in by_tag:
                self.assertIn("CVYR", by_tag[tag], msg=f"{tag} -> {by_tag[tag]}")
                self.assertNotIn("CONVEYOR", by_tag[tag], msg=f"{tag} -> {by_tag[tag]}")

    @unittest.skipUnless(HIERARCHY_CSV.exists(), "requires hierarchy orchestrator CSV")
    def test_no_unabbreviated_pump_motor_in_function_descriptions(self) -> None:
        rows = list(csv.DictReader(HIERARCHY_CSV.open(encoding="utf-8")))
        fns = collect_functions(rows)
        offenders = []
        for tag, _mask, desc in fns:
            words = set(desc.split())
            if "PUMP" in words or "MOTOR" in words or "CONVEYOR" in words:
                offenders.append((tag, desc))
        self.assertEqual(offenders, [], msg=json.dumps(offenders[:10], indent=2))


if __name__ == "__main__":
    unittest.main()
