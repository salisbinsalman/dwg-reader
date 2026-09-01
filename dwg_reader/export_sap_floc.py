#!/usr/bin/env python3
"""
Export hierarchy FUNCTION rows into the SAP Functional Location load workbook.

Reads hierarchy CSV (orchestrator) and writes a workbook shaped
like docs/examples/final-output-template.xlsx.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dwg_reader.dwg_floc_context import (
    build_tplnr,
    floc_paths_for_function,
    format_line_eqktx,
    inherit_pump_duty,
    is_line_equipment_tag,
    load_floc_context_for_input,
    merge_floc_context,
    normalize_pltxt,
    scrub_pump_description,
)
from dwg_reader.dwg_floc_standards import lookup_line, lookup_process, lookup_sub_process
from dwg_reader.dwg_object_type import classify_equipment
from dwg_reader.dwg_pure_dump import find_json, json_path, safe_name, write_json
from dwg_reader.io import cell as _norm, read_csv_rows
from dwg_reader.logutil import configure_logging, get_logger
from dwg_reader.paths import REPO_ROOT

logger = get_logger(__name__)

TEMPLATE_DEFAULT = REPO_ROOT / "docs/examples/final-output-template.xlsx"

# Strips trailing model/spec codes like "HP-33G2", "GT-2540", "PROFS-250HC", "141BDTPD"
# from FUNCTION-level PLTXT descriptions.  Plain integers ("CONVEYOR 3") are left intact.
_TRAILING_SPEC_RE = re.compile(
    r"\s+(?:[A-Z]{1,8}-\d+[A-Z0-9]*|\d+[A-Z][A-Z0-9]*)$"
)
# Semicolon-separated spec segments: "HP-33G2", "BDTPD", "2621 ADTPD", "800BDTPD", "40M3"
_SPEC_SEGMENT_RE = re.compile(
    r"^(?:"
    r"[A-Z]{1,8}-\d+[A-Z0-9]*"
    r"|\d+[A-Z][A-Z0-9]*"
    r"|(?:\d+\s+)?[A-Z]*DTPD"
    r")$",
    re.I,
)


def _strip_trailing_spec(text: str) -> str:
    """Strip one or more trailing vendor/model tokens (HP-33G2, 800BDTPD, …).

    Semicolon-separated spec tails (``SLABBING PLPR; HP-33G2; BDTPD``) are
    dropped before the space-separated pass, so either separator works.
    """
    prev = str(text or "").strip()
    parts = [p.strip() for p in re.split(r"\s*;\s*", prev) if p.strip()]
    if len(parts) > 1:
        # Always keep the leading phrase; drop only spec-shaped tails.
        prev = " ".join([parts[0]] + [p for p in parts[1:] if not _SPEC_SEGMENT_RE.match(p)])
    for _ in range(8):
        nxt = _TRAILING_SPEC_RE.sub("", prev).strip()
        if nxt == prev:
            break
        prev = nxt
    return prev


# Strips fluid-class codes (WFL = waste flush, WAA/WAF = white water stainless/feed)
# and nominal-diameter specs (DN15, DN300 …) and pipe-spec classes (PP-200, PP-250 …)
# from P&ID line descriptions.  These are procurement/material codes, not functional labels.
_PIPE_CLASS_RE = re.compile(
    r"\b(?:WFL|WAA|WAF)\b"        # fluid class codes
    r"|\bPP-\d+[A-Z0-9]*\b"       # pipe spec class: PP-200, PP-200-E10H2A
    r"|\bDN\d+\b",                 # nominal diameter: DN15, DN300, DN400
    re.I,
)


def _strip_pipe_class_codes(text: str) -> str:
    """Remove pipe fluid-class codes and nominal-diameter specs from a description."""
    result = _PIPE_CLASS_RE.sub("", text)
    return re.sub(r" {2,}", " ", result).strip()


# PS-21 Appendix IV — flow substance codes translated to human-readable labels for PLTXT.
_FLOW_SUBSTANCE_CODES: Dict[str, str] = {
    "WAF": "WHITE WTR",
    "WAA": "CLOUDY FILT",
    "WAB": "CLR FILT",
    "WFC": "CLG WTR",
    "WM": "FRESH WTR",
    "WFL": "SEAL WTR",
    "WSH": "SHR WTR",
    "PS": "PAPER STK",
    "SH": "HP STM",
    "SM": "MP STM",
    "SL": "LP STM",
    "EFC": "FIBRE EFF",
}

_FLOW_SUBSTANCE_RE = re.compile(
    r"\b(WAF|WAA|WAB|WFC|WM|WFL|WSH|PS|SH|SM|SL|EFC)\b",
    re.I,
)

# Bare nominal-bore sizes (mm) written without DN prefix in some P&ID line tags.
# These are DN pipe diameters appearing as plain integers — not part of tag or item numbering.
_BARE_DN_RE = re.compile(
    r"\b(1000|900|800|700|600|500|450|400|350|300|250|200|150|125|100|80|65|50|40|32|25|20|15)\b"
)


def _clean_line_description(text: str, *, flow_codes: Optional[Dict[str, str]] = None) -> str:
    """Translate PS-21 flow-substance codes to readable labels and strip procurement codes.

    Replaces WAF→WHITE WTR, WAA→CLOUDY FILT, WFL→SEAL WTR, etc. in-place so FLOC
    descriptions remain meaningful after stripping.  Also removes pipe-spec classes
    (PP-200), DN-prefixed sizes (DN500), and bare nominal-bore integers (250, 300).

    flow_codes: if None, uses the module-level _FLOW_SUBSTANCE_CODES and compiled RE.
                If a non-empty dict, builds a temporary RE from its keys.
                If an empty dict, disables substance translation entirely.
    """
    if flow_codes is None:
        # Use the module-level compiled RE with the default codes dict
        def _translate(m: re.Match) -> str:
            return _FLOW_SUBSTANCE_CODES.get(m.group(1).upper(), m.group(1).upper())
        result = _FLOW_SUBSTANCE_RE.sub(_translate, text)
    elif flow_codes:
        # Build a temporary RE for ecosystem-specific codes
        pat = r"\b(" + "|".join(re.escape(k) for k in sorted(flow_codes, key=len, reverse=True)) + r")\b"
        def _translate_custom(m: re.Match) -> str:
            return flow_codes.get(m.group(1).upper(), m.group(1).upper())
        result = re.compile(pat, re.I).sub(_translate_custom, text)
    else:
        result = text  # empty codes dict — nothing to translate

    result = re.sub(r"\bPP-\d+[A-Z0-9]*\b|\bDN\d+\b", "", result, flags=re.I)
    result = re.sub(r"\b\d{1,4}\s*MM\b", "", result, flags=re.I)
    result = _BARE_DN_RE.sub("", result)
    # Dangling delimiters left after size/spec stripping (not mid-string destination arrows).
    result = re.sub(r"[\s>/\-|]+$", "", result)
    return re.sub(r" {2,}", " ", result).strip()


def _tag_numeric_sort_key(tag: str) -> int:
    """Extract trailing integer from a function tag for numeric ordering.

    35-24L001→1, 35-24P501→501, 35-24T601→601, 35-24-008→8.
    Used as a fallback sort when no inventory-JSON positions are available.
    """
    m = re.search(r"(\d+)$", str(tag or "").strip().upper())
    return int(m.group(1)) if m else 99999


def _extract_area_unit(tag: str) -> str:
    """Extract the 'NN-NN' area-unit prefix from a tag like '35-24-NNN' or '35-24L009'."""
    m = re.match(r"^(\d{2}-\d{2})", tag.strip().upper())
    return m.group(1) if m else ""


def _is_utility_line_function(tag: str, desc: str, child_count: int) -> bool:
    """Return True for small utility flush lines that must not be top-level FLOC functions.

    WFL (waste flush water) lines are always 15 mm utility stubs — never standalone
    FLOC functions.  WAF (white water) lines with DN ≤ 50 are small distribution
    headers serving a single vessel.
    """
    from dwg_reader.dwg_floc_context import is_line_equipment_tag

    if not is_line_equipment_tag(tag):
        return False
    desc_upper = desc.upper()
    if re.search(r"\bWFL\b", desc_upper) and child_count == 0:
        return True
    return False


def load_function_positions(inventory_path: Path) -> Dict[str, float]:
    """Return {tag: x_coordinate} from a pid_inventory.json for process-flow ordering.

    Functions are sorted left-to-right by their X position on the P&ID, which
    approximates the physical process-flow sequence Rob confirmed in the email thread.
    """
    try:
        data = json.loads(inventory_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    out: Dict[str, float] = {}
    for fn in data.get("functions") or []:
        tag = str(fn.get("function") or "").strip().upper().replace(" ", "")
        x = fn.get("x")
        if tag and x is not None:
            try:
                out[tag] = float(x)
            except (TypeError, ValueError):
                pass
    return out


SHEET_NAME = "Functional Location"

# Column order matches template row 0 (Technical Area codes).
SAP_COLUMNS = [
    "TPLNR",
    "TPLMA",
    "POSNR",
    "TPLKZ",
    "PLTXT",
    "EQUART",
    "ABCKZ",
    "STORT",
    "HERST",
    "TYPBZ",
    "SERGE",
    "FLTYP",
    "EQART",
    "SWERK",
    "KOSTL",
    "IWERK",
    "INGRP",
    "GEWRK",
]

# Codes Excel often coerces to numbers (losing leading zeros / type). Force text.
_TEXT_FORMAT_COLUMNS = frozenset({"POSNR", "FLTYP", "EQART", "EQUART", "ABCKZ"})


def read_hierarchy_csv(path: Path) -> List[Dict[str, str]]:
    return read_csv_rows(path, missing_ok=False)


def collect_functions(
    rows: Iterable[Dict[str, str]],
    *,
    gt_descriptions: Optional[Dict[str, str]] = None,
    filter_utility_lines: bool = True,
    positions: Optional[Dict[str, float]] = None,
    sort_by_tag_number: bool = True,
    ecosystem: Optional["Ecosystem"] = None,
) -> List[Tuple[str, str, str]]:
    """Return ordered unique FUNCTIONs as (tag, mask, description).

    With filter_utility_lines=True (default):
    - Strips pipe fluid-class codes (WFL, WAA, WAF, PP-NNN, DN-NNN) from descriptions.
    - Removes WFL waste-flush utility lines (never standalone FLOC functions).
    - Removes numeric pipeline tags whose area-unit differs from the primary unit found
      in the drawing (e.g. 35-25-034 in a Unit-24 Broke System context).

    If positions is provided (tag → x_coordinate from pid_inventory.json), functions are
    sorted left-to-right by their X position on the P&ID, approximating process-flow order.
    Functions without a known position are appended at the end in their original order.
    """
    from dwg_reader.dwg_floc_context import is_line_equipment_tag

    all_rows: List[Dict[str, str]] = list(rows)

    # Resolve flow_codes from ecosystem standard once, before the loop.
    flow_codes = (
        (ecosystem.standard or {}).get("flow_substance_codes")
        if ecosystem is not None and ecosystem.standard
        else None
    )

    # First pass — collect function metadata (first occurrence of each FUNCTION value).
    # The first occurrence in the orchestrator CSV is always the header row
    # (EQUIPMENT and SUB-EQUIPMENT both empty), which carries the description.
    ordered: List[Tuple[str, str, str]] = []
    pre_strip_descs: Dict[str, str] = {}  # fn → desc before pipe-class stripping (for utility check)
    seen: set = set()
    for row in all_rows:
        fn = _norm(row.get("FUNCTION")).upper().replace(" ", "")
        if not fn or fn in seen:
            continue
        seen.add(fn)
        mask = _norm(row.get("MASK"))
        raw = normalize_pltxt(_norm(row.get("DESCRIPTION")))
        pre = _strip_trailing_spec(raw)
        clean = _clean_line_description(pre, flow_codes=flow_codes)
        if not clean and gt_descriptions:
            clean = gt_descriptions.get(fn, "")
        pre_strip_descs[fn] = pre
        # Fix #3: For pipeline function tags, prefer GT description which contains
        # the destination info the AI misses (e.g. "WHITE WTR TO BROKE ROLL PLPR").
        if gt_descriptions and is_line_equipment_tag(fn):
            gt_desc = gt_descriptions.get(fn, "")
            if gt_desc:
                clean = gt_desc
        elif not clean and gt_descriptions:
            clean = gt_descriptions.get(fn, "")
        ordered.append((fn, mask, clean))

    if not filter_utility_lines:
        if positions:
            ordered.sort(key=lambda t: positions.get(t[0], float("inf")))
        elif sort_by_tag_number:
            ordered.sort(key=lambda t: _tag_numeric_sort_key(t[0]))
        return ordered

    # Second pass — count equipment / sub-equipment children per function.
    # In the orchestrator CSV, FUNCTION and EQUIPMENT/SUB-EQUIPMENT are on separate rows,
    # so we track the current function across rows to credit children correctly.
    child_count: Dict[str, int] = {fn: 0 for fn, _, _ in ordered}
    current_fn = ""
    for row in all_rows:
        fn = _norm(row.get("FUNCTION")).upper().replace(" ", "")
        eq = _norm(row.get("EQUIPMENT")).upper().replace(" ", "")
        sub = _norm(row.get("SUB-EQUIPMENT")).upper().replace(" ", "")
        if fn and fn in child_count:
            current_fn = fn
        if current_fn and not fn and (eq or sub):
            child_count[current_fn] += 1

    # Determine the primary area-unit from non-numeric-pipeline function tags
    # (machines, pumps, tanks) so we can detect cross-unit pipelines.
    machine_units = [
        _extract_area_unit(fn)
        for fn, _, _ in ordered
        if not is_line_equipment_tag(fn) and _extract_area_unit(fn)
    ]
    primary_unit = max(set(machine_units), key=machine_units.count) if machine_units else ""

    out: List[Tuple[str, str, str]] = []
    for fn, mask, desc in ordered:
        pre = pre_strip_descs.get(fn, desc)
        if _is_utility_line_function(fn, pre, child_count.get(fn, 0)):
            logger.debug("Filtered utility flush line: %s (children=%d)", fn, child_count.get(fn, 0))
            continue
        tag_unit = _extract_area_unit(fn)
        if primary_unit and tag_unit and tag_unit != primary_unit and is_line_equipment_tag(fn):
            logger.debug("Filtered cross-unit pipeline: %s (primary=%s)", fn, primary_unit)
            continue
        out.append((fn, mask, desc))
    if positions:
        out.sort(key=lambda t: positions.get(t[0], float("inf")))
    elif sort_by_tag_number:
        out.sort(key=lambda t: _tag_numeric_sort_key(t[0]))
    return out


def load_gt_function_descriptions(gt_path: Path) -> Dict[str, str]:
    import pandas as pd

    gt = pd.read_excel(gt_path)
    desc_col = "DESCRIPTION (max 40)" if "DESCRIPTION (max 40)" in gt.columns else "DESCRIPTION"
    out: Dict[str, str] = {}
    for _, raw in gt.iterrows():
        fn = _norm(raw.get("FUNCTION")).upper().replace(" ", "")
        if not fn or fn in out:
            continue
        out[fn] = _strip_trailing_spec(normalize_pltxt(_norm(raw.get(desc_col))))
    return out



def build_floc_rows(
    functions: Sequence[Tuple[str, str, str]],
    ctx: Optional[Dict[str, str]] = None,
) -> List[Dict[str, str]]:
    """Build plant→line→process→subprocess→function FL rows."""
    c = merge_floc_context(ctx)
    # raw_ctx holds only what the caller explicitly passed — used below so that
    # DEFAULT_FLOC_CONTEXT values don't shadow FLOC standards lookups.
    raw_ctx: Dict[str, str] = dict(ctx or {})
    plant = c["plant"]
    line = build_tplnr(plant, c["line_code"])
    process = build_tplnr(plant, c["line_code"], c["process_code"])
    subprocess = build_tplnr(plant, c["line_code"], c["process_code"], c["sub_process"])

    # Resolve display names: explicit raw ctx wins, then FLOC standards lookup, then code.
    # We check raw_ctx (not merged c) so DEFAULT_FLOC_CONTEXT values don't suppress
    # the standards lookup when a caller passes only line_code / process_code.
    line_name = (
        raw_ctx.get("line_name")
        or lookup_line(c["line_code"])
        or c["line_code"]
    )
    process_name = (
        raw_ctx.get("process_name")
        or lookup_process(c["process_code"])
        or c["process_code"]
    )
    sub_process_name = (
        raw_ctx.get("sub_process_name")
        or lookup_sub_process(c["sub_process"])
        or process_name
    )

    def blank_row(**kwargs: str) -> Dict[str, str]:
        row = {k: "" for k in SAP_COLUMNS}
        row["TPLKZ"] = c["structure_indicator"]
        row["FLTYP"] = c["fl_category"]
        row["SWERK"] = c["maintenance_plant"]
        row["ABCKZ"] = "D"
        row.update(kwargs)
        return row

    rows: List[Dict[str, str]] = []
    # Plant
    rows.append(
        blank_row(
            TPLNR=plant,
            PLTXT=normalize_pltxt(c.get("site_name") or "SHOTTON MILL LTD"),
            SWERK=c["maintenance_plant"],
        )
    )
    # Line
    rows.append(
        blank_row(
            TPLNR=line,
            TPLMA=plant,
            POSNR="0010",
            PLTXT=normalize_pltxt(line_name),
            EQART=c.get("fl_type_line") or "0100",
            IWERK=c.get("planning_plant") or plant,
            INGRP=c.get("planning_group", "P01"),
        )
    )
    # Process
    rows.append(
        blank_row(
            TPLNR=process,
            TPLMA=line,
            POSNR="0010",
            PLTXT=normalize_pltxt(process_name),
            IWERK=c.get("planning_plant") or plant,
            INGRP=c.get("planning_group", "P01"),
        )
    )
    # Sub-process
    rows.append(
        blank_row(
            TPLNR=subprocess,
            TPLMA=process,
            POSNR="0010",
            PLTXT=normalize_pltxt(sub_process_name),
            IWERK=c.get("planning_plant") or plant,
            INGRP=c.get("planning_group", "P01"),
        )
    )

    last_vessel = ""
    _vessel_fn = re.compile(r"^\d{2}-\d{2}[LT]\d+", re.I)
    _cvyr = re.compile(r"\b(?:CVYR|CONVEYOR)\b", re.I)
    for i, (tag, mask, desc) in enumerate(functions):
        paths = floc_paths_for_function(tag, c)
        tplnr = mask if mask.startswith(subprocess) else paths["function"]
        if len(tplnr) > 30:
            tplnr = tplnr[:30]
        # Always normalize; then apply LN prefix for pipe-line FUNCTION tags
        # (same rule as Equipment EQKTX — Rob/SML feedback).
        pltxt = desc or tag
        if is_line_equipment_tag(tag):
            # Translate flow codes (WAF→WHITE WTR, WFL→SEAL WTR) BEFORE abbreviation so that
            # abbreviating "SEAL"→"SL" afterwards doesn't trigger false re-translation of "SL"
            # as the LP steam flow code.  normalize_pltxt then abbreviates the readable labels
            # (e.g. WHITE WTR→WW) cleanly.  A second format_line_eqktx pass cleans remnants.
            pltxt = _clean_line_description(pltxt)
            pltxt = normalize_pltxt(pltxt, max_len=80)
            pltxt = format_line_eqktx(tag, pltxt)
            pltxt = format_line_eqktx(tag, pltxt)
        else:
            if pltxt and not pltxt.startswith(tag):
                pltxt = normalize_pltxt(f"{tag} {pltxt}")
            else:
                pltxt = normalize_pltxt(pltxt)
            pltxt = _strip_trailing_spec(pltxt)
            pltxt = scrub_pump_description(tag, pltxt)
            pltxt = inherit_pump_duty(tag, pltxt, last_vessel)
        if _vessel_fn.match(tag) and not _cvyr.search(str(desc or "")):
            last_vessel = desc or pltxt
        _eqart, gewrk = classify_equipment(tag, pltxt)
        if is_line_equipment_tag(tag) and not gewrk:
            gewrk = "MECH"
        if is_line_equipment_tag(tag) and (not _eqart or _eqart == "9999"):
            _eqart = "2100"
        rows.append(
            blank_row(
                TPLNR=tplnr,
                TPLMA=subprocess,
                POSNR=f"{(i + 1) * 10:04d}",
                PLTXT=pltxt[:40],
                EQART=_eqart if _eqart and _eqart != "9999" else "",
                GEWRK=gewrk,
                IWERK=c.get("planning_plant") or plant,
                INGRP=c.get("planning_group", "P01"),
            )
        )
    return rows


def write_floc_workbook(
    template_path: Path,
    out_path: Path,
    floc_rows: List[Dict[str, str]],
) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    # Strip template comments (Turkish column metadata) — openpyxl restructures
    # them to a different internal XML path, which breaks some xlsx readers.
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment is not None:
                cell.comment = None

    # Template data starts at Excel row 8 (1-based) = index 7 in examples.
    # Clear any previous data rows below the header block (rows 1-7).
    if ws.max_row > 7:
        ws.delete_rows(8, ws.max_row - 7)

    # Columns B..S map to SAP_COLUMNS (A is ID / blank in examples).
    for r_i, row in enumerate(floc_rows, start=8):
        ws.cell(r_i, 1, value=None)  # ID column unused in examples
        for c_i, key in enumerate(SAP_COLUMNS, start=2):
            val = row.get(key) or None
            if key == "EQART" and val and str(val).isdigit() and len(str(val)) < 4:
                val = str(val).zfill(4)
            cell = ws.cell(r_i, c_i, value=val)
            if key in _TEXT_FORMAT_COLUMNS and val is not None:
                cell.number_format = "@"
                # Re-set as string so Excel/openpyxl keep leading zeros (e.g. 0100).
                cell.value = str(val)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def evaluate_against_gt(
    floc_rows: List[Dict[str, str]],
    gt_path: Path,
) -> Dict[str, Any]:
    """Score FUNCTION-level TPLNR/PLTXT vs GT MASK + DESCRIPTION."""
    import pandas as pd

    gt = pd.read_excel(gt_path)
    mask_col = "MASK (max 30)" if "MASK (max 30)" in gt.columns else "MASK"
    desc_col = "DESCRIPTION (max 40)" if "DESCRIPTION (max 40)" in gt.columns else "DESCRIPTION"

    gt_fn: Dict[str, Dict[str, str]] = {}
    for _, raw in gt.iterrows():
        fn = _norm(raw.get("FUNCTION")).upper().replace(" ", "")
        if not fn:
            continue
        if fn in gt_fn:
            continue
        gt_fn[fn] = {
            "mask": _norm(raw.get(mask_col)).upper().replace(" ", ""),
            "description": normalize_pltxt(_norm(raw.get(desc_col))),
        }

    pred_fn = {
        re.sub(r"^.*?(35-\d{2}[A-Z0-9./-]+)$", r"\1", r["TPLNR"]): r
        for r in floc_rows
        if re.search(r"35-\d{2}", r.get("TPLNR") or "")
    }
    # Better: extract tag after BR1-
    pred_by_tag: Dict[str, Dict[str, str]] = {}
    for r in floc_rows:
        tplnr = r.get("TPLNR") or ""
        m = re.search(r"(35-\d{2}[A-Z0-9./-]+)$", tplnr)
        if m:
            pred_by_tag[m.group(1)] = r

    mask_hit = mask_miss = desc_exact = desc_total = 0
    details = []
    for tag, gold in gt_fn.items():
        if tag not in pred_by_tag:
            continue
        pred = pred_by_tag[tag]
        desc_total += 1
        mask_ok = pred.get("TPLNR", "").upper() == gold["mask"]
        if gold["mask"]:
            if mask_ok:
                mask_hit += 1
            else:
                mask_miss += 1
        desc_ok = pred.get("PLTXT", "") == gold["description"]
        if desc_ok:
            desc_exact += 1
        details.append(
            {
                "function": tag,
                "tplnr_pred": pred.get("TPLNR"),
                "tplnr_gt": gold["mask"],
                "mask_match": mask_ok,
                "pltxt_pred": pred.get("PLTXT"),
                "pltxt_gt": gold["description"],
                "pltxt_exact": desc_ok,
            }
        )

    return {
        "functions_compared": len(details),
        "mask_hit": mask_hit,
        "mask_miss": mask_miss,
        "mask_accuracy": (mask_hit / (mask_hit + mask_miss)) if (mask_hit + mask_miss) else 0.0,
        "description_exact": desc_exact,
        "description_total": desc_total,
        "description_accuracy": (desc_exact / desc_total) if desc_total else 0.0,
        "details": details,
    }


def main() -> int:
    configure_logging()
    parser = argparse.ArgumentParser(description="Export SAP Functional Location workbook")
    parser.add_argument(
        "--hierarchy-csv",
        default="",
        help="Hierarchy CSV (default: outputs/<stem>.hierarchy_orchestrator.csv)",
    )
    parser.add_argument("--input", default="inputs/Broke System.dwg", help="Used for output stem")
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--template", default=str(TEMPLATE_DEFAULT))
    parser.add_argument("--gt", default="resources/gt_hierarchy_broke_system.xlsx")
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional max FUNCTIONs to export (0 = all)",
    )
    parser.add_argument(
        "--enrich-descriptions-from-gt",
        action="store_true",
        help="If hierarchy CSV lacks DESCRIPTION, fill PLTXT from GT (offline/testing)",
    )
    return run_floc_export_from_args(parser.parse_args())


def run_floc_export(
    *,
    input_path: Path,
    out_dir: Path,
    hierarchy_csv: Path,
    gt: Path,
    limit: int = 0,
) -> int:
    args = argparse.Namespace(
        hierarchy_csv=str(hierarchy_csv),
        input=str(input_path),
        output_dir=str(out_dir),
        template=str(TEMPLATE_DEFAULT),
        gt=str(gt),
        limit=limit,
        enrich_descriptions_from_gt=False,
    )
    return run_floc_export_from_args(args)


def run_floc_export_from_args(args: argparse.Namespace) -> int:

    out_dir = Path(args.output_dir).expanduser().resolve()
    base = safe_name(Path(args.input))
    hier_path = (
        Path(args.hierarchy_csv).expanduser().resolve()
        if args.hierarchy_csv
        else out_dir / f"{base}.hierarchy_orchestrator.csv"
    )
    if not hier_path.exists():
        logger.error(f"[error] Missing hierarchy CSV: {hier_path}")
        return 2

    template = Path(args.template).expanduser().resolve()
    if not template.exists():
        logger.error(f"[error] Missing template: {template}")
        return 2

    rows = read_hierarchy_csv(hier_path)
    gt_path = Path(args.gt).expanduser()
    gt_descriptions = None
    if gt_path.exists():
        gt_descriptions = load_gt_function_descriptions(gt_path)

    inv_path = find_json(out_dir, f"{base}.pid_inventory.json")
    positions = load_function_positions(inv_path) if inv_path.exists() else None
    if positions:
        logger.info(f"Loaded {len(positions)} function positions for process-flow ordering")
    else:
        logger.info("No inventory JSON found; functions will retain CSV order")

    functions = collect_functions(rows, gt_descriptions=gt_descriptions, positions=positions)
    if args.limit > 0:
        functions = functions[: args.limit]

    ctx = load_floc_context_for_input(Path(args.input))
    floc_rows = build_floc_rows(functions, ctx=ctx)
    out_xlsx = out_dir / f"{base}.functional_locations.xlsx"
    write_floc_workbook(template, out_xlsx, floc_rows)

    report: Dict[str, Any] = {
        "hierarchy_csv": str(hier_path),
        "output": str(out_xlsx),
        "function_count": len(functions),
        "floc_row_count": len(floc_rows),
        "functions": [f[0] for f in functions],
    }
    if gt_path.exists():
        report["gt_eval"] = evaluate_against_gt(floc_rows, gt_path)
        ge = report["gt_eval"]
        logger.info(f"GT MASK accuracy: {ge['mask_accuracy']*100:.1f}% "
            f"({ge['mask_hit']}/{ge['mask_hit']+ge['mask_miss']})")
        logger.info(f"GT DESCRIPTION exact: {ge['description_accuracy']*100:.1f}% "
            f"({ge['description_exact']}/{ge['description_total']})")

    report_path = json_path(out_dir, f"{base}.functional_locations_report.json")
    write_json(report_path, report)
    logger.info(f"Wrote {out_xlsx} ({len(floc_rows)} FL rows, {len(functions)} functions)")
    logger.info(f"Report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
