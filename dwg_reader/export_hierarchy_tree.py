#!/usr/bin/env python3
"""Colour-coded hierarchy tree (FLOC → equipment → sub-equipment).

Rob's working sheet uses yellow for FLOC rows, grey for equipment, and light
grey for sub-equipment. This writes both ``.hierarchy.xlsx`` and
``.hierarchy.html`` next to the SAP export workbooks.
"""

from __future__ import annotations

import argparse
import html
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from dwg_reader.dwg_floc_context import (
    floc_paths_for_function,
    load_floc_context_for_input,
    merge_floc_context,
)
from dwg_reader.dwg_pure_dump import json_path, safe_name, write_json
from dwg_reader.export_sap_equipment import resolve_hierarchy_csv
from dwg_reader.io import cell as _norm, read_csv_rows
from dwg_reader.logutil import configure_logging, get_logger

logger = get_logger(__name__)

TREE_COLUMNS = [
    "KIND",
    "SITE",
    "LINE",
    "PROCESS",
    "SUB-PROCESS",
    "FUNCTION",
    "EQUIPMENT",
    "SUB-EQUIPMENT",
    "MASK",
    "DESCRIPTION",
]

# Meeting 1: yellow = FLOC, grey = equipment, light grey = sub-equipment.
FILL_FLOC = "FFE599"
FILL_EQUIPMENT = "BFBFBF"
FILL_SUB = "D9D9D9"
FILL_HEADER = "1F4E79"
_KIND_FILL = {
    "FLOC": FILL_FLOC,
    "EQUIPMENT": FILL_EQUIPMENT,
    "SUB-EQUIPMENT": FILL_SUB,
}


def _kind_of(row: Dict[str, str]) -> str:
    if _norm(row.get("SUB-EQUIPMENT")):
        return "SUB-EQUIPMENT"
    if _norm(row.get("EQUIPMENT")):
        return "EQUIPMENT"
    return "FLOC"


def build_tree_rows(
    hierarchy_rows: Sequence[Dict[str, str]],
    *,
    ctx: Optional[Dict[str, str]] = None,
) -> List[Dict[str, str]]:
    """Expand hierarchy CSV into a review tree including plant/line/process/sub-process."""
    c = merge_floc_context(ctx)
    plant = c.get("plant") or ""
    line = c.get("line_code") or ""
    process = c.get("process_code") or ""
    sub = c.get("sub_process") or ""
    site_name = c.get("site_name") or ""
    line_name = c.get("line_name") or ""
    process_name = c.get("process_name") or ""
    sub_name = c.get("sub_process_name") or process_name

    paths0 = floc_paths_for_function("PLACEHOLDER", c)

    def tree_row(
        kind: str = "FLOC",
        *,
        line_code: str = "",
        process_code: str = "",
        sub_process: str = "",
        function: str = "",
        equipment: str = "",
        sub_equipment: str = "",
        mask: str = "",
        description: str = "",
    ) -> Dict[str, str]:
        return {
            "KIND": kind,
            "SITE": plant,
            "LINE": line_code,
            "PROCESS": process_code,
            "SUB-PROCESS": sub_process,
            "FUNCTION": function,
            "EQUIPMENT": equipment,
            "SUB-EQUIPMENT": sub_equipment,
            "MASK": mask,
            "DESCRIPTION": description,
        }

    out: List[Dict[str, str]] = [
        tree_row(mask=paths0["plant"], description=site_name),
        tree_row(line_code=line, mask=paths0["line"], description=line_name),
        tree_row(
            line_code=line,
            process_code=process,
            mask=paths0["process"],
            description=process_name,
        ),
        tree_row(
            line_code=line,
            process_code=process,
            sub_process=sub,
            mask=paths0["subprocess"],
            description=sub_name or process_name,
        ),
    ]

    current_fn = ""
    for row in hierarchy_rows:
        fn = _norm(row.get("FUNCTION")).upper().replace(" ", "")
        eq = _norm(row.get("EQUIPMENT")).upper().replace(" ", "")
        subeq = _norm(row.get("SUB-EQUIPMENT")).upper().replace(" ", "")
        desc = _norm(row.get("DESCRIPTION"))
        mask = _norm(row.get("MASK"))
        if fn and not eq and not subeq:
            current_fn = fn
            if not mask:
                mask = floc_paths_for_function(fn, c)["function"]
            out.append(
                tree_row(
                    line_code=line,
                    process_code=process,
                    sub_process=sub,
                    function=fn,
                    mask=mask,
                    description=desc,
                )
            )
            continue
        if not current_fn and not fn:
            continue
        use_fn = fn or current_fn
        if not mask:
            mask = floc_paths_for_function(use_fn, c)["function"]
        out.append(
            tree_row(
                kind=_kind_of({"FUNCTION": fn, "EQUIPMENT": eq, "SUB-EQUIPMENT": subeq}),
                line_code=line,
                process_code=process,
                sub_process=sub,
                function=use_fn,
                equipment=eq,
                sub_equipment=subeq,
                mask=mask,
                description=desc,
            )
        )
    return out


def _fill_for(kind: str) -> str:
    return _KIND_FILL.get(kind, FILL_FLOC)


def write_hierarchy_xlsx(path: Path, rows: List[Dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_fill = PatternFill("solid", fgColor=FILL_HEADER)
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    body_font = Font(name="Calibri", size=10)
    fills = {kind: PatternFill("solid", fgColor=colour) for kind, colour in _KIND_FILL.items()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Hierarchy"
    for c_i, key in enumerate(TREE_COLUMNS, start=1):
        cell = ws.cell(1, c_i, key)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for r_i, row in enumerate(rows, start=2):
        kind = row.get("KIND") or "FLOC"
        fill = fills.get(kind, fills["FLOC"])
        for c_i, key in enumerate(TREE_COLUMNS, start=1):
            cell = ws.cell(r_i, c_i, row.get(key) or "")
            cell.fill = fill
            cell.font = body_font
            cell.border = thin
    widths = {
        "A": 14, "B": 8, "C": 10, "D": 10, "E": 12,
        "F": 16, "G": 16, "H": 16, "I": 32, "J": 44,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A1:J{max(1, len(rows) + 1)}"
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_hierarchy_html(path: Path, rows: List[Dict[str, str]], *, title: str) -> None:
    def colour(kind: str) -> str:
        return f"#{_fill_for(kind)}"

    cells = []
    for row in rows:
        kind = row.get("KIND") or "FLOC"
        tds = "".join(
            f"<td>{html.escape(str(row.get(k) or ''))}</td>" for k in TREE_COLUMNS
        )
        cells.append(f'<tr style="background:{colour(kind)}">{tds}</tr>')
    headers = "".join(f"<th>{html.escape(k)}</th>" for k in TREE_COLUMNS)
    doc = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="utf-8"/>
<title>{html.escape(title)} hierarchy</title>
<style>
  body {{ font-family: Calibri, Segoe UI, sans-serif; margin: 24px; color: #222; }}
  h1 {{ font-size: 20px; }}
  .legend span {{ display: inline-block; padding: 2px 10px; margin-right: 12px; border: 1px solid #ccc; }}
  table {{ border-collapse: collapse; font-size: 13px; width: 100%; }}
  th {{ background: #{FILL_HEADER}; color: #fff; text-align: left; padding: 6px 8px; position: sticky; top: 0; }}
  td {{ padding: 4px 8px; border: 1px solid #ccc; white-space: nowrap; }}
</style>
</head><body>
<h1>{html.escape(title)}</h1>
<p class="legend">
  <span style="background:#{FILL_FLOC}">FLOC</span>
  <span style="background:#{FILL_EQUIPMENT}">Equipment</span>
  <span style="background:#{FILL_SUB}">Sub-equipment</span>
</p>
<table>
<thead><tr>{headers}</tr></thead>
<tbody>
{chr(10).join(cells)}
</tbody>
</table>
</body></html>
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(doc, encoding="utf-8")


def run_hierarchy_tree_export(
    *,
    input_path: Path,
    out_dir: Path,
    hierarchy_csv: Path,
) -> int:
    base = safe_name(input_path)
    ctx = load_floc_context_for_input(input_path)
    rows = build_tree_rows(read_csv_rows(hierarchy_csv, missing_ok=True), ctx=ctx)
    xlsx = out_dir / f"{base}.hierarchy.xlsx"
    html_path = out_dir / f"{base}.hierarchy.html"
    write_hierarchy_xlsx(xlsx, rows)
    write_hierarchy_html(html_path, rows, title=base)
    report = {
        "hierarchy_csv": str(hierarchy_csv),
        "xlsx": str(xlsx),
        "html": str(html_path),
        "row_count": len(rows),
        "floc": sum(1 for r in rows if r["KIND"] == "FLOC"),
        "equipment": sum(1 for r in rows if r["KIND"] == "EQUIPMENT"),
        "sub_equipment": sum(1 for r in rows if r["KIND"] == "SUB-EQUIPMENT"),
    }
    write_json(json_path(out_dir, f"{base}.hierarchy_tree_report.json"), report)
    logger.info(
        f"Wrote {xlsx.name} + {html_path.name} "
        f"({report['floc']} FLOC / {report['equipment']} EQ / {report['sub_equipment']} sub)"
    )
    return 0


def main() -> int:
    configure_logging()
    parser = argparse.ArgumentParser(description="Export colour-coded hierarchy tree")
    parser.add_argument("--hierarchy-csv", default="")
    parser.add_argument("--input", default="inputs/Broke System.dwg")
    parser.add_argument("--output-dir", default="outputs")
    args = parser.parse_args()
    out_dir = Path(args.output_dir).expanduser().resolve()
    input_path = Path(args.input).expanduser().resolve()
    hier = resolve_hierarchy_csv(out_dir, safe_name(input_path), args.hierarchy_csv)
    if not hier:
        logger.error(f"[error] Missing hierarchy CSV under {out_dir}")
        return 2
    return run_hierarchy_tree_export(
        input_path=input_path, out_dir=out_dir, hierarchy_csv=hier
    )


if __name__ == "__main__":
    raise SystemExit(main())
